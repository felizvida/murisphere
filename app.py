# Copyright 2026 Murisphere Contributors
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

from __future__ import annotations

import csv
from collections import deque
import fnmatch
import hashlib
import io
import json
import mimetypes
import os
import secrets
import threading
import time
from contextlib import closing
from dataclasses import dataclass
from datetime import UTC, date, datetime, timedelta
from functools import wraps
from pathlib import Path
from typing import Any
from urllib import error as urlerror
from urllib import request as urlrequest

from flask import Flask, Response, abort, g, jsonify, redirect, render_template, request, send_file, stream_with_context
from openpyxl import Workbook
from werkzeug.security import check_password_hash, generate_password_hash
import qrcode
from barcode import Code128
from barcode.writer import SVGWriter

import storage

APP_NAME = "Murisphere"
DB_PATH = os.getenv("MURISPHERE_DB", "murisphere.db")
ATTACHMENT_DIR = Path(os.getenv("MURISPHERE_ATTACHMENT_DIR", "uploads"))
APP_VERSION = Path("VERSION").read_text(encoding="utf-8").strip() if Path("VERSION").exists() else "dev"
DEFAULT_BIND_HOST = os.getenv("MURISPHERE_HOST", "0.0.0.0")
DEFAULT_BIND_PORT = int(os.getenv("MURISPHERE_PORT", "8000"))
RUNTIME_MODE = os.getenv("MURISPHERE_RUNTIME_MODE", "web")
TUTORIAL_DIR = Path("docs/tutorial")
TUTORIAL_HTML_PATH = TUTORIAL_DIR / "user_training_tutorial.html"
TUTORIAL_PDF_PATH = TUTORIAL_DIR / "user_training_tutorial.pdf"
TUTORIAL_CSS_PATH = TUTORIAL_DIR / "tutorial.css"
TUTORIAL_ASSET_DIR = TUTORIAL_DIR / "assets"

app = Flask(__name__, static_folder="static", template_folder="templates")
app.config["MAX_CONTENT_LENGTH"] = int(os.getenv("MURISPHERE_MAX_UPLOAD_BYTES", "5242880"))

LOGIN_RATE_LIMIT_MAX_FAILURES = int(os.getenv("MURISPHERE_LOGIN_RATE_LIMIT_MAX_FAILURES", "8"))
LOGIN_RATE_LIMIT_WINDOW_SEC = int(os.getenv("MURISPHERE_LOGIN_RATE_LIMIT_WINDOW_SEC", "300"))
LOGIN_RATE_LIMIT_BLOCK_SEC = int(os.getenv("MURISPHERE_LOGIN_RATE_LIMIT_BLOCK_SEC", "900"))
PUBLIC_SCAN_RATE_LIMIT_MAX = int(os.getenv("MURISPHERE_PUBLIC_SCAN_RATE_LIMIT_MAX", "120"))
PUBLIC_SCAN_RATE_LIMIT_WINDOW_SEC = int(os.getenv("MURISPHERE_PUBLIC_SCAN_RATE_LIMIT_WINDOW_SEC", "60"))

_rate_limit_lock = threading.Lock()
_login_failures: dict[str, deque[float]] = {}
_login_blocked_until: dict[str, float] = {}
_public_scan_hits: dict[str, deque[float]] = {}


@dataclass
class AuthContext:
    user_id: int
    email: str
    full_name: str
    role: str
    lab_id: int | None


PROJECT_CODE_LIST_SQL = storage.sql_list_agg("pj.project_code", ", ")
REQUEST_SLA_HOURS_SQL = storage.sql_hours_between("updated_at", "created_at")
GENOTYPING_PROVIDER_PRESETS: list[dict[str, Any]] = [
    {
        "key": "transnetyx",
        "name": "Transnetyx",
        "sampleProvider": "Transnetyx",
        "orderProvider": "Transnetyx",
        "defaultSampleType": "tail",
        "defaultMarkerPanel": "Cre Panel",
        "exportColumns": ["order_ref", "sample_code", "animal_code", "marker_panel", "result"],
        "importAliases": {
            "sampleCode": ["sample_code", "sampleid", "sample_id"],
            "animalCode": ["animal_code", "animal_id"],
            "result": ["result", "genotype_result", "call"],
            "markerPanel": ["marker_panel", "panel_name", "assay"],
        },
        "notes": "Phone-friendly default workflow with sample-code-based reconciliation.",
    },
    {
        "key": "charles-river",
        "name": "Charles River",
        "sampleProvider": "Charles River",
        "orderProvider": "Charles River",
        "defaultSampleType": "ear",
        "defaultMarkerPanel": "Mouse Line Verification",
        "exportColumns": ["order_ref", "tube_id", "animal_id", "panel_name", "result"],
        "importAliases": {
            "sampleCode": ["tube_id", "sample_code", "tube"],
            "animalCode": ["animal_id", "animal_code"],
            "result": ["result", "genotype_result", "call"],
            "markerPanel": ["panel_name", "marker_panel", "assay"],
        },
        "notes": "Tube-centric export/import aliases for external service workflows.",
    },
    {
        "key": "in-house-qpcr",
        "name": "In-House qPCR",
        "sampleProvider": "In-House qPCR",
        "orderProvider": "In-House qPCR",
        "defaultSampleType": "tissue",
        "defaultMarkerPanel": "qPCR Verification",
        "exportColumns": ["order_ref", "animal_code", "well_position", "target_assay", "marker_panel", "result"],
        "importAliases": {
            "sampleCode": ["sample_code", "well_position"],
            "animalCode": ["animal_code"],
            "result": ["result", "ct_call", "genotype_result"],
            "markerPanel": ["target_assay", "marker_panel"],
        },
        "notes": "Bench-oriented preset for internal assay plates and well-position tracking.",
    },
]

PROJECT_GENOTYPE_TEMPLATE_PRESETS: list[dict[str, Any]] = [
    {
        "presetKey": "balanced-pilot",
        "name": "Balanced Pilot Cohort",
        "description": "Quick pilot design for a small proof-of-concept cohort with matched driver and carrier animals.",
        "targetAnimals": 4,
        "targets": [
            {"genotypePattern": "Cre/+", "targetCount": 2, "priority": 1, "notes": "Driver-positive pilot animals"},
            {"genotypePattern": "fl/+", "targetCount": 2, "priority": 2, "notes": "Matched carrier controls"},
        ],
    },
    {
        "presetKey": "conditional-knockout",
        "name": "Conditional Knockout Study",
        "description": "Focus cohort on recombined conditional animals while keeping a smaller floxed-only control arm.",
        "targetAnimals": 6,
        "targets": [
            {"genotypePattern": "Cre/+;fl/fl", "targetCount": 4, "priority": 1, "notes": "Primary conditional knockout animals"},
            {"genotypePattern": "fl/fl", "targetCount": 2, "priority": 2, "notes": "Floxed controls without Cre"},
        ],
    },
    {
        "presetKey": "reporter-validation",
        "name": "Reporter Validation Set",
        "description": "Useful for imaging or validation projects that need reporter-positive animals alongside reporter-only comparators.",
        "targetAnimals": 6,
        "targets": [
            {"genotypePattern": "Cre/+;Ai14/+", "targetCount": 4, "priority": 1, "notes": "Reporter-positive experimental animals"},
            {"genotypePattern": "Ai14/+", "targetCount": 2, "priority": 2, "notes": "Reporter-only controls"},
        ],
    },
]

ASSIGNMENT_STATUS_STEPS: list[dict[str, str]] = [
    {"key": "reserved", "label": "Reserved", "color": "#4f8ef7"},
    {"key": "assigned", "label": "Assigned", "color": "#18a172"},
    {"key": "shipped", "label": "Shipped", "color": "#eb9c44"},
    {"key": "consumed", "label": "Consumed", "color": "#7c6cf2"},
    {"key": "released", "label": "Released", "color": "#64748b"},
]


def now_iso() -> str:
    return datetime.now(UTC).isoformat()


def _provider_preset_by_name(name: str | None) -> dict[str, Any]:
    normalized = str(name or "").strip().lower()
    for preset in GENOTYPING_PROVIDER_PRESETS:
        if normalized in {preset["key"], str(preset["name"]).strip().lower(), str(preset["orderProvider"]).strip().lower()}:
            return preset
    return GENOTYPING_PROVIDER_PRESETS[0]


def _genotype_template_preset_by_key(key: str | None) -> dict[str, Any] | None:
    normalized = str(key or "").strip().lower()
    for preset in PROJECT_GENOTYPE_TEMPLATE_PRESETS:
        if normalized == str(preset["presetKey"]).strip().lower():
            return preset
    return None


def _csv_pick(row: dict[str, Any], keys: list[str]) -> str:
    for key in keys:
        value = row.get(key)
        if value is None:
            continue
        text = str(value).strip()
        if text:
            return text
    return ""


def _match_genotype_pattern(genotype: str | None, pattern: str | None) -> bool:
    genotype_text = str(genotype or "").strip().lower()
    pattern_text = str(pattern or "").strip().lower()
    if not pattern_text:
        return False
    if "*" in pattern_text or "?" in pattern_text:
        return fnmatch.fnmatch(genotype_text, pattern_text)
    return genotype_text == pattern_text or pattern_text in genotype_text


def _normalize_target_rules(items: Any) -> list[dict[str, Any]]:
    if not isinstance(items, list):
        return []
    normalized: list[dict[str, Any]] = []
    seen_patterns: set[str] = set()
    for idx, item in enumerate(items, start=1):
        pattern = str((item or {}).get("genotypePattern") or (item or {}).get("genotype_pattern") or "").strip()
        folded = pattern.lower()
        if not pattern or folded in seen_patterns:
            continue
        seen_patterns.add(folded)
        raw_count = (item or {}).get("targetCount", (item or {}).get("target_count", 0))
        raw_priority = (item or {}).get("priority", idx)
        normalized.append(
            {
                "genotypePattern": pattern,
                "targetCount": max(0, int(raw_count or 0)),
                "priority": max(1, int(raw_priority or idx)),
                "notes": str((item or {}).get("notes") or "").strip(),
            }
        )
    return normalized


def _project_target_map(project_ids: list[int]) -> dict[int, list[dict[str, Any]]]:
    if not project_ids:
        return {}
    placeholders = ", ".join(["?"] * len(project_ids))
    rows = db().execute(
        f"""
        SELECT id, project_id, genotype_pattern, target_count, priority, notes
        FROM project_genotype_targets
        WHERE project_id IN ({placeholders})
        ORDER BY project_id ASC, priority ASC, id ASC
        """,
        project_ids,
    ).fetchall()
    targets: dict[int, list[dict[str, Any]]] = {}
    for row in rows:
        targets.setdefault(int(row["project_id"]), []).append(dict(row))
    return targets


def _replace_project_genotype_targets(project_id: int, targets: Any) -> tuple[list[dict[str, Any]], list[dict[str, Any]], int]:
    before = _project_target_map([project_id]).get(project_id, [])
    normalized = _normalize_target_rules(targets)
    db().execute("DELETE FROM project_genotype_targets WHERE project_id = ?", (project_id,))
    saved = 0
    for idx, item in enumerate(normalized, start=1):
        db().execute(
            """
            INSERT INTO project_genotype_targets (project_id, genotype_pattern, target_count, priority, notes, created_at)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (project_id, item["genotypePattern"], item["targetCount"], max(1, int(item["priority"] or idx)), item["notes"], now_iso()),
        )
        saved += 1
    after = _project_target_map([project_id]).get(project_id, [])
    return before, after, saved


def _custom_genotype_target_templates(user: AuthContext, template_ids: list[int] | None = None) -> list[dict[str, Any]]:
    filters: list[str] = []
    params: list[Any] = []
    if template_ids:
        filters.append("t.id IN (" + ", ".join(["?"] * len(template_ids)) + ")")
        params.extend(template_ids)
    if not is_admin(user):
        if user.lab_id is None:
            return []
        filters.append("t.lab_id = ?")
        params.append(user.lab_id)
    rows = db().execute(
        """
        SELECT t.id, t.lab_id, t.name, t.description, t.created_at,
               l.name AS lab_name, u.full_name AS created_by_name
        FROM genotype_target_templates t
        LEFT JOIN labs l ON l.id = t.lab_id
        LEFT JOIN users u ON u.id = t.created_by
        """
        + (" WHERE " + " AND ".join(filters) if filters else "")
        + """
        ORDER BY t.id DESC
        """,
        params,
    ).fetchall()
    if not rows:
        return []
    row_ids = [int(row["id"]) for row in rows]
    rules = db().execute(
        """
        SELECT id, template_id, genotype_pattern, target_count, priority, notes
        FROM genotype_target_template_rules
        WHERE template_id IN ("""
        + ", ".join(["?"] * len(row_ids))
        + """)
        ORDER BY template_id ASC, priority ASC, id ASC
        """,
        row_ids,
    ).fetchall()
    by_template: dict[int, list[dict[str, Any]]] = {}
    for rule in rules:
        by_template.setdefault(int(rule["template_id"]), []).append(
            {
                "id": int(rule["id"]),
                "genotypePattern": rule["genotype_pattern"],
                "targetCount": int(rule["target_count"] or 0),
                "priority": int(rule["priority"] or 1),
                "notes": rule["notes"],
            }
        )
    templates: list[dict[str, Any]] = []
    for row in rows:
        template_rules = by_template.get(int(row["id"]), [])
        templates.append(
            {
                "id": int(row["id"]),
                "source": "custom",
                "name": row["name"],
                "description": row["description"],
                "labId": row["lab_id"],
                "labName": row["lab_name"],
                "createdByName": row["created_by_name"],
                "targetAnimals": sum(int(rule["targetCount"]) for rule in template_rules),
                "targets": template_rules,
            }
        )
    return templates


def _visible_genotype_target_templates(user: AuthContext) -> list[dict[str, Any]]:
    templates = [
        {
            "source": "preset",
            "presetKey": preset["presetKey"],
            "name": preset["name"],
            "description": preset["description"],
            "labName": "Built-in",
            "targetAnimals": preset["targetAnimals"],
            "targets": _normalize_target_rules(preset["targets"]),
        }
        for preset in PROJECT_GENOTYPE_TEMPLATE_PRESETS
    ]
    templates.extend(_custom_genotype_target_templates(user))
    return templates


def _resolve_target_template(payload: dict[str, Any], user: AuthContext) -> dict[str, Any] | None:
    preset = _genotype_template_preset_by_key(payload.get("presetKey"))
    if preset:
        return {
            "source": "preset",
            "presetKey": preset["presetKey"],
            "name": preset["name"],
            "description": preset["description"],
            "targetAnimals": preset["targetAnimals"],
            "targets": _normalize_target_rules(preset["targets"]),
        }
    template_id = payload.get("templateId")
    if template_id is None:
        return None
    templates = _custom_genotype_target_templates(user, [int(template_id)])
    return templates[0] if templates else None


def _empty_assignment_status_counts() -> dict[str, int]:
    return {step["key"]: 0 for step in ASSIGNMENT_STATUS_STEPS}


def _project_assignment_status_counts(project_ids: list[int]) -> dict[int, dict[str, int]]:
    if not project_ids:
        return {}
    placeholders = ", ".join(["?"] * len(project_ids))
    rows = db().execute(
        f"""
        SELECT project_id, status, COUNT(*) AS count
        FROM project_animal_assignments
        WHERE project_id IN ({placeholders})
        GROUP BY project_id, status
        """,
        project_ids,
    ).fetchall()
    counts = {project_id: _empty_assignment_status_counts() for project_id in project_ids}
    for row in rows:
        project_counts = counts.setdefault(int(row["project_id"]), _empty_assignment_status_counts())
        project_counts[str(row["status"] or "reserved")] = int(row["count"] or 0)
    return counts


def _assignment_status_flow(project_id: int) -> list[dict[str, Any]]:
    counts = _project_assignment_status_counts([project_id]).get(project_id, _empty_assignment_status_counts())
    return [
        {
            "key": step["key"],
            "label": step["label"],
            "value": int(counts.get(step["key"], 0)),
            "color": step["color"],
        }
        for step in ASSIGNMENT_STATUS_STEPS
    ]


def _log_project_assignment_event(
    *,
    assignment_id: int | None,
    project_id: int,
    animal_id: int,
    event_type: str,
    from_status: str | None,
    to_status: str,
    notes: str | None,
    actor_user_id: int | None,
) -> None:
    db().execute(
        """
        INSERT INTO project_animal_assignment_events
            (assignment_id, project_id, animal_id, event_type, from_status, to_status, notes, actor_user_id, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (assignment_id, project_id, animal_id, event_type, from_status, to_status, notes, actor_user_id, now_iso()),
    )


def _project_assignment_timeline_payload(project_id: int) -> dict[str, Any]:
    flow = _assignment_status_flow(project_id)
    counts = {item["key"]: int(item["value"]) for item in flow}
    project = db().execute("SELECT project_code, title, target_animals FROM projects WHERE id = ?", (project_id,)).fetchone()
    completed = int(counts.get("consumed", 0)) + int(counts.get("released", 0))
    active = int(counts.get("reserved", 0)) + int(counts.get("assigned", 0)) + int(counts.get("shipped", 0))
    target = int(project["target_animals"] or 0) if project else 0
    completion_pct = round((completed / target) * 100, 1) if target > 0 else 0.0
    disposition = [
        {"label": "Consumed", "key": "consumed", "value": int(counts.get("consumed", 0)), "color": "#7c6cf2"},
        {"label": "Released", "key": "released", "value": int(counts.get("released", 0)), "color": "#64748b"},
    ]
    rows = db().execute(
        """
        SELECT e.id, e.event_type, e.from_status, e.to_status, e.notes, e.created_at,
               a.animal_code, c.cage_code, u.full_name AS actor_name
        FROM project_animal_assignment_events e
        JOIN animals a ON a.id = e.animal_id
        LEFT JOIN cages c ON c.id = a.cage_id
        LEFT JOIN users u ON u.id = e.actor_user_id
        WHERE e.project_id = ?
        ORDER BY e.created_at DESC, e.id DESC
        LIMIT 16
        """,
        (project_id,),
    ).fetchall()
    return {
        "statusFlow": flow,
        "statusCounts": counts,
        "activeAssignments": active,
        "completion": {
            "projectCode": project["project_code"] if project else None,
            "projectTitle": project["title"] if project else None,
            "targetAnimals": target,
            "completedAnimals": completed,
            "activeAnimals": active,
            "completionPct": completion_pct,
            "remainingAnimals": max(target - completed, 0),
            "state": "complete" if target and completed >= target else ("in_progress" if completed > 0 or active > 0 else "not_started"),
        },
        "dispositionFlow": disposition,
        "events": [
            {
                "id": int(row["id"]),
                "eventType": row["event_type"],
                "fromStatus": row["from_status"],
                "toStatus": row["to_status"],
                "notes": row["notes"],
                "createdAt": row["created_at"],
                "animalCode": row["animal_code"],
                "cageCode": row["cage_code"],
                "actorName": row["actor_name"],
            }
            for row in rows
        ],
    }


def db_target() -> str:
    return os.getenv("MURISPHERE_DATABASE_URL", DB_PATH)


def schema_path() -> str:
    return "schema_postgres.sql" if storage.is_postgres() else "schema.sql"


def token_digest(token: str) -> str:
    return hashlib.sha256(token.encode("utf-8")).hexdigest()


def tutorial_file_path(path: Path) -> Path:
    resolved = path.resolve()
    if not resolved.is_file():
        abort(404)
    return resolved


def tutorial_asset_path(name: str) -> Path:
    root = TUTORIAL_ASSET_DIR.resolve()
    resolved = (root / name).resolve()
    if root not in resolved.parents or not resolved.is_file():
        abort(404)
    return resolved


def tutorial_response(path: Path, mimetype: str) -> Response:
    return Response(path.read_bytes(), mimetype=mimetype)


def scoped_lab_clause(user: AuthContext, column: str) -> tuple[str, tuple[Any, ...]]:
    if is_admin(user):
        return ("", ())
    if user.lab_id is None:
        return (" WHERE 1 = 0", ())
    return (f" WHERE {column} = ?", (user.lab_id,))


def _client_ip() -> str:
    forwarded = request.headers.get("X-Forwarded-For", "")
    if forwarded:
        return forwarded.split(",")[0].strip() or "unknown"
    return request.remote_addr or "unknown"


def _prune_hits(bucket: deque[float], now_ts: float, window_sec: int) -> None:
    cutoff = now_ts - max(1, window_sec)
    while bucket and bucket[0] <= cutoff:
        bucket.popleft()


def _is_login_blocked(key: str) -> tuple[bool, int]:
    now_ts = time.monotonic()
    with _rate_limit_lock:
        blocked_until = _login_blocked_until.get(key, 0)
        if blocked_until <= now_ts:
            if key in _login_blocked_until:
                _login_blocked_until.pop(key, None)
            return (False, 0)
        retry = max(1, int(blocked_until - now_ts) + 1)
        return (True, retry)


def _record_login_failure(key: str) -> tuple[bool, int]:
    now_ts = time.monotonic()
    with _rate_limit_lock:
        bucket = _login_failures.setdefault(key, deque())
        _prune_hits(bucket, now_ts, LOGIN_RATE_LIMIT_WINDOW_SEC)
        bucket.append(now_ts)
        if len(bucket) >= max(1, LOGIN_RATE_LIMIT_MAX_FAILURES):
            blocked_until = now_ts + max(1, LOGIN_RATE_LIMIT_BLOCK_SEC)
            _login_blocked_until[key] = blocked_until
            bucket.clear()
            retry = max(1, int(blocked_until - now_ts) + 1)
            return (True, retry)
        return (False, 0)


def _clear_login_failures(key: str) -> None:
    with _rate_limit_lock:
        _login_failures.pop(key, None)
        _login_blocked_until.pop(key, None)


def _consume_public_scan_hit(client_ip: str) -> tuple[bool, int]:
    now_ts = time.monotonic()
    with _rate_limit_lock:
        bucket = _public_scan_hits.setdefault(client_ip, deque())
        _prune_hits(bucket, now_ts, PUBLIC_SCAN_RATE_LIMIT_WINDOW_SEC)
        if len(bucket) >= max(1, PUBLIC_SCAN_RATE_LIMIT_MAX):
            retry_after = max(1, int(PUBLIC_SCAN_RATE_LIMIT_WINDOW_SEC - (now_ts - bucket[0])) + 1)
            return (False, retry_after)
        bucket.append(now_ts)
        return (True, 0)


def reset_rate_limit_state() -> None:
    """Test hook to clear in-memory rate-limit trackers."""
    with _rate_limit_lock:
        _login_failures.clear()
        _login_blocked_until.clear()
        _public_scan_hits.clear()


def db() -> storage.Connection:
    if "db" not in g:
        conn = storage.connect(db_target())
        g.db = conn
    return g.db


@app.teardown_appcontext
def close_db(_exc: BaseException | None) -> None:
    conn = g.pop("db", None)
    if conn is not None:
        conn.close()


def init_db() -> None:
    with closing(storage.connect(db_target())) as conn:
        with open(schema_path(), "r", encoding="utf-8") as f:
            conn.executescript(f.read())
        _apply_schema_migrations(conn)
        conn.commit()

    conn = storage.connect(db_target())
    existing = conn.execute("SELECT id FROM users LIMIT 1").fetchone()
    if existing:
        conn.close()
        return

    conn.execute(
        "INSERT INTO facilities (name, timezone, created_at) VALUES (?, ?, ?)",
        ("North Campus Vivarium", "America/New_York", now_iso()),
    )
    conn.execute(
        "INSERT INTO labs (name, pi_name, facility_id, created_at) VALUES (?, ?, 1, ?)",
        ("Neurogenetics Lab", "Dr. A. Rivera", now_iso()),
    )
    conn.execute(
        "INSERT INTO rooms (name, facility_id, capacity, created_at) VALUES (?, 1, 240, ?)",
        ("Room A1", now_iso()),
    )
    conn.execute(
        "INSERT INTO racks (name, room_id, capacity, created_at) VALUES (?, 1, 120, ?)",
        ("Rack R1", now_iso()),
    )
    conn.execute(
        "INSERT INTO iacuc_protocols (protocol_number, title, lab_id, expires_on, created_at) VALUES (?, ?, 1, ?, ?)",
        ("IACUC-2026-014", "Synaptic Development Cohort", "2026-12-31", now_iso()),
    )

    users = [
        ("admin@murisphere.local", "Admin User", "Admin", None, "admin1234"),
        ("tech@murisphere.local", "Tech One", "Technician", 1, "tech1234"),
        ("pi@murisphere.local", "Principal Investigator", "PI", 1, "pi1234"),
    ]
    for email, name, role, lab_id, pwd in users:
        conn.execute(
            "INSERT INTO users (email, full_name, role, lab_id, password_hash, is_active, created_at) VALUES (?, ?, ?, ?, ?, 1, ?)",
            (email, name, role, lab_id, generate_password_hash(pwd), now_iso()),
        )

    cages = [
        ("C-A1-001", "C57BL/6J", "WT/WT", "Breeding", "2026-01-12", 1, 2, 1, 1, 1),
        ("C-A1-002", "Ai14 x Cre", "+/tg", "Holding", "2026-02-01", 2, 3, 1, 1, 1),
    ]
    for cage_id, strain, genotype, status, dob, males, females, room_id, rack_id, lab_id in cages:
        conn.execute(
            """
            INSERT INTO cages (
                cage_code, strain, genotype_summary, breeding_status, dob,
                male_count, female_count, room_id, rack_id, lab_id, protocol_id,
                qr_token, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1, ?, ?, ?)
            """,
            (
                cage_id,
                strain,
                genotype,
                status,
                dob,
                males,
                females,
                room_id,
                rack_id,
                lab_id,
                secrets.token_urlsafe(12),
                now_iso(),
                now_iso(),
            ),
        )

    conn.commit()
    conn.close()


def _apply_schema_migrations(conn: storage.Connection) -> None:
    litter_cols = set(storage.table_columns(conn, "litters"))
    if "weaned_on" not in litter_cols:
        conn.execute("ALTER TABLE litters ADD COLUMN weaned_on TEXT")


def audit_log(actor_id: int | None, entity_type: str, entity_id: int | str, action: str, before: Any, after: Any) -> None:
    db().execute(
        """
        INSERT INTO audit_logs (actor_user_id, entity_type, entity_id, action, before_json, after_json, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        (
            actor_id,
            entity_type,
            str(entity_id),
            action,
            json.dumps(before, default=str) if before is not None else None,
            json.dumps(after, default=str) if after is not None else None,
            now_iso(),
        ),
    )
    db().commit()


def current_user() -> AuthContext | None:
    token = request.headers.get("Authorization", "").replace("Bearer ", "").strip()
    if not token:
        token = request.cookies.get("murisphere_session", "").strip()
    if not token:
        return None
    row = db().execute(
        """
        SELECT u.id, u.email, u.full_name, u.role, u.lab_id
        FROM sessions s
        JOIN users u ON s.user_id = u.id
        WHERE s.token = ? AND s.expires_at > ?
        """,
        (token_digest(token), now_iso()),
    ).fetchone()
    if not row:
        return None
    return AuthContext(
        user_id=row["id"],
        email=row["email"],
        full_name=row["full_name"],
        role=row["role"],
        lab_id=row["lab_id"],
    )


def require_auth(roles: tuple[str, ...] | None = None):
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            user = current_user()
            if not user:
                return jsonify({"error": "Unauthorized"}), 401
            if roles and user.role not in roles:
                return jsonify({"error": "Forbidden"}), 403
            g.user = user
            return func(*args, **kwargs)

        return wrapper

    return decorator


def is_admin(user: AuthContext) -> bool:
    return user.role == "Admin"


def ensure_cage_scope(cage_id: int, user: AuthContext) -> storage.Row | None:
    if is_admin(user):
        return db().execute("SELECT * FROM cages WHERE id = ?", (cage_id,)).fetchone()
    if user.lab_id is None:
        return None
    return db().execute("SELECT * FROM cages WHERE id = ? AND lab_id = ?", (cage_id, user.lab_id)).fetchone()


def ensure_project_scope(project_id: int, user: AuthContext) -> storage.Row | None:
    if is_admin(user):
        return db().execute("SELECT * FROM projects WHERE id = ?", (project_id,)).fetchone()
    if user.lab_id is None:
        return None
    return db().execute("SELECT * FROM projects WHERE id = ? AND lab_id = ?", (project_id, user.lab_id)).fetchone()


def ensure_order_scope(order_id: int, user: AuthContext) -> storage.Row | None:
    if is_admin(user):
        return db().execute("SELECT * FROM animal_orders WHERE id = ?", (order_id,)).fetchone()
    if user.lab_id is None:
        return None
    return db().execute("SELECT * FROM animal_orders WHERE id = ? AND lab_id = ?", (order_id, user.lab_id)).fetchone()


def ensure_request_scope(request_id: int, user: AuthContext) -> storage.Row | None:
    if is_admin(user):
        return db().execute("SELECT * FROM facility_requests WHERE id = ?", (request_id,)).fetchone()
    if user.lab_id is None:
        return None
    return db().execute("SELECT * FROM facility_requests WHERE id = ? AND lab_id = ?", (request_id, user.lab_id)).fetchone()


def ensure_vet_case_scope(case_id: int, user: AuthContext) -> storage.Row | None:
    if is_admin(user):
        return db().execute("SELECT * FROM vet_cases WHERE id = ?", (case_id,)).fetchone()
    if user.lab_id is None:
        return None
    return db().execute("SELECT * FROM vet_cases WHERE id = ? AND lab_id = ?", (case_id, user.lab_id)).fetchone()


def ensure_animal_scope(animal_id: int, user: AuthContext) -> storage.Row | None:
    if is_admin(user):
        return db().execute("SELECT * FROM animals WHERE id = ?", (animal_id,)).fetchone()
    if user.lab_id is None:
        return None
    return db().execute(
        """
        SELECT a.*
        FROM animals a
        JOIN cages c ON c.id = a.cage_id
        WHERE a.id = ? AND c.lab_id = ?
        """,
        (animal_id, user.lab_id),
    ).fetchone()


def ensure_recommendation_scope(recommendation_id: int, user: AuthContext) -> storage.Row | None:
    if is_admin(user):
        return db().execute("SELECT * FROM workflow_recommendations WHERE id = ?", (recommendation_id,)).fetchone()
    if user.lab_id is None:
        return None
    return db().execute(
        "SELECT * FROM workflow_recommendations WHERE id = ? AND lab_id = ?",
        (recommendation_id, user.lab_id),
    ).fetchone()


def ensure_planner_scenario_scope(scenario_id: int, user: AuthContext) -> storage.Row | None:
    if is_admin(user):
        return db().execute("SELECT * FROM planner_scenarios WHERE id = ?", (scenario_id,)).fetchone()
    if user.lab_id is None:
        return None
    return db().execute("SELECT * FROM planner_scenarios WHERE id = ? AND lab_id = ?", (scenario_id, user.lab_id)).fetchone()


def cage_protocol_expired(cage_id: int) -> tuple[bool, str | None]:
    row = db().execute(
        """
        SELECT p.protocol_number, p.expires_on
        FROM cages c
        LEFT JOIN iacuc_protocols p ON p.id = c.protocol_id
        WHERE c.id = ?
        """,
        (cage_id,),
    ).fetchone()
    if not row or not row["expires_on"]:
        return False, None
    expired = row["expires_on"] < datetime.now(UTC).date().isoformat()
    if not expired:
        return False, None
    return True, f"Protocol {row['protocol_number']} expired on {row['expires_on']}"


def require_nonexpired_protocol(cage_id: int) -> Response | None:
    expired, msg = cage_protocol_expired(cage_id)
    if expired:
        return jsonify({"error": msg, "code": "PROTOCOL_EXPIRED"}), 409
    return None


SEVERITY_RANK = {"low": 1, "medium": 2, "high": 3}


def severity_at_least(actual: str, threshold: str) -> bool:
    return SEVERITY_RANK.get(actual, 0) >= SEVERITY_RANK.get(threshold, 0)


def escalation_delay_minutes(severity: str, escalation_level: int) -> int:
    base = {"high": 15, "medium": 60, "low": 240}.get(severity, 60)
    return max(5, base * max(1, escalation_level))


def derive_active_alerts(user: AuthContext) -> list[dict[str, Any]]:
    today = date.today().isoformat()
    now = now_iso()
    alerts: list[dict[str, Any]] = []

    scope_clause = "" if is_admin(user) else " AND c.lab_id = ?"
    scope_params: tuple[Any, ...] = () if is_admin(user) else (user.lab_id,)

    expired_rows = db().execute(
        """
        SELECT c.id AS cage_id, c.lab_id, c.cage_code, p.protocol_number, p.expires_on
        FROM cages c
        JOIN iacuc_protocols p ON p.id = c.protocol_id
        WHERE p.expires_on < ?
        """
        + scope_clause,
        (today, *scope_params),
    ).fetchall()
    for r in expired_rows:
        alerts.append(
            {
                "alert_key": f"protocol_expired:{r['cage_id']}",
                "lab_id": r["lab_id"],
                "cage_id": r["cage_id"],
                "severity": "high",
                "category": "protocol",
                "title": "Protocol Expired",
                "message": f"Cage {r['cage_code']} is on expired protocol {r['protocol_number']} ({r['expires_on']}).",
                "meta": {"protocolNumber": r["protocol_number"], "expiresOn": r["expires_on"]},
                "seen_at": now,
            }
        )

    task_rows = db().execute(
        """
        SELECT t.id, t.cage_id, c.lab_id, c.cage_code, t.task_type, t.due_on
        FROM task_assignments t
        JOIN cages c ON c.id = t.cage_id
        WHERE t.status IN ('pending', 'in_progress') AND t.due_on < ?
        """
        + ("" if is_admin(user) else " AND c.lab_id = ? "),
        (today,) if is_admin(user) else (today, user.lab_id),
    ).fetchall()
    for r in task_rows:
        alerts.append(
            {
                "alert_key": f"task_overdue:{r['id']}",
                "lab_id": r["lab_id"],
                "cage_id": r["cage_id"],
                "severity": "medium",
                "category": "task",
                "title": "Task Overdue",
                "message": f"Cage {r['cage_code']} has overdue task {r['task_type']} (due {r['due_on']}).",
                "meta": {"taskId": r["id"], "taskType": r["task_type"], "dueOn": r["due_on"]},
                "seen_at": now,
            }
        )

    dev_rows = db().execute(
        """
        SELECT d.id, d.cage_id, c.lab_id, c.cage_code, d.severity, d.summary
        FROM protocol_deviations d
        LEFT JOIN cages c ON c.id = d.cage_id
        JOIN iacuc_protocols p ON p.id = d.protocol_id
        WHERE d.status IN ('open', 'under_review')
        """
        + ("" if is_admin(user) else " AND p.lab_id = ? "),
        () if is_admin(user) else (user.lab_id,),
    ).fetchall()
    for r in dev_rows:
        sev = "high" if (r["severity"] or "").lower() in {"major", "critical", "high"} else "medium"
        alerts.append(
            {
                "alert_key": f"deviation_open:{r['id']}",
                "lab_id": r["lab_id"] if r["lab_id"] is not None else user.lab_id,
                "cage_id": r["cage_id"],
                "severity": sev,
                "category": "deviation",
                "title": "Protocol Deviation Open",
                "message": f"Deviation #{r['id']} is open: {r['summary'] or 'No summary'}",
                "meta": {"deviationId": r["id"]},
                "seen_at": now,
            }
        )

    necropsy_rows = db().execute(
        """
        SELECT m.id, m.cage_id, c.lab_id, c.cage_code
        FROM mortality_records m
        JOIN cages c ON c.id = m.cage_id
        WHERE m.necropsy_status = 'pending'
        """
        + ("" if is_admin(user) else " AND c.lab_id = ? "),
        () if is_admin(user) else (user.lab_id,),
    ).fetchall()
    for r in necropsy_rows:
        alerts.append(
            {
                "alert_key": f"necropsy_pending:{r['id']}",
                "lab_id": r["lab_id"],
                "cage_id": r["cage_id"],
                "severity": "high",
                "category": "mortality",
                "title": "Necropsy Pending",
                "message": f"Cage {r['cage_code']} has mortality record #{r['id']} pending necropsy.",
                "meta": {"mortalityId": r["id"]},
                "seen_at": now,
            }
        )

    vet_rows = db().execute(
        """
        SELECT v.id, v.cage_id, c.lab_id, c.cage_code, v.severity
        FROM vet_cases v
        JOIN cages c ON c.id = v.cage_id
        WHERE v.case_status = 'open'
        """
        + ("" if is_admin(user) else " AND c.lab_id = ? "),
        () if is_admin(user) else (user.lab_id,),
    ).fetchall()
    for r in vet_rows:
        sev = "high" if (r["severity"] or "").lower() in {"high", "critical"} else "medium"
        alerts.append(
            {
                "alert_key": f"vet_open:{r['id']}",
                "lab_id": r["lab_id"],
                "cage_id": r["cage_id"],
                "severity": sev,
                "category": "veterinary",
                "title": "Open Vet Case",
                "message": f"Cage {r['cage_code']} has open vet case #{r['id']}.",
                "meta": {"caseId": r["id"], "severity": r["severity"]},
                "seen_at": now,
            }
        )

    return alerts


def upsert_active_alerts(user: AuthContext) -> None:
    run_at = now_iso()
    active = derive_active_alerts(user)
    seen_keys = {a["alert_key"] for a in active}

    for a in active:
        existing = db().execute("SELECT id, status, escalation_level FROM alert_notifications WHERE alert_key = ?", (a["alert_key"],)).fetchone()
        if existing:
            db().execute(
                """
                UPDATE alert_notifications
                SET lab_id = ?, cage_id = ?, severity = ?, category = ?, title = ?, message = ?,
                    status = CASE WHEN status = 'resolved' THEN 'active' ELSE status END,
                    last_seen_at = ?, meta_json = ?
                WHERE id = ?
                """,
                (
                    a["lab_id"],
                    a["cage_id"],
                    a["severity"],
                    a["category"],
                    a["title"],
                    a["message"],
                    run_at,
                    json.dumps(a["meta"], default=str),
                    existing["id"],
                ),
            )
        else:
            db().execute(
                """
                INSERT INTO alert_notifications (
                    alert_key, lab_id, cage_id, severity, category, title, message, status,
                    first_seen_at, last_seen_at, next_notify_at, meta_json
                ) VALUES (?, ?, ?, ?, ?, ?, ?, 'active', ?, ?, ?, ?)
                """,
                (
                    a["alert_key"],
                    a["lab_id"],
                    a["cage_id"],
                    a["severity"],
                    a["category"],
                    a["title"],
                    a["message"],
                    run_at,
                    run_at,
                    run_at,
                    json.dumps(a["meta"], default=str),
                ),
            )

    if seen_keys:
        scope_sql = "" if is_admin(user) else " AND lab_id = ? "
        params: list[Any] = [run_at]
        if not is_admin(user):
            params.append(user.lab_id)
        placeholders = ", ".join(["?"] * len(seen_keys))
        params.extend(list(seen_keys))
        db().execute(
            f"""
            UPDATE alert_notifications
            SET status = 'resolved'
            WHERE status IN ('active', 'acknowledged')
              AND last_seen_at < ?
              {scope_sql}
              AND alert_key NOT IN ({placeholders})
            """,
            params,
        )
    else:
        db().execute(
            "UPDATE alert_notifications SET status = 'resolved' WHERE status IN ('active', 'acknowledged')"
            + ("" if is_admin(user) else " AND lab_id = ? "),
            () if is_admin(user) else (user.lab_id,),
        )
    db().commit()

def cage_payload(row: storage.Row) -> dict[str, Any]:
    return {
        "id": row["id"],
        "cageCode": row["cage_code"],
        "strain": row["strain"],
        "genotypeSummary": row["genotype_summary"],
        "breedingStatus": row["breeding_status"],
        "dob": row["dob"],
        "maleCount": row["male_count"],
        "femaleCount": row["female_count"],
        "room": row["room_name"],
        "rack": row["rack_name"],
        "lab": row["lab_name"],
        "protocol": row["protocol_number"],
        "projectCodes": row["project_codes"] if "project_codes" in row.keys() and row["project_codes"] else "",
        "notes": row["notes"],
        "qrToken": row["qr_token"],
        "updatedAt": row["updated_at"],
    }


@app.errorhandler(413)
def too_large(_err: Exception) -> Response:
    return jsonify({"error": "Upload too large"}), 413


def simple_pdf(lines: list[str]) -> bytes:
    def esc(s: str) -> str:
        return s.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")

    stream_lines = ["BT", "/F1 10 Tf", "50 790 Td", "12 TL"]
    for line in lines:
        stream_lines.append(f"({esc(line)}) Tj")
        stream_lines.append("T*")
    stream_lines.append("ET")
    content = "\n".join(stream_lines).encode("latin-1", errors="replace")

    objs = []
    objs.append(b"1 0 obj << /Type /Catalog /Pages 2 0 R >> endobj\n")
    objs.append(b"2 0 obj << /Type /Pages /Kids [3 0 R] /Count 1 >> endobj\n")
    objs.append(b"3 0 obj << /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] /Resources << /Font << /F1 4 0 R >> >> /Contents 5 0 R >> endobj\n")
    objs.append(b"4 0 obj << /Type /Font /Subtype /Type1 /BaseFont /Helvetica >> endobj\n")
    objs.append(f"5 0 obj << /Length {len(content)} >> stream\n".encode("ascii") + content + b"\nendstream endobj\n")

    body = b"%PDF-1.4\n"
    offsets = [0]
    for obj in objs:
        offsets.append(len(body))
        body += obj
    xref_start = len(body)
    xref = [f"xref\n0 {len(offsets)}\n".encode("ascii"), b"0000000000 65535 f \n"]
    for off in offsets[1:]:
        xref.append(f"{off:010d} 00000 n \n".encode("ascii"))
    trailer = f"trailer << /Size {len(offsets)} /Root 1 0 R >>\nstartxref\n{xref_start}\n%%EOF".encode("ascii")
    return body + b"".join(xref) + trailer


@app.route("/")
def index() -> str:
    return render_template("index.html", app_name=APP_NAME)


@app.route("/learn")
def tutorial_redirect() -> Response:
    return redirect("/learn/", code=308)


@app.route("/learn/")
def tutorial_page() -> Response:
    return tutorial_response(tutorial_file_path(TUTORIAL_HTML_PATH), "text/html")


@app.route("/learn/tutorial.css")
def tutorial_stylesheet() -> Response:
    return tutorial_response(tutorial_file_path(TUTORIAL_CSS_PATH), "text/css")


@app.route("/learn/user_training_tutorial.pdf")
def tutorial_pdf() -> Response:
    return tutorial_response(tutorial_file_path(TUTORIAL_PDF_PATH), "application/pdf")


@app.route("/learn/assets/<path:name>")
def tutorial_asset(name: str) -> Response:
    asset = tutorial_asset_path(name)
    guessed, _ = mimetypes.guess_type(asset.name)
    return tutorial_response(asset, guessed or "application/octet-stream")


@app.route("/scan/<token>")
def scan_page(token: str) -> str:
    return render_template("scan.html", app_name=APP_NAME, scan_token=token)


@app.get("/api/system/health")
def system_health() -> Response:
    return jsonify(
        {
            "ok": True,
            "app": APP_NAME,
            "version": APP_VERSION,
            "runtimeMode": RUNTIME_MODE,
            "storage": "postgres" if storage.is_postgres() else "sqlite",
            "timestamp": now_iso(),
        }
    )


@app.post("/api/auth/login")
def login() -> Response:
    payload = request.get_json(force=True)
    email = payload.get("email", "").strip().lower()
    password = payload.get("password", "")
    login_key = f"{_client_ip()}|{email}"
    blocked, retry_after = _is_login_blocked(login_key)
    if blocked:
        return jsonify({"error": "Too many login attempts. Retry later.", "retryAfterSec": retry_after}), 429

    row = db().execute(
        "SELECT id, email, full_name, role, lab_id, password_hash, is_active FROM users WHERE email = ?",
        (email,),
    ).fetchone()
    if not row or not row["is_active"] or not check_password_hash(row["password_hash"], password):
        now_blocked, block_retry = _record_login_failure(login_key)
        if now_blocked:
            return jsonify({"error": "Too many login attempts. Retry later.", "retryAfterSec": block_retry}), 429
        return jsonify({"error": "Invalid credentials"}), 401

    _clear_login_failures(login_key)
    token = secrets.token_urlsafe(24)
    expires_at = (datetime.now(UTC) + timedelta(hours=12)).isoformat()
    db().execute(
        "INSERT INTO sessions (token, user_id, expires_at, created_at) VALUES (?, ?, ?, ?)",
        (token_digest(token), row["id"], expires_at, now_iso()),
    )
    db().commit()

    resp = jsonify(
        {
            "token": token,
            "user": {
                "id": row["id"],
                "email": row["email"],
                "fullName": row["full_name"],
                "role": row["role"],
                "labId": row["lab_id"],
            },
            "appName": APP_NAME,
        }
    )
    resp.set_cookie(
        "murisphere_session",
        token,
        httponly=True,
        secure=os.getenv("MURISPHERE_COOKIE_SECURE", "0") == "1",
        samesite="Lax",
        max_age=12 * 60 * 60,
    )
    return resp


@app.post("/api/auth/logout")
@require_auth()
def logout() -> Response:
    token = request.headers.get("Authorization", "").replace("Bearer ", "").strip()
    if not token:
        token = request.cookies.get("murisphere_session", "").strip()
    if token:
        db().execute("DELETE FROM sessions WHERE token = ?", (token_digest(token),))
    db().commit()
    resp = jsonify({"ok": True})
    resp.delete_cookie("murisphere_session")
    return resp


@app.get("/api/auth/me")
@require_auth()
def me() -> Response:
    user = g.user
    return jsonify({"id": user.user_id, "email": user.email, "fullName": user.full_name, "role": user.role, "labId": user.lab_id})


def learning_modules() -> list[dict[str, Any]]:
    return [
        {
            "id": "orientation",
            "order": 1,
            "title": "Orientation and safe setup",
            "summary": "Start on the landing dashboard, verify scan setup, and get comfortable with the phone-first flow.",
            "actionLabel": "Open Dashboard",
            "tab": "dashboard",
        },
        {
            "id": "cage_cards",
            "order": 2,
            "title": "Cage card literacy",
            "summary": "Read the card the way a technician does: identify ownership, strain, genotype, population, and litter context fast.",
            "actionLabel": "Open Cages",
            "tab": "cages",
        },
        {
            "id": "scan_edit",
            "order": 3,
            "title": "Scan-to-edit workflow",
            "summary": "Jump from printed QR to the browser, update counts, and watch the audit trail behave like a real room workflow.",
            "actionLabel": "Open Scan/Edit",
            "tab": "scan",
        },
        {
            "id": "breeding",
            "order": 4,
            "title": "Breeding, litters, and pedigree",
            "summary": "Connect breeder pairs, litters, and inheritance so cage work stays tied to the biology.",
            "actionLabel": "Open Breeding",
            "tab": "breeding",
        },
        {
            "id": "projects",
            "order": 5,
            "title": "Projects and research readiness",
            "summary": "See how cages roll up into projects, quotas, and sample-driven research planning.",
            "actionLabel": "Open Projects",
            "tab": "projects",
        },
        {
            "id": "compliance",
            "order": 6,
            "title": "Compliance and abnormal conditions",
            "summary": "Practice identifying welfare pressure, protocol risk, and hard-stop situations before they disrupt work.",
            "actionLabel": "Open Compliance",
            "tab": "compliance",
        },
        {
            "id": "planner",
            "order": 7,
            "title": "Planner and manager workflows",
            "summary": "Use analytics, capacity, and seeded scenarios to reason about demand, risk, and cage growth.",
            "actionLabel": "Open Analytics",
            "tab": "analytics",
        },
    ]


def learning_counts(user: AuthContext) -> dict[str, int]:
    cage_where, cage_params = scoped_lab_clause(user, "lab_id")
    project_where, project_params = scoped_lab_clause(user, "lab_id")
    scenario_where, scenario_params = scoped_lab_clause(user, "lab_id")

    cages = db().execute(f"SELECT COUNT(*) AS n FROM cages{cage_where}", cage_params).fetchone()["n"]
    projects = db().execute(f"SELECT COUNT(*) AS n FROM projects{project_where}", project_params).fetchone()["n"]
    scenarios = db().execute(f"SELECT COUNT(*) AS n FROM planner_scenarios{scenario_where}", scenario_params).fetchone()["n"]
    if is_admin(user):
        labs = db().execute("SELECT COUNT(*) AS n FROM labs").fetchone()["n"]
        animals = db().execute("SELECT COUNT(*) AS n FROM animals").fetchone()["n"]
        litters = db().execute("SELECT COUNT(*) AS n FROM litters").fetchone()["n"]
        samples = db().execute("SELECT COUNT(*) AS n FROM sample_records").fetchone()["n"]
    else:
        labs = 1 if user.lab_id else 0
        animals = db().execute(
            """
            SELECT COUNT(*) AS n
            FROM animals a
            JOIN cages c ON c.id = a.cage_id
            WHERE c.lab_id = ?
            """,
            (user.lab_id,),
        ).fetchone()["n"]
        litters = db().execute(
            """
            SELECT COUNT(*) AS n
            FROM litters l
            JOIN cages c ON c.id = l.cage_id
            WHERE c.lab_id = ?
            """,
            (user.lab_id,),
        ).fetchone()["n"]
        samples = db().execute(
            """
            SELECT COUNT(*) AS n
            FROM sample_records s
            JOIN cages c ON c.id = s.cage_id
            WHERE c.lab_id = ?
            """,
            (user.lab_id,),
        ).fetchone()["n"]

    return {
        "labs": int(labs),
        "cages": int(cages),
        "projects": int(projects),
        "animals": int(animals),
        "litters": int(litters),
        "samples": int(samples),
        "plannerScenarios": int(scenarios),
    }


def learning_examples(user: AuthContext) -> dict[str, Any]:
    cage_where, cage_params = scoped_lab_clause(user, "c.lab_id")
    project_where, project_params = scoped_lab_clause(user, "p.lab_id")
    scenario_where, scenario_params = scoped_lab_clause(user, "ps.lab_id")

    cage_row = db().execute(
        f"""
        SELECT c.id, c.cage_code, c.breeding_status, c.strain, COUNT(l.id) AS litter_count
        FROM cages c
        LEFT JOIN litters l ON l.cage_id = c.id
        {cage_where}
        GROUP BY c.id, c.cage_code, c.breeding_status, c.strain
        ORDER BY litter_count DESC, c.id ASC
        LIMIT 1
        """,
        cage_params,
    ).fetchone()

    project_row = db().execute(
        f"""
        SELECT p.id, p.project_code, p.title, p.status
        FROM projects p
        {project_where}
        ORDER BY p.id ASC
        LIMIT 1
        """,
        project_params,
    ).fetchone()

    pedigree_row = db().execute(
        f"""
        SELECT a.id, a.animal_code, c.cage_code
        FROM animals a
        JOIN cages c ON c.id = a.cage_id
        {cage_where}{' AND' if cage_where else ' WHERE'} a.sire_id IS NOT NULL AND a.dam_id IS NOT NULL
        ORDER BY a.id ASC
        LIMIT 1
        """,
        cage_params,
    ).fetchone()

    sample_row = db().execute(
        f"""
        SELECT s.id, s.sample_code, s.status, a.animal_code
        FROM sample_records s
        JOIN cages c ON c.id = s.cage_id
        LEFT JOIN animals a ON a.id = s.animal_id
        {cage_where}
        ORDER BY s.id ASC
        LIMIT 1
        """,
        cage_params,
    ).fetchone()

    scenario_row = db().execute(
        f"""
        SELECT ps.id, ps.name, ps.status, l.name AS lab_name
        FROM planner_scenarios ps
        JOIN labs l ON l.id = ps.lab_id
        {scenario_where}
        ORDER BY ps.id ASC
        LIMIT 1
        """,
        scenario_params,
    ).fetchone()

    return {
        "cage": dict(cage_row) if cage_row else None,
        "project": dict(project_row) if project_row else None,
        "pedigreeAnimal": dict(pedigree_row) if pedigree_row else None,
        "sample": dict(sample_row) if sample_row else None,
        "plannerScenario": dict(scenario_row) if scenario_row else None,
    }


@app.get("/api/learning/overview")
@require_auth()
def learning_overview() -> Response:
    counts = learning_counts(g.user)
    availability = {
        "breedingPedigree": counts["animals"] > 0 and counts["litters"] > 0,
        "sampleGenotyping": counts["samples"] > 0,
        "planner": counts["plannerScenarios"] > 0,
        "projects": counts["projects"] > 0,
    }
    return jsonify(
        {
            "tutorialUrl": "/learn/",
            "tutorialPdfUrl": "/learn/user_training_tutorial.pdf",
            "counts": counts,
            "workflowAvailability": availability,
            "tutorialReady": all(availability.values()),
            "modules": learning_modules(),
            "examples": learning_examples(g.user),
        }
    )


@app.get("/api/projects")
@require_auth()
def list_projects() -> Response:
    params: tuple[Any, ...] = ()
    query = """
        SELECT p.id, p.lab_id, p.project_code, p.title, p.status, p.target_animals, p.created_at,
               l.name AS lab_name,
               COUNT(pc.cage_id) AS assigned_cages
        FROM projects p
        JOIN labs l ON l.id = p.lab_id
        LEFT JOIN project_cages pc ON pc.project_id = p.id
    """
    if not is_admin(g.user):
        query += " WHERE p.lab_id = ? "
        params = (g.user.lab_id,)
    query += " GROUP BY p.id, l.name ORDER BY p.created_at DESC LIMIT 500"
    rows = db().execute(query, params).fetchall()
    return jsonify([dict(r) for r in rows])


@app.post("/api/projects")
@require_auth(("PI", "Admin"))
def create_project() -> Response:
    payload = request.get_json(force=True)
    project_code = str(payload.get("projectCode", "")).strip()
    title = str(payload.get("title", "")).strip()
    if not project_code or not title:
        return jsonify({"error": "projectCode and title are required"}), 400
    status = str(payload.get("status", "active")).strip() or "active"
    target_animals = int(payload.get("targetAnimals", 0))
    if target_animals < 0:
        return jsonify({"error": "targetAnimals cannot be negative"}), 400

    lab_id = int(payload.get("labId", g.user.lab_id or 1))
    if not is_admin(g.user) and g.user.lab_id != lab_id:
        return jsonify({"error": "Forbidden"}), 403

    try:
        cur = db().execute(
            "INSERT INTO projects (lab_id, project_code, title, status, target_animals, created_at) VALUES (?, ?, ?, ?, ?, ?)",
            (lab_id, project_code, title, status, target_animals, now_iso()),
        )
        db().commit()
    except storage.IntegrityError:
        return jsonify({"error": "projectCode must be unique"}), 409

    project_id = cur.lastrowid
    audit_log(g.user.user_id, "project", project_id, "create", None, payload)
    return jsonify({"id": project_id}), 201


@app.patch("/api/projects/<int:project_id>")
@require_auth(("PI", "Admin"))
def update_project(project_id: int) -> Response:
    payload = request.get_json(force=True)
    row = ensure_project_scope(project_id, g.user)
    if not row:
        return jsonify({"error": "Not found"}), 404

    allowed = {
        "title": "title",
        "status": "status",
        "targetAnimals": "target_animals",
    }
    updates = {}
    for api_field, db_field in allowed.items():
        if api_field in payload:
            updates[db_field] = payload[api_field]
    if "target_animals" in updates and int(updates["target_animals"]) < 0:
        return jsonify({"error": "targetAnimals cannot be negative"}), 400
    if not updates:
        return jsonify({"error": "No changes supplied"}), 400

    before = dict(row)
    set_sql = ", ".join([f"{k} = ?" for k in updates])
    params = list(updates.values()) + [project_id]
    db().execute(f"UPDATE projects SET {set_sql} WHERE id = ?", params)
    db().commit()
    after = dict(db().execute("SELECT * FROM projects WHERE id = ?", (project_id,)).fetchone())
    audit_log(g.user.user_id, "project", project_id, "update", before, after)
    return jsonify({"ok": True})


@app.get("/api/projects/<int:project_id>/cages")
@require_auth()
def project_cages(project_id: int) -> Response:
    project = ensure_project_scope(project_id, g.user)
    if not project:
        return jsonify({"error": "Not found"}), 404
    rows = db().execute(
        """
        SELECT c.id, c.cage_code, c.strain, c.genotype_summary, c.breeding_status, c.male_count, c.female_count
        FROM project_cages pc
        JOIN cages c ON c.id = pc.cage_id
        WHERE pc.project_id = ?
        ORDER BY c.cage_code
        """,
        (project_id,),
    ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.post("/api/projects/<int:project_id>/assign-cages")
@require_auth(("PI", "Admin"))
def assign_project_cages(project_id: int) -> Response:
    project = ensure_project_scope(project_id, g.user)
    if not project:
        return jsonify({"error": "Not found"}), 404
    payload = request.get_json(force=True)
    cage_ids = payload.get("cageIds", [])
    if not cage_ids:
        return jsonify({"error": "Provide cageIds"}), 400

    assigned = 0
    for raw_id in cage_ids:
        cage_id = int(raw_id)
        cage = ensure_cage_scope(cage_id, g.user)
        if not cage:
            continue
        if int(cage["lab_id"]) != int(project["lab_id"]):
            continue
        try:
            db().execute(
                "INSERT INTO project_cages (project_id, cage_id, assigned_at) VALUES (?, ?, ?)",
                (project_id, cage_id, now_iso()),
            )
            assigned += 1
        except storage.IntegrityError:
            continue
    db().commit()
    audit_log(g.user.user_id, "project", project_id, "assign_cages", None, {"cageIds": cage_ids, "assigned": assigned})
    return jsonify({"assigned": assigned})


@app.get("/api/projects/<int:project_id>/genotype-targets")
@require_auth()
def project_genotype_targets(project_id: int) -> Response:
    project = ensure_project_scope(project_id, g.user)
    if not project:
        return jsonify({"error": "Not found"}), 404
    return jsonify(_project_target_map([project_id]).get(project_id, []))


@app.post("/api/projects/<int:project_id>/genotype-targets")
@require_auth(("PI", "Admin"))
def set_project_genotype_targets(project_id: int) -> Response:
    project = ensure_project_scope(project_id, g.user)
    if not project:
        return jsonify({"error": "Not found"}), 404
    payload = request.get_json(force=True)
    targets = payload.get("targets", [])
    if not isinstance(targets, list):
        return jsonify({"error": "targets must be a list"}), 400
    before, after, saved = _replace_project_genotype_targets(project_id, targets)
    db().commit()
    audit_log(g.user.user_id, "project", project_id, "set_genotype_targets", {"targets": before}, {"targets": after})
    return jsonify({"saved": saved})


@app.get("/api/genotyping/target-templates")
@require_auth()
def genotype_target_templates() -> Response:
    return jsonify(_visible_genotype_target_templates(g.user))


@app.post("/api/genotyping/target-templates")
@require_auth(("PI", "Admin"))
def create_genotype_target_template() -> Response:
    payload = request.get_json(force=True)
    name = str(payload.get("name") or "").strip()
    targets = payload.get("targets", [])
    if not name:
        return jsonify({"error": "Provide template name"}), 400
    normalized = _normalize_target_rules(targets)
    if not normalized:
        return jsonify({"error": "Provide at least one genotype target rule"}), 400
    lab_id = int(payload.get("labId") or g.user.lab_id or 0)
    if not lab_id:
        return jsonify({"error": "Provide labId"}), 400
    if not is_admin(g.user) and g.user.lab_id != lab_id:
        return jsonify({"error": "Forbidden"}), 403
    cur = db().execute(
        """
        INSERT INTO genotype_target_templates (lab_id, name, description, created_by, created_at)
        VALUES (?, ?, ?, ?, ?)
        """,
        (lab_id, name, payload.get("description"), g.user.user_id, now_iso()),
    )
    template_id = int(cur.lastrowid)
    for idx, item in enumerate(normalized, start=1):
        db().execute(
            """
            INSERT INTO genotype_target_template_rules (template_id, genotype_pattern, target_count, priority, notes)
            VALUES (?, ?, ?, ?, ?)
            """,
            (template_id, item["genotypePattern"], item["targetCount"], max(1, int(item["priority"] or idx)), item["notes"]),
        )
    db().commit()
    audit_log(
        g.user.user_id,
        "genotype_target_template",
        template_id,
        "create",
        None,
        {"name": name, "labId": lab_id, "targets": normalized},
    )
    return jsonify({"id": template_id}), 201


@app.post("/api/projects/<int:project_id>/apply-target-template")
@require_auth(("PI", "Admin"))
def apply_project_target_template(project_id: int) -> Response:
    project = ensure_project_scope(project_id, g.user)
    if not project:
        return jsonify({"error": "Not found"}), 404
    payload = request.get_json(force=True)
    template = _resolve_target_template(payload, g.user)
    if not template:
        return jsonify({"error": "Template not found"}), 404
    before, after, saved = _replace_project_genotype_targets(project_id, template["targets"])
    if template.get("targetAnimals") is not None:
        db().execute("UPDATE projects SET target_animals = ? WHERE id = ?", (int(template["targetAnimals"] or 0), project_id))
    db().commit()
    audit_log(
        g.user.user_id,
        "project",
        project_id,
        "apply_target_template",
        {"targets": before},
        {
            "targets": after,
            "template": {
                "source": template.get("source"),
                "presetKey": template.get("presetKey"),
                "id": template.get("id"),
                "name": template.get("name"),
            },
        },
    )
    return jsonify({"saved": saved, "templateName": template.get("name")})


@app.get("/api/projects/<int:project_id>/assignments")
@require_auth()
def project_animal_assignments(project_id: int) -> Response:
    project = ensure_project_scope(project_id, g.user)
    if not project:
        return jsonify({"error": "Not found"}), 404
    rows = db().execute(
        """
        SELECT pa.id, pa.status, pa.notes, pa.assigned_at,
               a.id AS animal_id, a.animal_code, a.sex, a.genotype, a.status AS animal_status,
               c.id AS cage_id, c.cage_code,
               u.full_name AS assigned_by_name
        FROM project_animal_assignments pa
        JOIN animals a ON a.id = pa.animal_id
        LEFT JOIN cages c ON c.id = a.cage_id
        LEFT JOIN users u ON u.id = pa.assigned_by
        WHERE pa.project_id = ?
        ORDER BY pa.assigned_at DESC, pa.id DESC
        """,
        (project_id,),
    ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.get("/api/projects/<int:project_id>/assignment-timeline")
@require_auth()
def project_assignment_timeline(project_id: int) -> Response:
    project = ensure_project_scope(project_id, g.user)
    if not project:
        return jsonify({"error": "Not found"}), 404
    return jsonify(_project_assignment_timeline_payload(project_id))


@app.post("/api/projects/<int:project_id>/reserve-animals")
@require_auth(("PI", "Admin"))
def reserve_project_animals(project_id: int) -> Response:
    project = ensure_project_scope(project_id, g.user)
    if not project:
        return jsonify({"error": "Not found"}), 404
    payload = request.get_json(force=True)
    animal_ids = [int(x) for x in payload.get("animalIds", [])]
    if not animal_ids:
        return jsonify({"error": "Provide animalIds"}), 400
    force = bool(payload.get("force", False))
    notes = payload.get("notes")
    targets = _project_target_map([project_id]).get(project_id, [])
    reserved = 0
    conflicts = []
    for animal_id in animal_ids:
        animal = ensure_animal_scope(animal_id, g.user)
        if not animal:
            conflicts.append({"animalId": animal_id, "reason": "not_found"})
            continue
        if animal["status"] != "Active":
            conflicts.append({"animalId": animal_id, "reason": "not_active"})
            continue
        genotype = str(animal["genotype"] or "").strip()
        if not genotype:
            conflicts.append({"animalId": animal_id, "reason": "genotype_missing"})
            continue
        if targets and not any(_match_genotype_pattern(genotype, target["genotype_pattern"]) for target in targets) and not force:
            conflicts.append({"animalId": animal_id, "reason": "target_mismatch"})
            continue
        existing = db().execute(
            """
            SELECT pa.project_id, pa.status, p.project_code
            FROM project_animal_assignments pa
            JOIN projects p ON p.id = pa.project_id
            WHERE pa.animal_id = ? AND pa.status <> 'released'
            """,
            (animal_id,),
        ).fetchone()
        if existing and int(existing["project_id"]) != project_id and not force:
            conflicts.append({"animalId": animal_id, "reason": "already_reserved", "projectCode": existing["project_code"]})
            continue
        db().execute(
            """
            INSERT INTO project_animal_assignments (project_id, animal_id, status, notes, assigned_at, assigned_by)
            VALUES (?, ?, 'reserved', ?, ?, ?)
            ON CONFLICT(animal_id) DO UPDATE SET
                project_id = excluded.project_id,
                status = 'reserved',
                notes = excluded.notes,
                assigned_at = excluded.assigned_at,
                assigned_by = excluded.assigned_by
            """,
            (project_id, animal_id, notes, now_iso(), g.user.user_id),
        )
        assignment = db().execute(
            """
            SELECT id
            FROM project_animal_assignments
            WHERE project_id = ? AND animal_id = ?
            """,
            (project_id, animal_id),
        ).fetchone()
        _log_project_assignment_event(
            assignment_id=int(assignment["id"]) if assignment else None,
            project_id=project_id,
            animal_id=animal_id,
            event_type="reserve",
            from_status=str(existing["status"]) if existing else None,
            to_status="reserved",
            notes=notes,
            actor_user_id=g.user.user_id,
        )
        reserved += 1
    db().commit()
    audit_log(g.user.user_id, "project", project_id, "reserve_animals", None, {"animalIds": animal_ids, "reserved": reserved, "conflicts": conflicts})
    return jsonify({"reserved": reserved, "conflicts": conflicts})


@app.post("/api/projects/<int:project_id>/assignment-status")
@require_auth(("PI", "Admin"))
def update_project_assignment_status(project_id: int) -> Response:
    project = ensure_project_scope(project_id, g.user)
    if not project:
        return jsonify({"error": "Not found"}), 404
    payload = request.get_json(force=True)
    animal_ids = [int(x) for x in payload.get("animalIds", [])]
    if not animal_ids:
        return jsonify({"error": "Provide animalIds"}), 400
    status = str(payload.get("status") or "").strip().lower()
    allowed = {step["key"] for step in ASSIGNMENT_STATUS_STEPS}
    if status not in allowed:
        return jsonify({"error": "Invalid status"}), 400
    notes = payload.get("notes")
    updated = 0
    conflicts = []
    for animal_id in animal_ids:
        row = db().execute(
            """
            SELECT id, status
            FROM project_animal_assignments
            WHERE project_id = ? AND animal_id = ?
            """,
            (project_id, animal_id),
        ).fetchone()
        if not row:
            conflicts.append({"animalId": animal_id, "reason": "assignment_missing"})
            continue
        before_status = str(row["status"] or "")
        if before_status == status:
            conflicts.append({"animalId": animal_id, "reason": "unchanged"})
            continue
        db().execute(
            """
            UPDATE project_animal_assignments
            SET status = ?, notes = COALESCE(?, notes), assigned_by = ?
            WHERE id = ?
            """,
            (status, notes, g.user.user_id, int(row["id"])),
        )
        _log_project_assignment_event(
            assignment_id=int(row["id"]),
            project_id=project_id,
            animal_id=animal_id,
            event_type=f"status_{status}",
            from_status=before_status,
            to_status=status,
            notes=notes,
            actor_user_id=g.user.user_id,
        )
        updated += 1
    db().commit()
    audit_log(
        g.user.user_id,
        "project",
        project_id,
        "assignment_status",
        None,
        {"animalIds": animal_ids, "status": status, "updated": updated, "conflicts": conflicts},
    )
    return jsonify({"updated": updated, "conflicts": conflicts})


@app.post("/api/projects/<int:project_id>/release-animals")
@require_auth(("PI", "Admin"))
def release_project_animals(project_id: int) -> Response:
    project = ensure_project_scope(project_id, g.user)
    if not project:
        return jsonify({"error": "Not found"}), 404
    payload = request.get_json(force=True)
    animal_ids = [int(x) for x in payload.get("animalIds", [])]
    if not animal_ids:
        return jsonify({"error": "Provide animalIds"}), 400
    released = 0
    for animal_id in animal_ids:
        row = db().execute(
            "SELECT id, status FROM project_animal_assignments WHERE project_id = ? AND animal_id = ?",
            (project_id, animal_id),
        ).fetchone()
        if not row or str(row["status"] or "") == "released":
            continue
        db().execute(
            "UPDATE project_animal_assignments SET status = 'released', notes = COALESCE(?, notes), assigned_by = ? WHERE id = ?",
            (payload.get("notes"), g.user.user_id, int(row["id"])),
        )
        _log_project_assignment_event(
            assignment_id=int(row["id"]),
            project_id=project_id,
            animal_id=animal_id,
            event_type="release",
            from_status=str(row["status"] or ""),
            to_status="released",
            notes=payload.get("notes"),
            actor_user_id=g.user.user_id,
        )
        released += 1
    db().commit()
    audit_log(g.user.user_id, "project", project_id, "release_animals", None, {"animalIds": animal_ids, "released": released})
    return jsonify({"released": released})


@app.get("/api/cages")
@require_auth()
def list_cages() -> Response:
    q = request.args.get("q", "").strip()
    status = request.args.get("status", "").strip()
    clauses = []
    params: list[Any] = []
    if q:
        clauses.append(
            """
            (
              c.cage_code LIKE ? OR
              c.strain LIKE ? OR
              c.genotype_summary LIKE ? OR
              l.name LIKE ? OR
              l.pi_name LIKE ? OR
              COALESCE(p.protocol_number, '') LIKE ? OR
              EXISTS (
                SELECT 1
                FROM project_cages pc_q
                JOIN projects p_q ON p_q.id = pc_q.project_id
                WHERE pc_q.cage_id = c.id
                  AND (p_q.project_code LIKE ? OR p_q.title LIKE ?)
              )
            )
            """
        )
        params.extend([f"%{q}%", f"%{q}%", f"%{q}%", f"%{q}%", f"%{q}%", f"%{q}%", f"%{q}%", f"%{q}%"])
    if status:
        clauses.append("c.breeding_status = ?")
        params.append(status)
    if not is_admin(g.user):
        clauses.append("c.lab_id = ?")
        params.append(g.user.lab_id)

    where_sql = f"WHERE {' AND '.join(clauses)}" if clauses else ""
    rows = db().execute(
        f"""
        SELECT
            c.*,
            r.name AS room_name,
            k.name AS rack_name,
            l.name AS lab_name,
            p.protocol_number,
            (
              SELECT {PROJECT_CODE_LIST_SQL}
              FROM project_cages pc_j
              JOIN projects pj ON pj.id = pc_j.project_id
              WHERE pc_j.cage_id = c.id
            ) AS project_codes
        FROM cages c
        LEFT JOIN rooms r ON c.room_id = r.id
        LEFT JOIN racks k ON c.rack_id = k.id
        LEFT JOIN labs l ON c.lab_id = l.id
        LEFT JOIN iacuc_protocols p ON c.protocol_id = p.id
        {where_sql}
        ORDER BY c.updated_at DESC
        LIMIT 500
        """,
        params,
    ).fetchall()
    return jsonify([cage_payload(row) for row in rows])


@app.get("/api/cages/<int:cage_id>")
@require_auth()
def get_cage(cage_id: int) -> Response:
    scope_clause = "" if is_admin(g.user) else "AND c.lab_id = ?"
    params: list[Any] = [cage_id]
    if not is_admin(g.user):
        params.append(g.user.lab_id)
    row = db().execute(
        f"""
        SELECT
            c.*,
            r.name AS room_name,
            k.name AS rack_name,
            l.name AS lab_name,
            p.protocol_number,
            (
              SELECT {PROJECT_CODE_LIST_SQL}
              FROM project_cages pc_j
              JOIN projects pj ON pj.id = pc_j.project_id
              WHERE pc_j.cage_id = c.id
            ) AS project_codes
        FROM cages c
        LEFT JOIN rooms r ON c.room_id = r.id
        LEFT JOIN racks k ON c.rack_id = k.id
        LEFT JOIN labs l ON c.lab_id = l.id
        LEFT JOIN iacuc_protocols p ON c.protocol_id = p.id
        WHERE c.id = ?
        """
        + scope_clause,
        params,
    ).fetchone()
    if not row:
        return jsonify({"error": "Not found"}), 404

    animals = db().execute(
        "SELECT id, animal_code, sex, dob, genotype, status FROM animals WHERE cage_id = ? ORDER BY id DESC",
        (cage_id,),
    ).fetchall()
    history = db().execute(
        "SELECT actor_user_id, action, created_at FROM audit_logs WHERE entity_type = 'cage' AND entity_id = ? ORDER BY id DESC LIMIT 30",
        (str(cage_id),),
    ).fetchall()

    return jsonify(
        {
            "cage": cage_payload(row),
            "animals": [dict(r) for r in animals],
            "history": [dict(r) for r in history],
        }
    )


@app.patch("/api/cages/<int:cage_id>")
@require_auth(("Technician", "PI", "Admin"))
def update_cage(cage_id: int) -> Response:
    payload = request.get_json(force=True)
    row = ensure_cage_scope(cage_id, g.user)
    if not row:
        return jsonify({"error": "Not found"}), 404
    blocked = require_nonexpired_protocol(cage_id)
    if blocked:
        return blocked

    allowed = {
        "maleCount": "male_count",
        "femaleCount": "female_count",
        "breedingStatus": "breeding_status",
        "notes": "notes",
        "genotypeSummary": "genotype_summary",
    }
    updates = {}
    for api_field, db_field in allowed.items():
        if api_field in payload:
            updates[db_field] = payload[api_field]

    if "male_count" in updates and int(updates["male_count"]) < 0:
        return jsonify({"error": "maleCount cannot be negative"}), 400
    if "female_count" in updates and int(updates["female_count"]) < 0:
        return jsonify({"error": "femaleCount cannot be negative"}), 400

    if not updates:
        return jsonify({"error": "No changes supplied"}), 400

    before = dict(row)
    updates["updated_at"] = now_iso()
    set_sql = ", ".join([f"{k} = ?" for k in updates])
    params = list(updates.values()) + [cage_id]
    db().execute(f"UPDATE cages SET {set_sql} WHERE id = ?", params)
    db().commit()

    after = dict(db().execute("SELECT * FROM cages WHERE id = ?", (cage_id,)).fetchone())
    audit_log(g.user.user_id, "cage", cage_id, "update", before, after)
    return jsonify({"ok": True})


@app.post("/api/cages")
@require_auth(("PI", "Admin"))
def create_cage() -> Response:
    payload = request.get_json(force=True)
    cage_code = payload["cageCode"].strip()
    strain = payload.get("strain", "Unknown")
    genotype = payload.get("genotypeSummary", "Unknown")
    breeding_status = payload.get("breedingStatus", "Holding")
    dob = payload.get("dob")
    male_count = int(payload.get("maleCount", 0))
    female_count = int(payload.get("femaleCount", 0))
    if male_count < 0 or female_count < 0:
        return jsonify({"error": "Counts cannot be negative"}), 400
    room_id = int(payload.get("roomId", 1))
    rack_id = int(payload.get("rackId", 1))
    lab_id = int(payload.get("labId", g.user.lab_id or 1))
    if not is_admin(g.user) and g.user.lab_id != lab_id:
        return jsonify({"error": "Forbidden"}), 403
    protocol_id = int(payload.get("protocolId", 1))

    cur = db().execute(
        """
        INSERT INTO cages (
            cage_code, strain, genotype_summary, breeding_status, dob, male_count, female_count,
            room_id, rack_id, lab_id, protocol_id, qr_token, created_at, updated_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            cage_code,
            strain,
            genotype,
            breeding_status,
            dob,
            male_count,
            female_count,
            room_id,
            rack_id,
            lab_id,
            protocol_id,
            secrets.token_urlsafe(12),
            now_iso(),
            now_iso(),
        ),
    )
    db().commit()
    cage_id = cur.lastrowid
    audit_log(g.user.user_id, "cage", cage_id, "create", None, payload)
    return jsonify({"id": cage_id}), 201


@app.get("/api/scan/<code>")
@require_auth()
def scan_cage(code: str) -> Response:
    started = datetime.now(UTC)
    scope_clause = "" if is_admin(g.user) else " AND c.lab_id = ?"
    params: list[Any] = [code, code]
    if not is_admin(g.user):
        params.append(g.user.lab_id)
    row = db().execute(
        """
        SELECT c.*, r.name AS room_name, k.name AS rack_name, l.name AS lab_name, p.protocol_number
        FROM cages c
        LEFT JOIN rooms r ON c.room_id = r.id
        LEFT JOIN racks k ON c.rack_id = k.id
        LEFT JOIN labs l ON c.lab_id = l.id
        LEFT JOIN iacuc_protocols p ON c.protocol_id = p.id
        WHERE (c.cage_code = ? OR c.qr_token = ?)
        """
        + scope_clause
        + """
        """,
        params,
    ).fetchone()
    if not row:
        return jsonify({"error": "Cage not found"}), 404
    elapsed_ms = int((datetime.now(UTC) - started).total_seconds() * 1000)
    return jsonify({"cage": cage_payload(row), "lookupMs": elapsed_ms})


@app.get("/api/public/scan/<token>")
def public_scan(token: str) -> Response:
    allowed, retry_after = _consume_public_scan_hit(_client_ip())
    if not allowed:
        return jsonify({"error": "Too many scan requests. Retry later.", "retryAfterSec": retry_after}), 429

    row = db().execute(
        """
        SELECT c.*, r.name AS room_name, k.name AS rack_name, l.name AS lab_name, p.protocol_number
        FROM cages c
        LEFT JOIN rooms r ON c.room_id = r.id
        LEFT JOIN racks k ON c.rack_id = k.id
        LEFT JOIN labs l ON c.lab_id = l.id
        LEFT JOIN iacuc_protocols p ON c.protocol_id = p.id
        WHERE c.qr_token = ?
        """,
        (token,),
    ).fetchone()
    if not row:
        return jsonify({"error": "Cage not found"}), 404
    return jsonify(
        {
            "cage": {
                "id": row["id"],
                "cageCode": row["cage_code"],
                "strain": row["strain"],
                "genotypeSummary": row["genotype_summary"],
                "breedingStatus": row["breeding_status"],
                "dob": row["dob"],
                "maleCount": row["male_count"],
                "femaleCount": row["female_count"],
                "protocol": row["protocol_number"],
                "room": row["room_name"],
                "rack": row["rack_name"],
                "lab": row["lab_name"],
                "updatedAt": row["updated_at"],
            }
        }
    )


@app.get("/api/assets/qrcode.png")
def qrcode_asset() -> Response:
    value = request.args.get("v", "").strip()
    if not value:
        return jsonify({"error": "Missing query parameter: v"}), 400
    if len(value) > 2048:
        return jsonify({"error": "Value too long"}), 400

    qr = qrcode.QRCode(
        version=None,
        error_correction=qrcode.constants.ERROR_CORRECT_M,
        box_size=8,
        border=2,
    )
    qr.add_data(value)
    qr.make(fit=True)
    image = qr.make_image(fill_color="#0f2027", back_color="#ffffff")
    bio = io.BytesIO()
    image.save(bio, format="PNG")
    bio.seek(0)
    return send_file(bio, mimetype="image/png")


@app.get("/api/assets/barcode.svg")
def barcode_asset() -> Response:
    value = request.args.get("v", "").strip()
    if not value:
        return jsonify({"error": "Missing query parameter: v"}), 400
    if len(value) > 80:
        return jsonify({"error": "Value too long"}), 400

    code = Code128(value, writer=SVGWriter())
    svg_payload = code.render(
        writer_options={
            "module_width": 0.28,
            "module_height": 14.0,
            "quiet_zone": 1.0,
            "font_size": 8,
            "text_distance": 1.0,
            "write_text": True,
        }
    )
    return Response(svg_payload, mimetype="image/svg+xml")


@app.post("/api/cages/<int:cage_id>/wean")
@require_auth(("Technician", "PI", "Admin"))
def wean(cage_id: int) -> Response:
    payload = request.get_json(force=True)
    male = int(payload.get("male", 0))
    female = int(payload.get("female", 0))
    event_date = payload.get("date", datetime.now(UTC).date().isoformat())
    litter_id = payload.get("litterId")
    if male < 0 or female < 0:
        return jsonify({"error": "Counts cannot be negative"}), 400
    if not ensure_cage_scope(cage_id, g.user):
        return jsonify({"error": "Not found"}), 404
    blocked = require_nonexpired_protocol(cage_id)
    if blocked:
        return blocked

    litter_row = None
    if litter_id is not None:
        try:
            litter_id = int(litter_id)
        except (TypeError, ValueError):
            return jsonify({"error": "litterId must be an integer"}), 400
        litter_row = db().execute("SELECT id FROM litters WHERE id = ? AND cage_id = ?", (litter_id, cage_id)).fetchone()
        if not litter_row:
            return jsonify({"error": "Litter not found for cage"}), 404

    cur = db().execute(
        "INSERT INTO lifecycle_events (cage_id, event_type, details_json, event_date, created_by, created_at) VALUES (?, 'weaning', ?, ?, ?, ?)",
        (cage_id, json.dumps({"male": male, "female": female, "litterId": litter_id}), event_date, g.user.user_id, now_iso()),
    )
    if litter_row:
        db().execute("UPDATE litters SET weaned_on = ? WHERE id = ?", (event_date, litter_id))
    db().execute(
        "UPDATE cages SET male_count = male_count + ?, female_count = female_count + ?, updated_at = ? WHERE id = ?",
        (male, female, now_iso(), cage_id),
    )
    db().commit()
    audit_log(
        g.user.user_id,
        "lifecycle_event",
        cur.lastrowid,
        "wean",
        None,
        {"cage_id": cage_id, "male": male, "female": female, "litterId": litter_id, "date": event_date},
    )
    return jsonify({"ok": True})


@app.post("/api/cages/<int:cage_id>/transfer")
@require_auth(("Technician", "PI", "Admin"))
def transfer(cage_id: int) -> Response:
    payload = request.get_json(force=True)
    target_room_id = int(payload["roomId"])
    target_rack_id = int(payload["rackId"])

    scoped = ensure_cage_scope(cage_id, g.user)
    if not scoped:
        return jsonify({"error": "Not found"}), 404
    blocked = require_nonexpired_protocol(cage_id)
    if blocked:
        return blocked
    before = db().execute("SELECT room_id, rack_id FROM cages WHERE id = ?", (cage_id,)).fetchone()
    if not before:
        return jsonify({"error": "Not found"}), 404
    db().execute(
        "UPDATE cages SET room_id = ?, rack_id = ?, updated_at = ? WHERE id = ?",
        (target_room_id, target_rack_id, now_iso(), cage_id),
    )
    db().execute(
        "INSERT INTO lifecycle_events (cage_id, event_type, details_json, event_date, created_by, created_at) VALUES (?, 'transfer', ?, ?, ?, ?)",
        (
            cage_id,
            json.dumps({"fromRoom": before["room_id"], "fromRack": before["rack_id"], "toRoom": target_room_id, "toRack": target_rack_id}),
            datetime.now(UTC).date().isoformat(),
            g.user.user_id,
            now_iso(),
        ),
    )
    db().commit()
    audit_log(g.user.user_id, "cage", cage_id, "transfer", dict(before), {"room_id": target_room_id, "rack_id": target_rack_id})
    return jsonify({"ok": True})


@app.post("/api/cages/<int:cage_id>/note")
@require_auth(("Technician", "PI", "Admin"))
def add_note(cage_id: int) -> Response:
    payload = request.get_json(force=True)
    text = payload.get("text", "").strip()
    if not ensure_cage_scope(cage_id, g.user):
        return jsonify({"error": "Not found"}), 404
    blocked = require_nonexpired_protocol(cage_id)
    if blocked:
        return blocked
    if not text:
        return jsonify({"error": "Note cannot be empty"}), 400
    db().execute(
        "INSERT INTO notes (entity_type, entity_id, text, created_by, created_at) VALUES ('cage', ?, ?, ?, ?)",
        (str(cage_id), text, g.user.user_id, now_iso()),
    )
    db().commit()
    audit_log(g.user.user_id, "cage", cage_id, "note", None, {"text": text})
    return jsonify({"ok": True})


@app.post("/api/cages/bulk-actions")
@require_auth(("PI", "Admin"))
def bulk_cage_actions() -> Response:
    payload = request.get_json(force=True)
    action = str(payload.get("action", "")).strip()
    cage_ids = payload.get("cageIds", [])
    if action not in {"retire_breeders", "transfer"}:
        return jsonify({"error": "Unsupported action"}), 400
    if not cage_ids:
        return jsonify({"error": "Provide cageIds"}), 400

    updated = 0
    if action == "retire_breeders":
        for raw_id in cage_ids:
            cage_id = int(raw_id)
            row = ensure_cage_scope(cage_id, g.user)
            if not row:
                continue
            blocked = require_nonexpired_protocol(cage_id)
            if blocked:
                continue
            before = dict(row)
            db().execute(
                "UPDATE cages SET breeding_status = 'Retired', updated_at = ? WHERE id = ?",
                (now_iso(), cage_id),
            )
            db().execute(
                "INSERT INTO lifecycle_events (cage_id, event_type, details_json, event_date, created_by, created_at) VALUES (?, 'retire', ?, ?, ?, ?)",
                (cage_id, json.dumps({"reason": payload.get("reason", "bulk_retire")}), datetime.now(UTC).date().isoformat(), g.user.user_id, now_iso()),
            )
            updated += 1
            audit_log(g.user.user_id, "cage", cage_id, "retire", before, {"breeding_status": "Retired"})
    else:
        room_id = int(payload.get("roomId", 0))
        rack_id = int(payload.get("rackId", 0))
        if room_id <= 0 or rack_id <= 0:
            return jsonify({"error": "roomId and rackId are required for transfer"}), 400
        for raw_id in cage_ids:
            cage_id = int(raw_id)
            row = ensure_cage_scope(cage_id, g.user)
            if not row:
                continue
            blocked = require_nonexpired_protocol(cage_id)
            if blocked:
                continue
            before = {"room_id": row["room_id"], "rack_id": row["rack_id"]}
            db().execute(
                "UPDATE cages SET room_id = ?, rack_id = ?, updated_at = ? WHERE id = ?",
                (room_id, rack_id, now_iso(), cage_id),
            )
            db().execute(
                "INSERT INTO lifecycle_events (cage_id, event_type, details_json, event_date, created_by, created_at) VALUES (?, 'transfer', ?, ?, ?, ?)",
                (
                    cage_id,
                    json.dumps({"fromRoom": row["room_id"], "fromRack": row["rack_id"], "toRoom": room_id, "toRack": rack_id}),
                    datetime.now(UTC).date().isoformat(),
                    g.user.user_id,
                    now_iso(),
                ),
            )
            updated += 1
            audit_log(g.user.user_id, "cage", cage_id, "transfer", before, {"room_id": room_id, "rack_id": rack_id})
    db().commit()
    return jsonify({"updated": updated})


@app.post("/api/litters")
@require_auth(("Technician", "PI", "Admin"))
def create_litter() -> Response:
    payload = request.get_json(force=True)
    cage_id = int(payload["cageId"])
    birth_date = payload["birthDate"]
    weaned_on = payload.get("weanedOn") or payload.get("dow")
    size = int(payload.get("size", 0))
    survived = int(payload.get("survived", size))
    if size < 0 or survived < 0:
        return jsonify({"error": "Litter counts cannot be negative"}), 400
    try:
        date.fromisoformat(str(birth_date))
    except ValueError:
        return jsonify({"error": "birthDate must be ISO date format YYYY-MM-DD"}), 400
    if weaned_on:
        try:
            date.fromisoformat(str(weaned_on))
        except ValueError:
            return jsonify({"error": "weanedOn/dow must be ISO date format YYYY-MM-DD"}), 400
    if not ensure_cage_scope(cage_id, g.user):
        return jsonify({"error": "Not found"}), 404
    blocked = require_nonexpired_protocol(cage_id)
    if blocked:
        return blocked

    cur = db().execute(
        "INSERT INTO litters (cage_id, birth_date, litter_size, survived_count, weaned_on, created_at) VALUES (?, ?, ?, ?, ?, ?)",
        (cage_id, birth_date, size, survived, weaned_on, now_iso()),
    )
    litter_id = cur.lastrowid
    count_m = int(payload.get("male", 0))
    count_f = int(payload.get("female", 0))

    for i in range(count_m):
        db().execute(
            "INSERT INTO animals (animal_code, sex, dob, strain, genotype, status, cage_id, litter_id, created_at, updated_at) VALUES (?, 'M', ?, ?, ?, 'Active', ?, ?, ?, ?)",
            (f"M-{litter_id}-{i+1:03d}", birth_date, payload.get("strain", "Unknown"), payload.get("genotype", "Pending"), cage_id, litter_id, now_iso(), now_iso()),
        )
    for i in range(count_f):
        db().execute(
            "INSERT INTO animals (animal_code, sex, dob, strain, genotype, status, cage_id, litter_id, created_at, updated_at) VALUES (?, 'F', ?, ?, ?, 'Active', ?, ?, ?, ?)",
            (f"F-{litter_id}-{i+1:03d}", birth_date, payload.get("strain", "Unknown"), payload.get("genotype", "Pending"), cage_id, litter_id, now_iso(), now_iso()),
        )

    db().execute(
        "UPDATE cages SET male_count = male_count + ?, female_count = female_count + ?, updated_at = ? WHERE id = ?",
        (count_m, count_f, now_iso(), cage_id),
    )
    db().commit()

    audit_log(g.user.user_id, "litter", litter_id, "create", None, payload)
    return jsonify({"id": litter_id})


@app.post("/api/breeding/events")
@require_auth(("Technician", "PI", "Admin"))
def breeding_event() -> Response:
    payload = request.get_json(force=True)
    cage_id = int(payload["cageId"])
    if not ensure_cage_scope(cage_id, g.user):
        return jsonify({"error": "Not found"}), 404
    blocked = require_nonexpired_protocol(cage_id)
    if blocked:
        return blocked
    cur = db().execute(
        "INSERT INTO breeding_events (cage_id, event_type, event_date, details_json, assigned_to, created_at) VALUES (?, ?, ?, ?, ?, ?)",
        (
            cage_id,
            payload["eventType"],
            payload["eventDate"],
            json.dumps(payload.get("details", {})),
            payload.get("assignedTo"),
            now_iso(),
        ),
    )
    db().commit()
    audit_log(g.user.user_id, "breeding_event", cur.lastrowid, "create", None, payload)
    return jsonify({"id": cur.lastrowid})


@app.get("/api/calendar")
@require_auth()
def calendar() -> Response:
    start = request.args.get("start", datetime.now(UTC).date().isoformat())
    end = request.args.get("end", (datetime.now(UTC).date() + timedelta(days=30)).isoformat())

    scope_clause = ""
    params: list[Any] = [start, end]
    if not is_admin(g.user):
        scope_clause = " AND c.lab_id = ?"
        params.append(g.user.lab_id)
    events = db().execute(
        """
        SELECT b.id, b.cage_id, c.cage_code, b.event_type, b.event_date, b.assigned_to, b.details_json
        FROM breeding_events b
        JOIN cages c ON b.cage_id = c.id
        WHERE b.event_date BETWEEN ? AND ?
        """
        + scope_clause
        + """
        ORDER BY b.event_date ASC
        """,
        params,
    ).fetchall()
    return jsonify([dict(e) for e in events])


@app.post("/api/genotyping/upload")
@require_auth(("PI", "Admin"))
def genotype_upload() -> Response:
    if "file" not in request.files:
        return jsonify({"error": "Upload a CSV file"}), 400
    f = request.files["file"]
    content = f.read().decode("utf-8")
    reader = csv.DictReader(io.StringIO(content))
    inserted = 0
    for row in reader:
        animal_code = row.get("animal_code")
        genotype = row.get("genotype_result")
        if not animal_code or not genotype:
            continue
        animal = db().execute("SELECT id, genotype FROM animals WHERE animal_code = ?", (animal_code,)).fetchone()
        if not animal:
            continue
        if not is_admin(g.user):
            scoped = db().execute(
                """
                SELECT a.id
                FROM animals a
                JOIN cages c ON c.id = a.cage_id
                WHERE a.id = ? AND c.lab_id = ?
                """,
                (animal["id"], g.user.lab_id),
            ).fetchone()
            if not scoped:
                continue
        db().execute(
            "INSERT INTO genotype_results (animal_id, result, source, created_at) VALUES (?, ?, ?, ?)",
            (animal["id"], genotype, "CSV", now_iso()),
        )
        db().execute("UPDATE animals SET genotype = ?, updated_at = ? WHERE id = ?", (genotype, now_iso(), animal["id"]))
        inserted += 1
    db().commit()
    return jsonify({"updatedAnimals": inserted})


@app.get("/api/animals")
@require_auth()
def list_animals() -> Response:
    q = request.args.get("q", "").strip()
    sex = request.args.get("sex", "").strip()
    status = request.args.get("status", "").strip()
    genotype = request.args.get("genotype", "").strip()

    clauses = ["1 = 1"]
    params: list[Any] = []
    if q:
        clauses.append("(a.animal_code LIKE ? OR a.strain LIKE ? OR COALESCE(a.genotype, '') LIKE ?)")
        params.extend([f"%{q}%", f"%{q}%", f"%{q}%"])
    if sex:
        clauses.append("a.sex = ?")
        params.append(sex)
    if status:
        clauses.append("a.status = ?")
        params.append(status)
    if genotype:
        clauses.append("COALESCE(a.genotype, '') LIKE ?")
        params.append(f"%{genotype}%")
    if not is_admin(g.user):
        clauses.append("c.lab_id = ?")
        params.append(g.user.lab_id)

    rows = db().execute(
        f"""
        SELECT a.id, a.animal_code, a.sex, a.dob, a.strain, a.genotype, a.status, a.cage_id, c.cage_code
        FROM animals a
        LEFT JOIN cages c ON c.id = a.cage_id
        WHERE {' AND '.join(clauses)}
        ORDER BY a.id DESC
        LIMIT 1000
        """,
        params,
    ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.get("/api/animals/<int:animal_id>/pedigree")
@require_auth()
def animal_pedigree(animal_id: int) -> Response:
    generations = max(1, min(int(request.args.get("generations", 3)), 6))
    root = db().execute(
        """
        SELECT a.id, a.animal_code, a.sex, a.dob, a.strain, a.genotype, a.status, a.sire_id, a.dam_id
        FROM animals a
        LEFT JOIN cages c ON c.id = a.cage_id
        WHERE a.id = ?
        """
        + ("" if is_admin(g.user) else " AND c.lab_id = ? "),
        (animal_id,) if is_admin(g.user) else (animal_id, g.user.lab_id),
    ).fetchone()
    if not root:
        return jsonify({"error": "Not found"}), 404

    queue: list[tuple[int, int]] = [(animal_id, 0)]
    seen: set[int] = set()
    nodes: dict[int, dict[str, Any]] = {}
    edges: list[dict[str, Any]] = []

    while queue:
        aid, depth = queue.pop(0)
        if aid in seen or depth > generations:
            continue
        seen.add(aid)
        row = db().execute(
            """
            SELECT a.id, a.animal_code, a.sex, a.dob, a.strain, a.genotype, a.status, a.sire_id, a.dam_id
            FROM animals a
            LEFT JOIN cages c ON c.id = a.cage_id
            WHERE a.id = ?
            """
            + ("" if is_admin(g.user) else " AND c.lab_id = ? "),
            (aid,) if is_admin(g.user) else (aid, g.user.lab_id),
        ).fetchone()
        if not row:
            continue
        nodes[aid] = dict(row)
        if depth == generations:
            continue
        for rel, pid in (("sire", row["sire_id"]), ("dam", row["dam_id"])):
            if pid:
                edges.append({"from": aid, "to": pid, "relation": rel})
                queue.append((int(pid), depth + 1))

    return jsonify({"rootId": animal_id, "generations": generations, "nodes": list(nodes.values()), "edges": edges})


@app.get("/api/breeding/productivity")
@require_auth()
def breeding_productivity() -> Response:
    min_litters = int(request.args.get("minLitters", 0))
    scope = ""
    params: list[Any] = []
    if not is_admin(g.user):
        scope = " WHERE c.lab_id = ? "
        params.append(g.user.lab_id)

    rows = db().execute(
        f"""
        SELECT c.id AS cage_id, c.cage_code, c.breeding_status,
               COUNT(l.id) AS litter_count,
               COALESCE(AVG(l.survived_count), 0) AS avg_survived,
               MAX(l.birth_date) AS last_litter_date
        FROM cages c
        LEFT JOIN litters l ON l.cage_id = c.id
        {scope}
        GROUP BY c.id
        HAVING COUNT(l.id) >= ?
        ORDER BY litter_count DESC, avg_survived DESC
        LIMIT 500
        """,
        params + [min_litters],
    ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.get("/api/breeding/non-productive")
@require_auth()
def breeding_non_productive() -> Response:
    stale_days = int(request.args.get("staleDays", 45))
    cutoff = (datetime.now(UTC).date() - timedelta(days=stale_days)).isoformat()

    params: list[Any] = []
    scope = ""
    if not is_admin(g.user):
        scope = " AND c.lab_id = ? "
        params.append(g.user.lab_id)
    params.append(cutoff)
    rows = db().execute(
        """
        SELECT c.id AS cage_id, c.cage_code, c.breeding_status, MAX(l.birth_date) AS last_litter_date, COUNT(l.id) AS litter_count
        FROM cages c
        LEFT JOIN litters l ON l.cage_id = c.id
        WHERE c.breeding_status IN ('Breeding', 'Timed Mating', 'Holding')
        """
        + scope
        + """
        GROUP BY c.id
        HAVING (MAX(l.birth_date) IS NULL OR MAX(l.birth_date) < ?)
        ORDER BY last_litter_date ASC
        LIMIT 500
        """,
        params,
    ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.get("/api/tasks/reminders")
@require_auth()
def task_reminders() -> Response:
    window_days = int(request.args.get("windowDays", 14))
    today = datetime.now(UTC).date().isoformat()
    horizon = (datetime.now(UTC).date() + timedelta(days=window_days)).isoformat()
    params: list[Any] = [today, horizon]
    scope = ""
    if not is_admin(g.user):
        scope = " AND c.lab_id = ? "
        params.append(g.user.lab_id)
    rows = db().execute(
        """
        SELECT b.id, b.event_type, b.event_date, b.cage_id, c.cage_code, b.assigned_to, u.full_name AS assignee
        FROM breeding_events b
        JOIN cages c ON c.id = b.cage_id
        LEFT JOIN users u ON u.id = b.assigned_to
        WHERE b.event_date <= ?
          AND b.event_date >= ?
        """
        + scope
        + """
        ORDER BY b.event_date ASC
        LIMIT 500
        """,
        [horizon, today] + ([g.user.lab_id] if not is_admin(g.user) else []),
    ).fetchall()
    out = []
    for r in rows:
        d = dict(r)
        d["overdue"] = d["event_date"] < today
        out.append(d)
    return jsonify(out)


@app.get("/api/genotyping/mendelian")
@require_auth()
def mendelian_tracking() -> Response:
    scope = ""
    params: list[Any] = []
    if not is_admin(g.user):
        scope = " AND c.lab_id = ? "
        params.append(g.user.lab_id)

    rows = db().execute(
        """
        SELECT l.id AS litter_id, l.birth_date, c.cage_code, a.genotype, COUNT(a.id) AS n
        FROM litters l
        JOIN cages c ON c.id = l.cage_id
        LEFT JOIN animals a ON a.litter_id = l.id
        WHERE 1 = 1
        """
        + scope
        + """
        GROUP BY l.id, a.genotype
        ORDER BY l.birth_date DESC
        LIMIT 2000
        """,
        params,
    ).fetchall()

    by_litter: dict[int, dict[str, Any]] = {}
    for r in rows:
        lid = r["litter_id"]
        if lid not in by_litter:
            by_litter[lid] = {
                "litterId": lid,
                "birthDate": r["birth_date"],
                "cageCode": r["cage_code"],
                "observed": {},
            }
        gkey = r["genotype"] or "Unknown"
        by_litter[lid]["observed"][gkey] = int(r["n"])

    result = []
    for payload in by_litter.values():
        observed = payload["observed"]
        total = sum(observed.values())
        if total <= 0:
            continue
        expected = {k: round(1.0 / len(observed), 4) for k in observed.keys()}
        observed_ratio = {k: round(v / total, 4) for k, v in observed.items()}
        payload["expectedRatio"] = expected
        payload["observedRatio"] = observed_ratio
        payload["totalGenotyped"] = total
        result.append(payload)
    return jsonify(result)


@app.get("/api/genotyping/alerts")
@require_auth()
def genotyping_alerts() -> Response:
    threshold = float(request.args.get("threshold", 0.25))
    tracked = mendelian_tracking().get_json()
    alerts = []
    for row in tracked:
        max_dev = 0.0
        for gkey, exp in row["expectedRatio"].items():
            obs = row["observedRatio"].get(gkey, 0.0)
            max_dev = max(max_dev, abs(obs - exp))
        if max_dev >= threshold:
            alerts.append(
                {
                    "litterId": row["litterId"],
                    "cageCode": row["cageCode"],
                    "birthDate": row["birthDate"],
                    "maxDeviation": round(max_dev, 4),
                    "threshold": threshold,
                }
            )
    return jsonify(alerts)


def _sample_workflow_state(sample_status: str | None, item_result: str | None, order_status: str | None) -> str:
    status = (sample_status or "").strip().lower()
    order = (order_status or "").strip().lower()
    if item_result:
        return "resulted"
    if status == "rejected":
        return "blocked"
    if order in {"received", "closed"} and not item_result:
        return "missing_result"
    if status == "received":
        return "with_provider"
    if status == "shipped":
        return "in_transit"
    return "ready_to_ship"


def _order_scope_row(order_id: int, user: AuthContext) -> storage.Row | None:
    return db().execute(
        """
        SELECT o.*, l.name AS lab_name, p.project_code
        FROM genotyping_orders o
        JOIN labs l ON l.id = o.lab_id
        LEFT JOIN projects p ON p.id = o.project_id
        WHERE o.id = ?
        """
        + ("" if is_admin(user) else " AND o.lab_id = ?"),
        (order_id,) if is_admin(user) else (order_id, user.lab_id),
    ).fetchone()


def _order_items_with_sample_context(order_id: int) -> list[dict[str, Any]]:
    rows = db().execute(
        """
        SELECT i.id, i.sample_id, i.animal_id, i.marker_panel, i.result, i.result_at,
               s.sample_code, s.status AS sample_status, s.provider AS sample_provider, s.tracking_number, s.collected_on,
               a.animal_code
        FROM genotyping_order_items i
        LEFT JOIN sample_records s ON s.id = i.sample_id
        LEFT JOIN animals a ON a.id = i.animal_id
        WHERE i.order_id = ?
        ORDER BY i.id ASC
        """,
        (order_id,),
    ).fetchall()
    return [dict(r) for r in rows]


def _serialize_genotyping_reconciliation(order: storage.Row | dict[str, Any], items: list[dict[str, Any]]) -> dict[str, Any]:
    order_status = order["status"] if order and "status" in order.keys() else None
    counts = {
        "resulted": 0,
        "missing_result": 0,
        "with_provider": 0,
        "in_transit": 0,
        "ready_to_ship": 0,
        "blocked": 0,
    }
    sample_status_counts: dict[str, int] = {}
    enriched: list[dict[str, Any]] = []
    for item in items:
        workflow_state = _sample_workflow_state(item.get("sample_status"), item.get("result"), order_status)
        counts[workflow_state] = counts.get(workflow_state, 0) + 1
        sample_status = item.get("sample_status") or "unknown"
        sample_status_counts[sample_status] = sample_status_counts.get(sample_status, 0) + 1
        payload = dict(item)
        payload["workflowState"] = workflow_state
        enriched.append(payload)
    return {
        "summary": {
            "expectedItems": len(items),
            "resultedItems": counts["resulted"],
            "missingResultItems": counts["missing_result"],
            "withProviderItems": counts["with_provider"],
            "inTransitItems": counts["in_transit"],
            "readyToShipItems": counts["ready_to_ship"],
            "blockedItems": counts["blocked"],
            "completionPct": round((counts["resulted"] / max(1, len(items))) * 100, 1),
            "sampleStatuses": sample_status_counts,
        },
        "items": enriched,
    }


def _append_sample_event(sample_id: int, event_type: str, actor_user_id: int | None, details: dict[str, Any] | None) -> None:
    db().execute(
        "INSERT INTO sample_events (sample_id, event_type, event_time, actor_user_id, details_json) VALUES (?, ?, ?, ?, ?)",
        (sample_id, event_type, now_iso(), actor_user_id, json.dumps(details or {}, default=str)),
    )


def _apply_order_results(
    order: storage.Row | dict[str, Any],
    results: list[dict[str, Any]],
    *,
    order_status: str,
    genotype_source: str,
    actor_user_id: int | None,
) -> int:
    updated = 0
    for item in results:
        result = str(item.get("result") or "").strip()
        if not result:
            continue
        resolved: storage.Row | None = None
        sample_code = str(item.get("sampleCode") or "").strip()
        if sample_code:
            resolved = db().execute(
                """
                SELECT i.id, i.sample_id, i.animal_id, s.status AS sample_status
                FROM genotyping_order_items i
                JOIN sample_records s ON s.id = i.sample_id
                WHERE i.order_id = ? AND s.sample_code = ?
                """,
                (order["id"], sample_code),
            ).fetchone()
        elif item.get("animalCode"):
            resolved = db().execute(
                """
                SELECT i.id, i.sample_id, i.animal_id, s.status AS sample_status
                FROM genotyping_order_items i
                JOIN animals a ON a.id = i.animal_id
                LEFT JOIN sample_records s ON s.id = i.sample_id
                WHERE i.order_id = ? AND a.animal_code = ?
                """,
                (order["id"], str(item.get("animalCode")).strip()),
            ).fetchone()
        if not resolved or not resolved["animal_id"]:
            continue
        db().execute(
            "UPDATE genotyping_order_items SET result = ?, result_at = ?, marker_panel = COALESCE(?, marker_panel) WHERE id = ?",
            (result, now_iso(), item.get("markerPanel"), resolved["id"]),
        )
        db().execute(
            "INSERT INTO genotype_results (animal_id, result, source, created_at) VALUES (?, ?, ?, ?)",
            (resolved["animal_id"], result, genotype_source, now_iso()),
        )
        db().execute("UPDATE animals SET genotype = ?, updated_at = ? WHERE id = ?", (result, now_iso(), resolved["animal_id"]))
        if resolved["sample_id"] and (resolved["sample_status"] or "") != "resulted":
            db().execute("UPDATE sample_records SET status = 'resulted' WHERE id = ?", (resolved["sample_id"],))
            _append_sample_event(
                int(resolved["sample_id"]),
                "resulted",
                actor_user_id,
                {"orderRef": order["order_ref"], "source": genotype_source, "result": result},
            )
        updated += 1
    db().execute("UPDATE genotyping_orders SET status = ?, updated_at = ? WHERE id = ?", (order_status, now_iso(), order["id"]))
    return updated


@app.get("/api/genotyping/dashboard")
@require_auth()
def genotyping_dashboard() -> Response:
    sample_scope = ""
    order_scope = ""
    animal_scope = ""
    sample_params: list[Any] = []
    order_params: list[Any] = []
    animal_params: list[Any] = []
    if not is_admin(g.user):
        sample_scope = " WHERE c.lab_id = ? "
        order_scope = " WHERE o.lab_id = ? "
        animal_scope = " WHERE c.lab_id = ? "
        sample_params.append(g.user.lab_id)
        order_params.append(g.user.lab_id)
        animal_params.append(g.user.lab_id)

    sample_rows = db().execute(
        f"""
        SELECT s.status, COUNT(*) AS n
        FROM sample_records s
        LEFT JOIN cages c ON c.id = s.cage_id
        {sample_scope}
        GROUP BY s.status
        ORDER BY n DESC, s.status ASC
        """,
        sample_params,
    ).fetchall()
    order_rows = db().execute(
        f"""
        SELECT o.status, COUNT(*) AS n
        FROM genotyping_orders o
        {order_scope}
        GROUP BY o.status
        ORDER BY n DESC, o.status ASC
        """,
        order_params,
    ).fetchall()
    provider_rows = db().execute(
        f"""
        SELECT o.provider,
               COUNT(DISTINCT o.id) AS order_count,
               COUNT(i.id) AS item_count,
               SUM(CASE WHEN i.result IS NOT NULL THEN 1 ELSE 0 END) AS resulted_count
        FROM genotyping_orders o
        LEFT JOIN genotyping_order_items i ON i.order_id = o.id
        {order_scope}
        GROUP BY o.provider
        ORDER BY order_count DESC, o.provider ASC
        LIMIT 12
        """,
        order_params,
    ).fetchall()
    genotype_rows = db().execute(
        f"""
        SELECT a.genotype, COUNT(*) AS n
        FROM animals a
        LEFT JOIN cages c ON c.id = a.cage_id
        {animal_scope}
        {" AND " if animal_scope else " WHERE "} COALESCE(a.genotype, '') <> ''
        GROUP BY a.genotype
        ORDER BY n DESC, a.genotype ASC
        LIMIT 12
        """,
        animal_params,
    ).fetchall()
    turnaround_rows = db().execute(
        f"""
        SELECT s.collected_on, s.status
        FROM sample_records s
        LEFT JOIN cages c ON c.id = s.cage_id
        {sample_scope}
        ORDER BY s.id DESC
        LIMIT 1000
        """,
        sample_params,
    ).fetchall()
    sample_activity_rows = db().execute(
        f"""
        SELECT 'sample' AS kind, s.sample_code AS ref_code, e.event_type AS label, e.event_time AS happened_at
        FROM sample_events e
        JOIN sample_records s ON s.id = e.sample_id
        LEFT JOIN cages c ON c.id = s.cage_id
        {sample_scope}
        ORDER BY e.id DESC
        LIMIT 8
        """,
        sample_params,
    ).fetchall()
    order_activity_rows = db().execute(
        f"""
        SELECT 'order' AS kind, o.order_ref AS ref_code, o.status AS label, o.updated_at AS happened_at
        FROM genotyping_orders o
        {order_scope}
        ORDER BY o.id DESC
        LIMIT 8
        """,
        order_params,
    ).fetchall()

    turnaround_buckets = {"0-2d": 0, "3-7d": 0, "8-14d": 0, "15d+": 0}
    today = datetime.now(UTC).date()
    for row in turnaround_rows:
        try:
            collected = date.fromisoformat(str(row["collected_on"]))
            age = max(0, (today - collected).days)
        except ValueError:
            continue
        if age <= 2:
            turnaround_buckets["0-2d"] += 1
        elif age <= 7:
            turnaround_buckets["3-7d"] += 1
        elif age <= 14:
            turnaround_buckets["8-14d"] += 1
        else:
            turnaround_buckets["15d+"] += 1

    recent_activity = sorted(
        [dict(r) for r in sample_activity_rows] + [dict(r) for r in order_activity_rows],
        key=lambda row: str(row.get("happened_at") or ""),
        reverse=True,
    )[:10]

    return jsonify(
        {
            "sampleStatus": [{"label": r["status"], "value": int(r["n"])} for r in sample_rows],
            "orderStatus": [{"label": r["status"], "value": int(r["n"])} for r in order_rows],
            "providers": [
                {
                    "provider": r["provider"] or "Unspecified",
                    "orders": int(r["order_count"]),
                    "items": int(r["item_count"] or 0),
                    "resulted": int(r["resulted_count"] or 0),
                    "pending": max(0, int(r["item_count"] or 0) - int(r["resulted_count"] or 0)),
                }
                for r in provider_rows
            ],
            "genotypeDistribution": [{"label": r["genotype"], "value": int(r["n"])} for r in genotype_rows],
            "turnaround": [{"label": label, "value": value} for label, value in turnaround_buckets.items()],
            "recentActivity": recent_activity,
        }
    )


@app.get("/api/genotyping/providers")
@require_auth()
def genotyping_provider_presets() -> Response:
    return jsonify(GENOTYPING_PROVIDER_PRESETS)


@app.get("/api/genotyping/cohorts")
@require_auth()
def genotyping_cohort_insights() -> Response:
    cage_clause, cage_params = scoped_lab_clause(g.user, "c.lab_id")
    pair_clause, pair_params = scoped_lab_clause(g.user, "bp.lab_id")
    project_filters = ["p.status = 'active'"]
    project_params: list[Any] = []
    if not is_admin(g.user):
        project_filters.append("p.lab_id = ?")
        project_params.append(g.user.lab_id)

    projects = db().execute(
        """
        SELECT p.id, p.lab_id, p.project_code, p.title, p.target_animals, l.name AS lab_name
        FROM projects p
        JOIN labs l ON l.id = p.lab_id
        WHERE """
        + " AND ".join(project_filters)
        + """
        ORDER BY p.id DESC
        LIMIT 200
        """,
        project_params,
    ).fetchall()
    project_ids = [int(row["id"]) for row in projects]
    project_targets = _project_target_map(project_ids)

    assignment_filters = ["1 = 1"]
    assignment_params: list[Any] = []
    if not is_admin(g.user):
        assignment_filters.append("p.lab_id = ?")
        assignment_params.append(g.user.lab_id)
    assignment_rows = db().execute(
        """
        SELECT pa.project_id, pa.animal_id, pa.status, p.project_code
        FROM project_animal_assignments pa
        JOIN projects p ON p.id = pa.project_id
        WHERE """
        + " AND ".join(assignment_filters),
        assignment_params,
    ).fetchall()
    assignment_by_animal = {int(row["animal_id"]): dict(row) for row in assignment_rows if str(row["status"] or "") != "released"}
    reserved_counts: dict[int, int] = {}
    status_counts: dict[int, dict[str, int]] = {project_id: _empty_assignment_status_counts() for project_id in project_ids}
    for row in assignment_rows:
        project_id = int(row["project_id"])
        status = str(row["status"] or "reserved")
        project_status_counts = status_counts.setdefault(project_id, _empty_assignment_status_counts())
        project_status_counts[status] = project_status_counts.get(status, 0) + 1
        if status == "reserved":
            reserved_counts[project_id] = reserved_counts.get(project_id, 0) + 1

    animal_rows = db().execute(
        f"""
        SELECT a.id, a.animal_code, a.sex, a.genotype, a.cage_id, c.cage_code, c.lab_id, l.name AS lab_name
        FROM animals a
        JOIN cages c ON c.id = a.cage_id
        JOIN labs l ON l.id = c.lab_id
        {cage_clause}
        {" AND " if cage_clause else " WHERE "} a.status = 'Active' AND COALESCE(a.genotype, '') <> ''
        ORDER BY a.id DESC
        LIMIT 2000
        """,
        cage_params,
    ).fetchall()

    animals: list[dict[str, Any]] = []
    for row in animal_rows:
        animal = {
            "id": int(row["id"]),
            "animalCode": row["animal_code"],
            "sex": row["sex"],
            "genotype": row["genotype"],
            "cageId": int(row["cage_id"]),
            "cageCode": row["cage_code"],
            "labId": int(row["lab_id"]),
            "labName": row["lab_name"],
            "matchingProjects": [],
            "assignment": assignment_by_animal.get(int(row["id"])),
        }
        for project in projects:
            if not is_admin(g.user) and int(g.user.lab_id or -1) != int(row["lab_id"]):
                continue
            if int(project["id"]) not in project_ids:
                continue
            targets = project_targets.get(int(project["id"]), [])
            if targets and not any(_match_genotype_pattern(row["genotype"], target["genotype_pattern"]) for target in targets):
                continue
            animal["matchingProjects"].append(
                {
                    "id": int(project["id"]),
                    "projectCode": project["project_code"],
                    "title": project["title"],
                    "labName": project["lab_name"],
                }
            )
        animals.append(animal)

    cohort_projects = []
    for row in projects:
        targets = project_targets.get(int(row["id"]), [])
        matching_animals = [animal for animal in animals if any(project["id"] == int(row["id"]) for project in animal["matchingProjects"])]
        ready = len(matching_animals)
        reserved = reserved_counts.get(int(row["id"]), 0)
        target_total = sum(int(target["target_count"] or 0) for target in targets) if targets else int(row["target_animals"] or 0)
        pressure = max(target_total - reserved, 0)
        if ready > reserved and pressure > 0:
            action = "assign_now"
        elif ready == 0 and target_total > 0:
            action = "breed_more"
        else:
            action = "monitor"
        cohort_projects.append(
            {
                "id": int(row["id"]),
                "labId": int(row["lab_id"]),
                "projectCode": row["project_code"],
                "title": row["title"],
                "labName": row["lab_name"],
                "targetAnimals": int(row["target_animals"] or 0),
                "targetRules": [
                    {
                        "id": int(target["id"]),
                        "genotypePattern": target["genotype_pattern"],
                        "targetCount": int(target["target_count"] or 0),
                        "priority": int(target["priority"] or 1),
                        "notes": target["notes"],
                    }
                    for target in targets
                ],
                "matchedReadyAnimals": ready,
                "reservedAnimals": reserved,
                "assignmentPressure": pressure,
                "recommendedAction": action,
                "statusFlow": [
                    {
                        "key": step["key"],
                        "label": step["label"],
                        "value": int(status_counts.get(int(row["id"]), _empty_assignment_status_counts()).get(step["key"], 0)),
                        "color": step["color"],
                    }
                    for step in ASSIGNMENT_STATUS_STEPS
                ],
            }
        )

    breeder_rows = db().execute(
        f"""
        SELECT bp.id, bp.status, c.id AS cage_id, c.cage_code, sire.animal_code AS sire_code, dam.animal_code AS dam_code,
               COUNT(DISTINCT l.id) AS litter_count,
               COALESCE(AVG(l.survived_count), 0) AS avg_survived,
               COUNT(DISTINCT CASE WHEN a.status = 'Active' AND COALESCE(a.genotype, '') <> '' THEN a.id END) AS ready_animals
        FROM breeding_pairs bp
        JOIN cages c ON c.id = bp.cage_id
        JOIN animals sire ON sire.id = bp.sire_id
        JOIN animals dam ON dam.id = bp.dam_id
        LEFT JOIN litters l ON l.cage_id = bp.cage_id
        LEFT JOIN animals a ON a.cage_id = bp.cage_id
        {pair_clause}
        GROUP BY bp.id, bp.status, c.id, c.cage_code, sire.animal_code, dam.animal_code
        ORDER BY bp.id DESC
        LIMIT 20
        """,
        pair_params,
    ).fetchall()
    breeder_signals = []
    for row in breeder_rows:
        ready = int(row["ready_animals"] or 0)
        litter_count = int(row["litter_count"] or 0)
        avg_survived = float(row["avg_survived"] or 0)
        if row["status"] == "active" and ready >= 4:
            signal = "cohort_ready_pause_soon"
            note = "Enough genotype-ready output is visible to consider pausing this pair after the current cohort."
        elif row["status"] == "active" and litter_count == 0:
            signal = "await_first_output"
            note = "Pair is active but has not produced a tracked litter yet."
        elif row["status"] == "active" and avg_survived < 2:
            signal = "review_low_yield"
            note = "Low average survival suggests this pair should be reviewed before expanding demand."
        else:
            signal = "maintain"
            note = "Current output supports continued monitoring rather than immediate breeder changes."
        breeder_signals.append(
            {
                "pairId": int(row["id"]),
                "status": row["status"],
                "cageId": int(row["cage_id"]),
                "cageCode": row["cage_code"],
                "sireCode": row["sire_code"],
                "damCode": row["dam_code"],
                "litterCount": litter_count,
                "avgSurvived": round(avg_survived, 2),
                "readyAnimals": ready,
                "signal": signal,
                "note": note,
            }
        )

    unassigned_ready = [animal for animal in animals if not animal["assignment"]]
    return jsonify(
        {
            "projects": cohort_projects[:12],
            "readyAnimals": animals[:20],
            "unassignedReadyCount": len(unassigned_ready),
            "breederSignals": breeder_signals,
        }
    )


@app.get("/api/forecast/cage-space")
@require_auth(("PI", "Admin"))
def forecast_cage_space() -> Response:
    days = int(request.args.get("days", 30))
    if is_admin(g.user):
        rooms = db().execute(
            """
            SELECT r.id, r.name, r.capacity, COUNT(c.id) AS occupied
            FROM rooms r
            LEFT JOIN cages c ON c.room_id = r.id
            GROUP BY r.id
            """
        ).fetchall()
        created_recent = db().execute(
            "SELECT COUNT(*) AS c FROM cages WHERE created_at >= ?",
            ((datetime.now(UTC) - timedelta(days=30)).isoformat(),),
        ).fetchone()["c"]
    else:
        rooms = db().execute(
            """
            SELECT r.id, r.name, r.capacity, SUM(CASE WHEN c.lab_id = ? THEN 1 ELSE 0 END) AS occupied
            FROM rooms r
            LEFT JOIN cages c ON c.room_id = r.id
            GROUP BY r.id
            """,
            (g.user.lab_id,),
        ).fetchall()
        created_recent = db().execute(
            "SELECT COUNT(*) AS c FROM cages WHERE created_at >= ? AND lab_id = ?",
            ((datetime.now(UTC) - timedelta(days=30)).isoformat(), g.user.lab_id),
        ).fetchone()["c"]

    retire_recent = db().execute(
        """
        SELECT COUNT(*) AS c
        FROM lifecycle_events le
        JOIN cages c ON c.id = le.cage_id
        WHERE le.event_type = 'retire' AND le.created_at >= ?
        """
        + ("" if is_admin(g.user) else " AND c.lab_id = ? "),
        ((datetime.now(UTC) - timedelta(days=30)).isoformat(),)
        if is_admin(g.user)
        else ((datetime.now(UTC) - timedelta(days=30)).isoformat(), g.user.lab_id),
    ).fetchone()["c"]

    net_per_day = (created_recent - retire_recent) / 30.0
    out = []
    for r in rooms:
        projected = int(round(r["occupied"] + net_per_day * days))
        out.append(
            {
                "roomId": r["id"],
                "roomName": r["name"],
                "capacity": r["capacity"],
                "occupiedNow": r["occupied"],
                "projectedOccupied": projected,
                "projectedUtilizationPct": round((projected * 100.0) / r["capacity"], 2) if r["capacity"] else None,
            }
        )
    return jsonify({"days": days, "netCageDeltaPerDay": round(net_per_day, 3), "rooms": out})


@app.get("/api/forecast/consolidation")
@require_auth(("PI", "Admin"))
def forecast_consolidation() -> Response:
    max_animals = int(request.args.get("maxAnimals", 2))
    scope = ""
    params: list[Any] = [max_animals]
    if not is_admin(g.user):
        scope = " AND c.lab_id = ? "
        params.append(g.user.lab_id)
    rows = db().execute(
        """
        SELECT c.id, c.cage_code, c.strain, c.genotype_summary, c.room_id, c.rack_id, (c.male_count + c.female_count) AS total_animals
        FROM cages c
        WHERE c.breeding_status = 'Holding'
          AND (c.male_count + c.female_count) <= ?
        """
        + scope
        + """
        ORDER BY c.room_id, c.strain, c.genotype_summary, total_animals ASC
        LIMIT 2000
        """,
        params,
    ).fetchall()

    grouped: dict[tuple[Any, ...], list[storage.Row]] = {}
    for r in rows:
        key = (r["room_id"], r["strain"], r["genotype_summary"])
        grouped.setdefault(key, []).append(r)

    recommendations = []
    for key, candidates in grouped.items():
        while len(candidates) >= 2:
            a = candidates.pop(0)
            b = candidates.pop(0)
            recommendations.append(
                {
                    "roomId": key[0],
                    "strain": key[1],
                    "genotype": key[2],
                    "fromCages": [a["cage_code"], b["cage_code"]],
                    "combinedAnimals": int(a["total_animals"]) + int(b["total_animals"]),
                }
            )
    return jsonify(recommendations)


@app.get("/api/analytics/summary")
@require_auth()
def analytics_summary() -> Response:
    if is_admin(g.user):
        total_cages = db().execute("SELECT COUNT(*) AS c FROM cages").fetchone()["c"]
        total_animals = db().execute("SELECT COUNT(*) AS c FROM animals WHERE status = 'Active'").fetchone()["c"]
        sex = db().execute("SELECT sex, COUNT(*) AS c FROM animals WHERE status = 'Active' GROUP BY sex").fetchall()
    else:
        total_cages = db().execute("SELECT COUNT(*) AS c FROM cages WHERE lab_id = ?", (g.user.lab_id,)).fetchone()["c"]
        total_animals = db().execute(
            """
            SELECT COUNT(*) AS c
            FROM animals a
            JOIN cages c ON c.id = a.cage_id
            WHERE a.status = 'Active' AND c.lab_id = ?
            """,
            (g.user.lab_id,),
        ).fetchone()["c"]
        sex = db().execute(
            """
            SELECT a.sex, COUNT(*) AS c
            FROM animals a
            JOIN cages c ON c.id = a.cage_id
            WHERE a.status = 'Active' AND c.lab_id = ?
            GROUP BY a.sex
            """,
            (g.user.lab_id,),
        ).fetchall()
    sex_map = {r["sex"]: r["c"] for r in sex}

    litters = db().execute("SELECT litter_size, survived_count FROM litters ORDER BY id DESC LIMIT 200").fetchall()
    survival = 0.0
    if litters:
        total_litter_size = sum(r["litter_size"] for r in litters)
        total_survived = sum(r["survived_count"] for r in litters)
        survival = round((total_survived / total_litter_size) * 100, 2) if total_litter_size else 0.0

    if is_admin(g.user):
        room_capacity = db().execute(
            """
            SELECT r.id, r.name, r.capacity, COUNT(c.id) AS occupied
            FROM rooms r
            LEFT JOIN cages c ON c.room_id = r.id
            GROUP BY r.id
            """
        ).fetchall()
    else:
        room_capacity = db().execute(
            """
            SELECT r.id, r.name, r.capacity, SUM(CASE WHEN c.lab_id = ? THEN 1 ELSE 0 END) AS occupied
            FROM rooms r
            LEFT JOIN cages c ON c.room_id = r.id
            GROUP BY r.id
            """,
            (g.user.lab_id,),
        ).fetchall()

    upcoming_tasks = db().execute(
        """
        SELECT b.event_type, b.event_date, c.cage_code
        FROM breeding_events b
        JOIN cages c ON b.cage_id = c.id
        WHERE b.event_date BETWEEN ? AND ?
        """
        + ("" if is_admin(g.user) else " AND c.lab_id = ? ")
        + """
        ORDER BY b.event_date ASC
        LIMIT 20
        """,
        (datetime.now(UTC).date().isoformat(), (datetime.now(UTC).date() + timedelta(days=14)).isoformat())
        if is_admin(g.user)
        else (datetime.now(UTC).date().isoformat(), (datetime.now(UTC).date() + timedelta(days=14)).isoformat(), g.user.lab_id),
    ).fetchall()

    cohort_scope = "" if is_admin(g.user) else " WHERE p.lab_id = ? "
    cohort_params: tuple[Any, ...] = () if is_admin(g.user) else (g.user.lab_id,)
    cohort_rows = db().execute(
        """
        SELECT pa.status, COUNT(*) AS count
        FROM project_animal_assignments pa
        JOIN projects p ON p.id = pa.project_id
        """
        + cohort_scope
        + """
        GROUP BY pa.status
        """,
        cohort_params,
    ).fetchall()
    cohort_counts = {str(row["status"] or "reserved"): int(row["count"] or 0) for row in cohort_rows}

    lab_rows = db().execute(
        """
        SELECT l.id AS lab_id, l.name AS lab_name,
               COUNT(DISTINCT CASE WHEN p.status = 'active' THEN p.id END) AS active_projects,
               COALESCE(SUM(CASE WHEN pa.status = 'reserved' THEN 1 ELSE 0 END), 0) AS reserved_count,
               COALESCE(SUM(CASE WHEN pa.status = 'assigned' THEN 1 ELSE 0 END), 0) AS assigned_count,
               COALESCE(SUM(CASE WHEN pa.status = 'shipped' THEN 1 ELSE 0 END), 0) AS shipped_count,
               COALESCE(SUM(CASE WHEN pa.status = 'consumed' THEN 1 ELSE 0 END), 0) AS consumed_count,
               COALESCE(SUM(CASE WHEN pa.status = 'released' THEN 1 ELSE 0 END), 0) AS released_count
        FROM labs l
        LEFT JOIN projects p ON p.lab_id = l.id
        LEFT JOIN project_animal_assignments pa ON pa.project_id = p.id
        """
        + ("" if is_admin(g.user) else " WHERE l.id = ? ")
        + """
        GROUP BY l.id, l.name
        ORDER BY (COALESCE(SUM(CASE WHEN pa.status IN ('reserved', 'assigned', 'shipped') THEN 1 ELSE 0 END), 0)
               + COALESCE(SUM(CASE WHEN pa.status IN ('consumed', 'released') THEN 1 ELSE 0 END), 0)) DESC, l.name ASC
        LIMIT 20
        """,
        () if is_admin(g.user) else (g.user.lab_id,),
    ).fetchall()
    cohort_labs = []
    for row in lab_rows:
        active_count = int(row["reserved_count"] or 0) + int(row["assigned_count"] or 0) + int(row["shipped_count"] or 0)
        completed_count = int(row["consumed_count"] or 0) + int(row["released_count"] or 0)
        cohort_labs.append(
            {
                "labId": int(row["lab_id"]),
                "labName": row["lab_name"],
                "activeProjects": int(row["active_projects"] or 0),
                "activeAssignments": active_count,
                "completedAssignments": completed_count,
                "completionPct": round((completed_count / max(active_count + completed_count, 1)) * 100, 1),
                "statusFlow": [
                    {"key": "reserved", "label": "Reserved", "value": int(row["reserved_count"] or 0), "color": "#4f8ef7"},
                    {"key": "assigned", "label": "Assigned", "value": int(row["assigned_count"] or 0), "color": "#18a172"},
                    {"key": "shipped", "label": "Shipped", "value": int(row["shipped_count"] or 0), "color": "#eb9c44"},
                    {"key": "consumed", "label": "Consumed", "value": int(row["consumed_count"] or 0), "color": "#7c6cf2"},
                    {"key": "released", "label": "Released", "value": int(row["released_count"] or 0), "color": "#64748b"},
                ],
            }
        )
    completed_total = int(cohort_counts.get("consumed", 0)) + int(cohort_counts.get("released", 0))
    active_total = int(cohort_counts.get("reserved", 0)) + int(cohort_counts.get("assigned", 0)) + int(cohort_counts.get("shipped", 0))

    return jsonify(
        {
            "totalCages": total_cages,
            "totalActiveAnimals": total_animals,
            "sexRatio": {"M": sex_map.get("M", 0), "F": sex_map.get("F", 0)},
            "pupSurvivalPct": survival,
            "roomCapacity": [dict(r) for r in room_capacity],
            "upcomingTasks": [dict(r) for r in upcoming_tasks],
            "cohortFlow": [
                {"key": step["key"], "label": step["label"], "value": int(cohort_counts.get(step["key"], 0)), "color": step["color"]}
                for step in ASSIGNMENT_STATUS_STEPS
            ],
            "cohortDisposition": [
                {"label": "Consumed", "key": "consumed", "value": int(cohort_counts.get("consumed", 0)), "color": "#7c6cf2"},
                {"label": "Released", "key": "released", "value": int(cohort_counts.get("released", 0)), "color": "#64748b"},
            ],
            "cohortCompletion": {
                "activeAssignments": active_total,
                "completedAssignments": completed_total,
                "completionPct": round((completed_total / max(active_total + completed_total, 1)) * 100, 1),
            },
            "cohortLabs": cohort_labs,
        }
    )


@app.post("/api/forecast/demand")
@require_auth(("PI", "Admin"))
def forecast() -> Response:
    payload = request.get_json(force=True)
    needed_by = payload.get("neededBy")
    requested = int(payload.get("animalsNeeded", 0))
    if is_admin(g.user):
        active = db().execute("SELECT COUNT(*) AS c FROM animals WHERE status = 'Active'").fetchone()["c"]
    else:
        active = db().execute(
            """
            SELECT COUNT(*) AS c
            FROM animals a
            JOIN cages c ON c.id = a.cage_id
            WHERE a.status = 'Active' AND c.lab_id = ?
            """,
            (g.user.lab_id,),
        ).fetchone()["c"]
    deficit = max(requested - active, 0)

    if is_admin(g.user):
        avg_litter_survival = db().execute(
            "SELECT AVG(CAST(survived_count AS FLOAT) / NULLIF(litter_size, 0)) AS v FROM litters"
        ).fetchone()["v"]
    else:
        avg_litter_survival = db().execute(
            """
            SELECT AVG(CAST(l.survived_count AS FLOAT) / NULLIF(l.litter_size, 0)) AS v
            FROM litters l
            JOIN cages c ON c.id = l.cage_id
            WHERE c.lab_id = ?
            """,
            (g.user.lab_id,),
        ).fetchone()["v"]
    avg_litter_survival = float(avg_litter_survival) if avg_litter_survival else 0.75
    expected_per_litter = max(int(round(6 * avg_litter_survival)), 1)

    litters_required = (deficit + expected_per_litter - 1) // expected_per_litter
    return jsonify(
        {
            "neededBy": needed_by,
            "requested": requested,
            "activeNow": active,
            "deficit": deficit,
            "estimatedLittersRequired": litters_required,
            "assumptions": {
                "baselinePupsPerLitter": 6,
                "avgSurvivalFraction": round(avg_litter_survival, 2),
            },
        }
    )


@app.post("/api/import/excel")
@require_auth(("Admin",))
def import_excel() -> Response:
    if "file" not in request.files:
        return jsonify({"error": "File missing"}), 400

    from openpyxl import load_workbook

    wb = load_workbook(request.files["file"], data_only=True)
    ws = wb.active

    created = 0
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        cage_code, strain, genotype, breeding_status, male, female = row[:6]
        if not cage_code:
            continue
        try:
            db().execute(
                """
                INSERT INTO cages (
                    cage_code, strain, genotype_summary, breeding_status, male_count, female_count,
                    room_id, rack_id, lab_id, protocol_id, qr_token, created_at, updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, 1, 1, 1, 1, ?, ?, ?)
                """,
                (
                    str(cage_code),
                    str(strain or "Unknown"),
                    str(genotype or "Unknown"),
                    str(breeding_status or "Holding"),
                    int(male or 0),
                    int(female or 0),
                    secrets.token_urlsafe(12),
                    now_iso(),
                    now_iso(),
                ),
            )
            created += 1
        except storage.IntegrityError:
            app.logger.warning("Skipping duplicate cage code row %s", idx)
    db().commit()
    audit_log(g.user.user_id, "import", "excel", "bulk_import", None, {"created": created})
    return jsonify({"created": created})


@app.get("/api/reports/cages.csv")
@require_auth()
def report_cages_csv() -> Response:
    query = "SELECT cage_code, strain, genotype_summary, breeding_status, male_count, female_count, dob FROM cages"
    params: tuple[Any, ...] = ()
    if not is_admin(g.user):
        query += " WHERE lab_id = ?"
        params = (g.user.lab_id,)
    query += " ORDER BY cage_code"
    rows = db().execute(query, params).fetchall()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["cage_code", "strain", "genotype", "status", "male", "female", "dob"])
    for r in rows:
        writer.writerow([r["cage_code"], r["strain"], r["genotype_summary"], r["breeding_status"], r["male_count"], r["female_count"], r["dob"]])
    return Response(output.getvalue(), mimetype="text/csv", headers={"Content-Disposition": "attachment; filename=cages_report.csv"})


@app.get("/api/reports/cages.xlsx")
@require_auth()
def report_cages_xlsx() -> Response:
    query = "SELECT cage_code, strain, genotype_summary, breeding_status, male_count, female_count, dob FROM cages"
    params: tuple[Any, ...] = ()
    if not is_admin(g.user):
        query += " WHERE lab_id = ?"
        params = (g.user.lab_id,)
    query += " ORDER BY cage_code"
    rows = db().execute(query, params).fetchall()
    wb = Workbook()
    ws = wb.active
    ws.title = "Cages"
    ws.append(["cage_code", "strain", "genotype", "status", "male", "female", "dob"])
    for r in rows:
        ws.append([r["cage_code"], r["strain"], r["genotype_summary"], r["breeding_status"], r["male_count"], r["female_count"], r["dob"]])
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(
        bio,
        as_attachment=True,
        download_name="cages_report.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.get("/api/reports/cages.pdf")
@require_auth()
def report_cages_pdf() -> Response:
    query = "SELECT cage_code, strain, genotype_summary, breeding_status, male_count, female_count, dob FROM cages"
    params: tuple[Any, ...] = ()
    if not is_admin(g.user):
        query += " WHERE lab_id = ?"
        params = (g.user.lab_id,)
    query += " ORDER BY cage_code"
    rows = db().execute(query, params).fetchall()
    lines = [f"Murisphere Cage Report - generated {datetime.now(UTC).date().isoformat()}", ""]
    for r in rows:
        lines.append(
            f"{r['cage_code']} | {r['strain']} | {r['genotype_summary']} | {r['breeding_status']} | M:{r['male_count']} F:{r['female_count']} | DOB:{r['dob']}"
        )
    return Response(simple_pdf(lines), mimetype="application/pdf", headers={"Content-Disposition": "attachment; filename=cages_report.pdf"})


@app.post("/api/cages/cards")
@require_auth()
def cage_cards() -> Response:
    payload = request.get_json(force=True)
    ids = payload.get("ids", [])
    if not ids:
        return jsonify({"error": "Provide at least one cage ID"}), 400
    placeholders = ",".join("?" for _ in ids)
    scope_clause = ""
    params: list[Any] = list(ids)
    if not is_admin(g.user):
        scope_clause = " AND c.lab_id = ?"
        params.append(g.user.lab_id)
    rows = db().execute(
        f"""
        SELECT c.id, c.cage_code, c.strain, c.genotype_summary, c.breeding_status, c.dob,
               c.male_count, c.female_count,
               l.name AS lab_name, l.pi_name AS pi_name,
               p.protocol_number, p.title AS protocol_title, p.expires_on AS protocol_expires_on,
               r.name AS room_name, k.name AS rack_name, c.qr_token,
               (
                   SELECT {PROJECT_CODE_LIST_SQL}
                   FROM project_cages pc
                   JOIN projects pj ON pj.id = pc.project_id
                   WHERE pc.cage_id = c.id
               ) AS project_codes
        FROM cages c
        LEFT JOIN labs l ON c.lab_id = l.id
        LEFT JOIN iacuc_protocols p ON c.protocol_id = p.id
        LEFT JOIN rooms r ON c.room_id = r.id
        LEFT JOIN racks k ON c.rack_id = k.id
        WHERE c.id IN ({placeholders}) {scope_clause}
        """,
        params,
    ).fetchall()

    cage_ids = [int(r["id"]) for r in rows]
    animals_by_cage: dict[int, list[dict[str, Any]]] = {cage_id: [] for cage_id in cage_ids}
    litters_by_cage: dict[int, list[dict[str, Any]]] = {cage_id: [] for cage_id in cage_ids}
    if cage_ids:
        animal_placeholders = ",".join("?" for _ in cage_ids)
        animal_rows = db().execute(
            f"""
            SELECT cage_id, animal_code, sex, dob, genotype, status
            FROM animals
            WHERE cage_id IN ({animal_placeholders})
            ORDER BY cage_id, created_at ASC, id ASC
            """,
            cage_ids,
        ).fetchall()
        for a in animal_rows:
            animals_by_cage[int(a["cage_id"])].append(
                {
                    "animalCode": a["animal_code"],
                    "sex": a["sex"],
                    "dob": a["dob"],
                    "genotype": a["genotype"],
                    "status": a["status"],
                }
            )

        litter_rows = db().execute(
            f"""
            SELECT l.id, l.cage_id, l.birth_date, l.litter_size, l.survived_count, l.weaned_on,
                   COALESCE(SUM(CASE WHEN a.sex = 'M' THEN 1 ELSE 0 END), 0) AS male_count,
                   COALESCE(SUM(CASE WHEN a.sex = 'F' THEN 1 ELSE 0 END), 0) AS female_count
            FROM litters l
            LEFT JOIN animals a ON a.litter_id = l.id
            WHERE l.cage_id IN ({animal_placeholders})
            GROUP BY l.id, l.cage_id, l.birth_date, l.litter_size, l.survived_count, l.weaned_on
            ORDER BY l.cage_id, l.birth_date DESC, l.id DESC
            """,
            cage_ids,
        ).fetchall()
        for l in litter_rows:
            litters_by_cage[int(l["cage_id"])].append(
                {
                    "litterId": l["id"],
                    "birthDate": l["birth_date"],
                    "born": l["litter_size"],
                    "survived": l["survived_count"],
                    "maleCount": l["male_count"],
                    "femaleCount": l["female_count"],
                    "dow": l["weaned_on"],
                }
            )

    cards = []
    for r in rows:
        projects = []
        if r["project_codes"]:
            projects = [x.strip() for x in str(r["project_codes"]).split(",") if x.strip()]
        cards.append(
            {
                "cageId": r["id"],
                "cageCode": r["cage_code"],
                "strain": r["strain"],
                "genotype": r["genotype_summary"],
                "groupOwner": r["pi_name"],
                "groupName": r["lab_name"],
                "piLab": r["lab_name"],
                "breedingStatus": r["breeding_status"],
                "dob": r["dob"],
                "animalCount": {"M": r["male_count"], "F": r["female_count"]},
                "protocol": r["protocol_number"],
                "protocolDescription": r["protocol_title"],
                "protocolExpiresOn": r["protocol_expires_on"],
                "projects": projects,
                "roomName": r["room_name"],
                "rackName": r["rack_name"],
                "location": f"{r['room_name']} / {r['rack_name']}",
                "animals": animals_by_cage.get(int(r["id"]), []),
                "litters": litters_by_cage.get(int(r["id"]), []),
                "qrValue": r["qr_token"],
                "scanUrl": f"/scan/{r['qr_token']}",
            }
        )
    return jsonify(cards)


@app.get("/api/compliance/protocol-alerts")
@require_auth()
def protocol_alerts() -> Response:
    cutoff = (datetime.now(UTC).date() + timedelta(days=45)).isoformat()
    rows = db().execute(
        "SELECT protocol_number, title, expires_on FROM iacuc_protocols WHERE expires_on <= ? ORDER BY expires_on ASC",
        (cutoff,),
    ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.get("/api/facility/capacity")
@require_auth(("Admin", "PI"))
def facility_capacity() -> Response:
    if is_admin(g.user):
        rooms = db().execute(
            """
            SELECT r.id, r.name, r.capacity,
                   COUNT(c.id) AS occupied,
                   ROUND((COUNT(c.id) * 100.0) / NULLIF(r.capacity, 0), 2) AS utilization_pct
            FROM rooms r
            LEFT JOIN cages c ON c.room_id = r.id
            GROUP BY r.id
            ORDER BY utilization_pct DESC
            """
        ).fetchall()
    else:
        facility = db().execute(
            """
            SELECT l.facility_id
            FROM labs l
            WHERE l.id = ?
            """,
            (g.user.lab_id,),
        ).fetchone()
        if not facility:
            return jsonify([])
        rooms = db().execute(
            """
            SELECT r.id, r.name, r.capacity,
                   COUNT(c.id) AS occupied,
                   ROUND((COUNT(c.id) * 100.0) / NULLIF(r.capacity, 0), 2) AS utilization_pct
            FROM rooms r
            LEFT JOIN cages c ON c.room_id = r.id
            WHERE r.facility_id = ?
            GROUP BY r.id
            ORDER BY utilization_pct DESC
            """,
            (facility["facility_id"],),
        ).fetchall()
    return jsonify([dict(r) for r in rooms])


@app.get("/api/facilities")
@require_auth(("Admin", "PI"))
def list_facilities() -> Response:
    rows = db().execute(
        """
        SELECT f.id, f.name, f.timezone,
               COUNT(DISTINCT l.id) AS labs,
               COUNT(DISTINCT r.id) AS rooms,
               COUNT(DISTINCT c.id) AS cages
        FROM facilities f
        LEFT JOIN labs l ON l.facility_id = f.id
        LEFT JOIN rooms r ON r.facility_id = f.id
        LEFT JOIN cages c ON c.room_id = r.id
        GROUP BY f.id
        ORDER BY f.name
        """
    ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.get("/api/facility/quotas")
@require_auth(("Admin", "PI"))
def facility_quotas() -> Response:
    if is_admin(g.user):
        rows = db().execute(
            """
            SELECT l.id AS lab_id,
                   l.name AS lab_name,
                   COALESCE(lp.size_tier, 'unassigned') AS size_tier,
                   COALESCE(lp.expected_cage_load, 0) AS expected_cage_load,
                   COALESCE(lp.active_project_count, 0) AS expected_active_projects,
                   COUNT(DISTINCT c.id) AS current_cages,
                   COUNT(DISTINCT p.id) AS current_projects
            FROM labs l
            LEFT JOIN lab_profiles lp ON lp.lab_id = l.id
            LEFT JOIN cages c ON c.lab_id = l.id
            LEFT JOIN projects p ON p.lab_id = l.id
            GROUP BY l.id, l.name, lp.size_tier, lp.expected_cage_load, lp.active_project_count
            ORDER BY current_cages DESC
            """
        ).fetchall()
    else:
        rows = db().execute(
            """
            SELECT l.id AS lab_id,
                   l.name AS lab_name,
                   COALESCE(lp.size_tier, 'unassigned') AS size_tier,
                   COALESCE(lp.expected_cage_load, 0) AS expected_cage_load,
                   COALESCE(lp.active_project_count, 0) AS expected_active_projects,
                   COUNT(DISTINCT c.id) AS current_cages,
                   COUNT(DISTINCT p.id) AS current_projects
            FROM labs l
            LEFT JOIN lab_profiles lp ON lp.lab_id = l.id
            LEFT JOIN cages c ON c.lab_id = l.id
            LEFT JOIN projects p ON p.lab_id = l.id
            WHERE l.id = ?
            GROUP BY l.id, l.name, lp.size_tier, lp.expected_cage_load, lp.active_project_count
            """,
            (g.user.lab_id,),
        ).fetchall()

    results = []
    for r in rows:
        expected_load = int(r["expected_cage_load"] or 0)
        current = int(r["current_cages"] or 0)
        remaining = expected_load - current
        utilization = round((current * 100.0) / expected_load, 2) if expected_load > 0 else None
        results.append(
            {
                "labId": r["lab_id"],
                "labName": r["lab_name"],
                "sizeTier": r["size_tier"],
                "expectedCageLoad": expected_load,
                "expectedActiveProjects": int(r["expected_active_projects"] or 0),
                "currentCages": current,
                "currentProjects": int(r["current_projects"] or 0),
                "remainingQuota": remaining,
                "utilizationPct": utilization,
            }
        )
    return jsonify(results)


@app.get("/api/facility/chargeback")
@require_auth(("Admin", "PI"))
def facility_chargeback() -> Response:
    period_days = int(request.args.get("periodDays", 30))
    rate_per_cage_day = float(request.args.get("ratePerCageDay", 0.85))
    if is_admin(g.user):
        rows = db().execute(
            """
            SELECT l.id AS lab_id, l.name AS lab_name, COUNT(c.id) AS cage_count
            FROM labs l
            LEFT JOIN cages c ON c.lab_id = l.id
            GROUP BY l.id, l.name
            ORDER BY cage_count DESC
            """
        ).fetchall()
    else:
        rows = db().execute(
            """
            SELECT l.id AS lab_id, l.name AS lab_name, COUNT(c.id) AS cage_count
            FROM labs l
            LEFT JOIN cages c ON c.lab_id = l.id
            WHERE l.id = ?
            GROUP BY l.id, l.name
            """,
            (g.user.lab_id,),
        ).fetchall()

    out = []
    for r in rows:
        cage_days = int(r["cage_count"]) * period_days
        amount = round(cage_days * rate_per_cage_day, 2)
        out.append(
            {
                "labId": r["lab_id"],
                "labName": r["lab_name"],
                "periodDays": period_days,
                "cageCount": int(r["cage_count"]),
                "cageDays": cage_days,
                "ratePerCageDay": rate_per_cage_day,
                "estimatedCharge": amount,
            }
        )
    return jsonify(out)


@app.get("/api/reports/breeder-productivity.csv")
@require_auth(("PI", "Admin"))
def report_breeder_productivity() -> Response:
    rows = breeding_productivity().get_json()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["cage_code", "breeding_status", "litter_count", "avg_survived", "last_litter_date"])
    for r in rows:
        writer.writerow([r["cage_code"], r["breeding_status"], r["litter_count"], r["avg_survived"], r["last_litter_date"]])
    return Response(output.getvalue(), mimetype="text/csv", headers={"Content-Disposition": "attachment; filename=breeder_productivity.csv"})


@app.get("/api/reports/survival.csv")
@require_auth(("PI", "Admin"))
def report_survival() -> Response:
    query = "SELECT l.id, l.birth_date, l.litter_size, l.survived_count, c.cage_code FROM litters l JOIN cages c ON c.id = l.cage_id"
    params: tuple[Any, ...] = ()
    if not is_admin(g.user):
        query += " WHERE c.lab_id = ?"
        params = (g.user.lab_id,)
    query += " ORDER BY l.birth_date DESC LIMIT 2000"
    rows = db().execute(query, params).fetchall()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["litter_id", "birth_date", "cage_code", "litter_size", "survived_count", "survival_pct"])
    for r in rows:
        pct = round((r["survived_count"] * 100.0) / r["litter_size"], 2) if r["litter_size"] else 0.0
        writer.writerow([r["id"], r["birth_date"], r["cage_code"], r["litter_size"], r["survived_count"], pct])
    return Response(output.getvalue(), mimetype="text/csv", headers={"Content-Disposition": "attachment; filename=survival_report.csv"})


@app.get("/api/reports/protocol-usage.csv")
@require_auth(("PI", "Admin"))
def report_protocol_usage() -> Response:
    query = """
        SELECT p.protocol_number, p.title, l.name AS lab_name, COUNT(c.id) AS cages
        FROM iacuc_protocols p
        JOIN labs l ON l.id = p.lab_id
        LEFT JOIN cages c ON c.protocol_id = p.id
    """
    params: tuple[Any, ...] = ()
    if not is_admin(g.user):
        query += " WHERE p.lab_id = ? "
        params = (g.user.lab_id,)
    query += " GROUP BY p.id, l.name ORDER BY cages DESC"
    rows = db().execute(query, params).fetchall()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["protocol_number", "title", "lab", "cages"])
    for r in rows:
        writer.writerow([r["protocol_number"], r["title"], r["lab_name"], r["cages"]])
    return Response(output.getvalue(), mimetype="text/csv", headers={"Content-Disposition": "attachment; filename=protocol_usage.csv"})


@app.get("/api/billing/rules")
@require_auth(("Admin", "PI"))
def billing_rules_list() -> Response:
    query = """
        SELECT br.id, br.lab_id, br.room_id, br.line_type, br.rate, br.service_name, br.active, br.created_at,
               l.name AS lab_name, r.name AS room_name
        FROM billing_rules br
        LEFT JOIN labs l ON l.id = br.lab_id
        LEFT JOIN rooms r ON r.id = br.room_id
    """
    params: tuple[Any, ...] = ()
    if not is_admin(g.user):
        query += " WHERE br.lab_id = ? OR br.lab_id IS NULL "
        params = (g.user.lab_id,)
    query += " ORDER BY br.id DESC"
    rows = db().execute(query, params).fetchall()
    return jsonify([dict(r) for r in rows])


@app.post("/api/billing/rules")
@require_auth(("Admin",))
def billing_rule_create() -> Response:
    payload = request.get_json(force=True)
    line_type = str(payload.get("lineType", "per_diem")).strip()
    rate = float(payload.get("rate", 0))
    if line_type not in {"per_diem", "service"}:
        return jsonify({"error": "Invalid lineType"}), 400
    if rate < 0:
        return jsonify({"error": "rate must be non-negative"}), 400
    cur = db().execute(
        "INSERT INTO billing_rules (lab_id, room_id, line_type, rate, service_name, active, created_at) VALUES (?, ?, ?, ?, ?, ?, ?)",
        (
            payload.get("labId"),
            payload.get("roomId"),
            line_type,
            rate,
            payload.get("serviceName"),
            1 if payload.get("active", True) else 0,
            now_iso(),
        ),
    )
    db().commit()
    audit_log(g.user.user_id, "billing_rule", cur.lastrowid, "create", None, payload)
    return jsonify({"id": cur.lastrowid}), 201


@app.post("/api/billing/run")
@require_auth(("Admin", "PI"))
def billing_run() -> Response:
    payload = request.get_json(force=True)
    period_start = str(payload.get("periodStart", "")).strip()
    period_end = str(payload.get("periodEnd", "")).strip()
    if not period_start or not period_end:
        return jsonify({"error": "periodStart and periodEnd are required"}), 400
    if period_end < period_start:
        return jsonify({"error": "periodEnd must be >= periodStart"}), 400

    # Ensure period exists and is open
    period = db().execute(
        "SELECT id, status FROM billing_periods WHERE period_start = ? AND period_end = ?",
        (period_start, period_end),
    ).fetchone()
    if not period:
        db().execute(
            "INSERT INTO billing_periods (period_start, period_end, status, created_at) VALUES (?, ?, 'open', ?)",
            (period_start, period_end, now_iso()),
        )
        db().commit()
        period = db().execute(
            "SELECT id, status FROM billing_periods WHERE period_start = ? AND period_end = ?",
            (period_start, period_end),
        ).fetchone()
    if period["status"] == "closed":
        return jsonify({"error": "Billing period is closed"}), 409

    # Derive per-diem rate per cage (lab-specific override > global)
    cages = db().execute(
        """
        SELECT c.id, c.lab_id, c.cage_code
        FROM cages c
        """
        + ("" if is_admin(g.user) else " WHERE c.lab_id = ? "),
        () if is_admin(g.user) else (g.user.lab_id,),
    ).fetchall()

    try:
        days = (datetime.fromisoformat(period_end) - datetime.fromisoformat(period_start)).days + 1
    except ValueError:
        return jsonify({"error": "Invalid date format; use YYYY-MM-DD"}), 400
    days = max(days, 1)
    created = 0
    for c in cages:
        rule = db().execute(
            """
            SELECT rate
            FROM billing_rules
            WHERE active = 1 AND line_type = 'per_diem'
              AND (lab_id = ? OR lab_id IS NULL)
            ORDER BY CASE WHEN lab_id = ? THEN 0 ELSE 1 END, id DESC
            LIMIT 1
            """,
            (c["lab_id"], c["lab_id"]),
        ).fetchone()
        rate = float(rule["rate"]) if rule else 0.85
        qty = float(days)
        amount = round(qty * rate, 2)
        db().execute(
            storage.sql_upsert(
                "billing_entries",
                [
                    "period_start",
                    "period_end",
                    "lab_id",
                    "cage_id",
                    "line_type",
                    "quantity",
                    "rate",
                    "amount",
                    "description",
                    "created_at",
                ],
                ["period_start", "period_end", "lab_id", "cage_id", "line_type", "description"],
                ["quantity", "rate", "amount", "created_at"],
            ),
            (
                period_start,
                period_end,
                c["lab_id"],
                c["id"],
                "per_diem",
                qty,
                rate,
                amount,
                f"Cage {c['cage_code']} per-diem",
                now_iso(),
            ),
        )
        created += 1
    db().commit()
    audit_log(g.user.user_id, "billing_period", f"{period_start}:{period_end}", "run", None, {"entries": created})
    return jsonify({"entriesUpserted": created, "periodStart": period_start, "periodEnd": period_end})


@app.post("/api/billing/close-period")
@require_auth(("Admin",))
def billing_close_period() -> Response:
    payload = request.get_json(force=True)
    period_start = str(payload.get("periodStart", "")).strip()
    period_end = str(payload.get("periodEnd", "")).strip()
    row = db().execute(
        "SELECT id, status FROM billing_periods WHERE period_start = ? AND period_end = ?",
        (period_start, period_end),
    ).fetchone()
    if not row:
        return jsonify({"error": "Period not found"}), 404
    if row["status"] == "closed":
        return jsonify({"ok": True, "status": "closed"})
    db().execute(
        "UPDATE billing_periods SET status = 'closed', closed_by = ?, closed_at = ? WHERE id = ?",
        (g.user.user_id, now_iso(), row["id"]),
    )
    db().commit()
    audit_log(g.user.user_id, "billing_period", row["id"], "close", None, {"periodStart": period_start, "periodEnd": period_end})
    return jsonify({"ok": True, "status": "closed"})


@app.get("/api/billing/statements.csv")
@require_auth(("Admin", "PI"))
def billing_statements_csv() -> Response:
    period_start = request.args.get("periodStart", "").strip()
    period_end = request.args.get("periodEnd", "").strip()
    clauses = []
    params: list[Any] = []
    if period_start:
        clauses.append("be.period_start = ?")
        params.append(period_start)
    if period_end:
        clauses.append("be.period_end = ?")
        params.append(period_end)
    if not is_admin(g.user):
        clauses.append("be.lab_id = ?")
        params.append(g.user.lab_id)
    where = f"WHERE {' AND '.join(clauses)}" if clauses else ""
    rows = db().execute(
        f"""
        SELECT be.period_start, be.period_end, l.name AS lab_name, be.line_type,
               SUM(be.quantity) AS quantity, AVG(be.rate) AS rate, SUM(be.amount) AS amount
        FROM billing_entries be
        JOIN labs l ON l.id = be.lab_id
        {where}
        GROUP BY be.period_start, be.period_end, be.lab_id, l.name, be.line_type
        ORDER BY be.period_start DESC, lab_name
        """,
        params,
    ).fetchall()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["period_start", "period_end", "lab", "line_type", "quantity", "avg_rate", "amount"])
    for r in rows:
        writer.writerow([r["period_start"], r["period_end"], r["lab_name"], r["line_type"], r["quantity"], r["rate"], r["amount"]])
    return Response(output.getvalue(), mimetype="text/csv", headers={"Content-Disposition": "attachment; filename=billing_statements.csv"})


@app.get("/api/requests")
@require_auth()
def facility_requests_list() -> Response:
    params: list[Any] = []
    where = ""
    if not is_admin(g.user):
        where = "WHERE fr.lab_id = ?"
        params.append(g.user.lab_id)
    rows = db().execute(
        f"""
        SELECT fr.id, fr.request_type, fr.status, fr.details_json, fr.created_at, fr.updated_at,
               l.name AS lab_name, p.project_code, u.full_name AS requested_by_name
        FROM facility_requests fr
        JOIN labs l ON l.id = fr.lab_id
        LEFT JOIN projects p ON p.id = fr.project_id
        LEFT JOIN users u ON u.id = fr.requested_by
        {where}
        ORDER BY fr.id DESC
        LIMIT 500
        """,
        params,
    ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.post("/api/requests")
@require_auth(("Technician", "PI", "Admin"))
def facility_request_create() -> Response:
    payload = request.get_json(force=True)
    lab_id = int(payload.get("labId", g.user.lab_id or 1))
    if not is_admin(g.user) and g.user.lab_id != lab_id:
        return jsonify({"error": "Forbidden"}), 403
    cur = db().execute(
        """
        INSERT INTO facility_requests
        (request_type, lab_id, project_id, status, details_json, requested_by, created_at, updated_at)
        VALUES (?, ?, ?, 'submitted', ?, ?, ?, ?)
        """,
        (
            payload.get("requestType", "general"),
            lab_id,
            payload.get("projectId"),
            json.dumps(payload.get("details", {})),
            g.user.user_id,
            now_iso(),
            now_iso(),
        ),
    )
    db().commit()
    audit_log(g.user.user_id, "facility_request", cur.lastrowid, "create", None, payload)
    return jsonify({"id": cur.lastrowid}), 201


@app.post("/api/requests/<int:request_id>/status")
@require_auth(("PI", "Admin"))
def facility_request_status(request_id: int) -> Response:
    payload = request.get_json(force=True)
    status = str(payload.get("status", "")).strip()
    if status not in {"submitted", "approved", "fulfilled", "rejected"}:
        return jsonify({"error": "Invalid status"}), 400
    row = db().execute("SELECT * FROM facility_requests WHERE id = ?", (request_id,)).fetchone()
    if not row:
        return jsonify({"error": "Not found"}), 404
    if not is_admin(g.user) and int(row["lab_id"]) != int(g.user.lab_id or -1):
        return jsonify({"error": "Forbidden"}), 403
    db().execute(
        "UPDATE facility_requests SET status = ?, reviewed_by = ?, updated_at = ? WHERE id = ?",
        (status, g.user.user_id, now_iso(), request_id),
    )
    db().commit()
    audit_log(g.user.user_id, "facility_request", request_id, "status", {"status": row["status"]}, {"status": status})
    return jsonify({"ok": True})


@app.get("/api/operations/sla")
@require_auth(("PI", "Admin"))
def operations_sla() -> Response:
    rows = db().execute(
        f"""
        SELECT request_type, status, AVG({REQUEST_SLA_HOURS_SQL}) AS avg_hours, COUNT(*) AS n
        FROM facility_requests
        """
        + ("" if is_admin(g.user) else " WHERE lab_id = ? ")
        + """
        GROUP BY request_type, status
        ORDER BY n DESC
        """,
        () if is_admin(g.user) else (g.user.lab_id,),
    ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.get("/api/facility/benchmark")
@require_auth(("Admin",))
def facility_benchmark() -> Response:
    rows = db().execute(
        """
        SELECT f.id AS facility_id, f.name AS facility_name,
               COUNT(DISTINCT l.id) AS labs,
               COUNT(DISTINCT c.id) AS cages,
               COUNT(DISTINCT p.id) AS projects,
               COUNT(DISTINCT a.id) AS active_animals
        FROM facilities f
        LEFT JOIN labs l ON l.facility_id = f.id
        LEFT JOIN projects p ON p.lab_id = l.id
        LEFT JOIN cages c ON c.lab_id = l.id
        LEFT JOIN animals a ON a.cage_id = c.id AND a.status = 'Active'
        GROUP BY f.id
        ORDER BY cages DESC
        """
    ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.post("/api/integrations/export-jobs")
@require_auth(("PI", "Admin"))
def create_export_job() -> Response:
    payload = request.get_json(force=True)
    job_type = str(payload.get("jobType", "")).strip()
    if not job_type:
        return jsonify({"error": "jobType is required"}), 400
    cur = db().execute(
        "INSERT INTO export_jobs (job_type, target_url, status, payload_json, created_by, created_at) VALUES (?, ?, 'pending', ?, ?, ?)",
        (job_type, payload.get("targetUrl"), json.dumps(payload.get("payload", {})), g.user.user_id, now_iso()),
    )
    db().commit()
    audit_log(g.user.user_id, "export_job", cur.lastrowid, "create", None, payload)
    return jsonify({"id": cur.lastrowid}), 201


@app.post("/api/integrations/export-jobs/<int:job_id>/run")
@require_auth(("PI", "Admin"))
def run_export_job(job_id: int) -> Response:
    row = db().execute(
        """
        SELECT ej.*, u.lab_id AS creator_lab_id
        FROM export_jobs ej
        LEFT JOIN users u ON u.id = ej.created_by
        WHERE ej.id = ?
        """,
        (job_id,),
    ).fetchone()
    if not row:
        return jsonify({"error": "Not found"}), 404
    if not is_admin(g.user) and int(row["creator_lab_id"] or -1) != int(g.user.lab_id or -1):
        return jsonify({"error": "Not found"}), 404
    payload = row["payload_json"] or "{}"
    target = row["target_url"]
    try:
        if target:
            req = urlrequest.Request(
                target,
                data=payload.encode("utf-8"),
                headers={"Content-Type": "application/json"},
                method="POST",
            )
            with urlrequest.urlopen(req, timeout=8) as resp:
                if resp.status >= 400:
                    raise RuntimeError(f"dispatch_failed_{resp.status}")
        db().execute("UPDATE export_jobs SET status = 'sent', sent_at = ? WHERE id = ?", (now_iso(), job_id))
        db().commit()
        audit_log(g.user.user_id, "export_job", job_id, "run", {"status": row["status"]}, {"status": "sent"})
        return jsonify({"ok": True, "status": "sent"})
    except (urlerror.URLError, TimeoutError, RuntimeError) as exc:
        db().execute("UPDATE export_jobs SET status = 'failed' WHERE id = ?", (job_id,))
        db().commit()
        audit_log(g.user.user_id, "export_job", job_id, "run_failed", {"status": row["status"]}, {"status": "failed", "error": str(exc)})
        return jsonify({"ok": False, "status": "failed", "error": str(exc)}), 502


@app.get("/api/integrations/export-jobs")
@require_auth(("PI", "Admin"))
def list_export_jobs() -> Response:
    rows = db().execute(
        """
        SELECT ej.id, ej.job_type, ej.target_url, ej.status, ej.created_at, ej.sent_at
        FROM export_jobs ej
        LEFT JOIN users u ON u.id = ej.created_by
        """
        + ("" if is_admin(g.user) else " WHERE u.lab_id = ? ")
        + """
        ORDER BY ej.id DESC
        LIMIT 500
        """,
        () if is_admin(g.user) else (g.user.lab_id,),
    ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.post("/api/census/sessions")
@require_auth(("Technician", "PI", "Admin"))
def create_census_session() -> Response:
    payload = request.get_json(force=True)
    room_id = payload.get("roomId")
    cur = db().execute(
        "INSERT INTO cage_census_sessions (room_id, started_by, started_at, status, notes) VALUES (?, ?, ?, 'active', ?)",
        (room_id, g.user.user_id, now_iso(), payload.get("notes")),
    )
    db().commit()
    audit_log(g.user.user_id, "census_session", cur.lastrowid, "create", None, payload)
    return jsonify({"id": cur.lastrowid}), 201


@app.post("/api/census/sessions/<int:session_id>/scan")
@require_auth(("Technician", "PI", "Admin"))
def census_scan(session_id: int) -> Response:
    session = db().execute("SELECT * FROM cage_census_sessions WHERE id = ?", (session_id,)).fetchone()
    if not session:
        return jsonify({"error": "Session not found"}), 404
    if session["status"] != "active":
        return jsonify({"error": "Session is not active"}), 409

    payload = request.get_json(force=True)
    code = str(payload.get("code", "")).strip()
    cage = db().execute(
        "SELECT id, lab_id FROM cages WHERE cage_code = ? OR qr_token = ?",
        (code, code),
    ).fetchone()
    if not cage:
        return jsonify({"error": "Cage not found"}), 404
    if not is_admin(g.user) and int(cage["lab_id"]) != int(g.user.lab_id or -1):
        return jsonify({"error": "Not found"}), 404

    db().execute(
        storage.sql_upsert(
            "cage_census_scans",
            [
                "session_id",
                "cage_id",
                "scanned_at",
                "scanned_by",
                "observed_male_count",
                "observed_female_count",
                "observed_status",
            ],
            ["session_id", "cage_id"],
            [
                "scanned_at",
                "scanned_by",
                "observed_male_count",
                "observed_female_count",
                "observed_status",
            ],
        ),
        (
            session_id,
            cage["id"],
            now_iso(),
            g.user.user_id,
            payload.get("maleCount"),
            payload.get("femaleCount"),
            payload.get("breedingStatus"),
        ),
    )
    db().commit()
    return jsonify({"ok": True, "cageId": cage["id"]})


@app.post("/api/census/sessions/<int:session_id>/complete")
@require_auth(("Technician", "PI", "Admin"))
def complete_census_session(session_id: int) -> Response:
    row = db().execute("SELECT * FROM cage_census_sessions WHERE id = ?", (session_id,)).fetchone()
    if not row:
        return jsonify({"error": "Session not found"}), 404
    db().execute(
        "UPDATE cage_census_sessions SET status = 'completed', ended_at = ?, notes = COALESCE(?, notes) WHERE id = ?",
        (now_iso(), (request.get_json(silent=True) or {}).get("notes"), session_id),
    )
    db().commit()
    audit_log(g.user.user_id, "census_session", session_id, "complete", {"status": row["status"]}, {"status": "completed"})
    return jsonify({"ok": True})


@app.get("/api/census/sessions/<int:session_id>")
@require_auth()
def get_census_session(session_id: int) -> Response:
    session = db().execute(
        "SELECT id, room_id, started_by, started_at, ended_at, status, notes FROM cage_census_sessions WHERE id = ?",
        (session_id,),
    ).fetchone()
    if not session:
        return jsonify({"error": "Session not found"}), 404
    scans = db().execute(
        """
        SELECT s.id, s.scanned_at, s.observed_male_count, s.observed_female_count, s.observed_status, c.cage_code
        FROM cage_census_scans s
        JOIN cages c ON c.id = s.cage_id
        WHERE s.session_id = ?
        ORDER BY s.scanned_at ASC
        """,
        (session_id,),
    ).fetchall()
    return jsonify({"session": dict(session), "scans": [dict(r) for r in scans]})


@app.post("/api/orders")
@require_auth(("Technician", "PI", "Admin"))
def create_order() -> Response:
    payload = request.get_json(force=True)
    lab_id = int(payload.get("labId", g.user.lab_id or 1))
    if not is_admin(g.user) and int(g.user.lab_id or -1) != lab_id:
        return jsonify({"error": "Forbidden"}), 403
    quantity = int(payload.get("quantity", 0))
    if quantity <= 0:
        return jsonify({"error": "quantity must be positive"}), 400
    cur = db().execute(
        """
        INSERT INTO animal_orders
        (lab_id, project_id, vendor, strain, sex, quantity, requested_date, needed_by, status, created_by, created_at, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'submitted', ?, ?, ?)
        """,
        (
            lab_id,
            payload.get("projectId"),
            payload.get("vendor"),
            payload.get("strain"),
            payload.get("sex"),
            quantity,
            payload.get("requestedDate", datetime.now(UTC).date().isoformat()),
            payload.get("neededBy"),
            g.user.user_id,
            now_iso(),
            now_iso(),
        ),
    )
    db().commit()
    audit_log(g.user.user_id, "animal_order", cur.lastrowid, "create", None, payload)
    return jsonify({"id": cur.lastrowid}), 201


@app.get("/api/orders")
@require_auth()
def list_orders() -> Response:
    status = request.args.get("status", "").strip()
    clauses = []
    params: list[Any] = []
    if status:
        clauses.append("o.status = ?")
        params.append(status)
    if not is_admin(g.user):
        clauses.append("o.lab_id = ?")
        params.append(g.user.lab_id)
    where = f"WHERE {' AND '.join(clauses)}" if clauses else ""
    rows = db().execute(
        f"""
        SELECT o.*, l.name AS lab_name, p.project_code, u.full_name AS creator_name
        FROM animal_orders o
        JOIN labs l ON l.id = o.lab_id
        LEFT JOIN projects p ON p.id = o.project_id
        LEFT JOIN users u ON u.id = o.created_by
        {where}
        ORDER BY o.id DESC
        LIMIT 500
        """,
        params,
    ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.post("/api/orders/<int:order_id>/status")
@require_auth(("PI", "Admin"))
def order_status(order_id: int) -> Response:
    row = ensure_order_scope(order_id, g.user)
    if not row:
        return jsonify({"error": "Not found"}), 404
    payload = request.get_json(force=True)
    status = str(payload.get("status", "")).strip()
    if status not in {"submitted", "approved", "ordered", "received", "cancelled"}:
        return jsonify({"error": "Invalid status"}), 400
    received_qty = payload.get("receivedQuantity")
    db().execute(
        "UPDATE animal_orders SET status = ?, reviewed_by = ?, received_quantity = COALESCE(?, received_quantity), updated_at = ? WHERE id = ?",
        (status, g.user.user_id, received_qty, now_iso(), order_id),
    )
    db().commit()
    audit_log(g.user.user_id, "animal_order", order_id, "status", {"status": row["status"]}, {"status": status})
    return jsonify({"ok": True})


@app.post("/api/protocols/<int:protocol_id>/versions")
@require_auth(("PI", "Admin"))
def create_protocol_version(protocol_id: int) -> Response:
    protocol = db().execute("SELECT * FROM iacuc_protocols WHERE id = ?", (protocol_id,)).fetchone()
    if not protocol:
        return jsonify({"error": "Not found"}), 404
    if not is_admin(g.user) and int(protocol["lab_id"]) != int(g.user.lab_id or -1):
        return jsonify({"error": "Forbidden"}), 403
    payload = request.get_json(force=True)
    prev = db().execute("SELECT COALESCE(MAX(version_number), 0) AS v FROM protocol_versions WHERE protocol_id = ?", (protocol_id,)).fetchone()
    version = int(payload.get("versionNumber", int(prev["v"]) + 1))
    cur = db().execute(
        """
        INSERT INTO protocol_versions (protocol_id, version_number, title, details_json, effective_on, created_by, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        (
            protocol_id,
            version,
            payload.get("title", protocol["title"]),
            json.dumps(payload.get("details", {})),
            payload.get("effectiveOn", datetime.now(UTC).date().isoformat()),
            g.user.user_id,
            now_iso(),
        ),
    )
    db().commit()
    audit_log(g.user.user_id, "protocol_version", cur.lastrowid, "create", None, payload)
    return jsonify({"id": cur.lastrowid}), 201


@app.get("/api/protocols/<int:protocol_id>/versions")
@require_auth()
def list_protocol_versions(protocol_id: int) -> Response:
    protocol = db().execute("SELECT * FROM iacuc_protocols WHERE id = ?", (protocol_id,)).fetchone()
    if not protocol:
        return jsonify({"error": "Not found"}), 404
    if not is_admin(g.user) and int(protocol["lab_id"]) != int(g.user.lab_id or -1):
        return jsonify({"error": "Not found"}), 404
    rows = db().execute(
        "SELECT id, version_number, title, details_json, effective_on, created_at FROM protocol_versions WHERE protocol_id = ? ORDER BY version_number DESC",
        (protocol_id,),
    ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.post("/api/billing/adjustments")
@require_auth(("Admin",))
def add_billing_adjustment() -> Response:
    payload = request.get_json(force=True)
    cur = db().execute(
        """
        INSERT INTO billing_adjustments (period_start, period_end, lab_id, amount, reason, created_by, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        (
            payload["periodStart"],
            payload["periodEnd"],
            int(payload["labId"]),
            float(payload["amount"]),
            str(payload.get("reason", "")).strip(),
            g.user.user_id,
            now_iso(),
        ),
    )
    db().commit()
    audit_log(g.user.user_id, "billing_adjustment", cur.lastrowid, "create", None, payload)
    return jsonify({"id": cur.lastrowid}), 201


@app.post("/api/billing/review")
@require_auth(("PI", "Admin"))
def billing_review() -> Response:
    payload = request.get_json(force=True)
    lab_id = int(payload.get("labId", g.user.lab_id or 0))
    if not is_admin(g.user) and lab_id != int(g.user.lab_id or -1):
        return jsonify({"error": "Forbidden"}), 403
    status = str(payload.get("reviewStatus", "draft")).strip()
    if status not in {"draft", "approved", "rejected"}:
        return jsonify({"error": "Invalid reviewStatus"}), 400
    db().execute(
        storage.sql_upsert(
            "billing_reviews",
            ["period_start", "period_end", "lab_id", "review_status", "note", "reviewed_by", "reviewed_at"],
            ["period_start", "period_end", "lab_id"],
            ["review_status", "note", "reviewed_by", "reviewed_at"],
        ),
        (
            payload["periodStart"],
            payload["periodEnd"],
            lab_id,
            status,
            payload.get("note"),
            g.user.user_id,
            now_iso(),
        ),
    )
    db().commit()
    return jsonify({"ok": True})


@app.get("/api/billing/rate-model")
@require_auth(("PI", "Admin"))
def billing_rate_model() -> Response:
    labor_per_day = float(request.args.get("laborPerDay", 0.42))
    housing_per_day = float(request.args.get("housingPerDay", 0.28))
    overhead_per_day = float(request.args.get("overheadPerDay", 0.15))
    margin_pct = float(request.args.get("marginPct", 10.0))
    base = labor_per_day + housing_per_day + overhead_per_day
    recommended = round(base * (1.0 + margin_pct / 100.0), 4)
    return jsonify(
        {
            "inputs": {
                "laborPerDay": labor_per_day,
                "housingPerDay": housing_per_day,
                "overheadPerDay": overhead_per_day,
                "marginPct": margin_pct,
            },
            "baseRate": round(base, 4),
            "recommendedPerDiemRate": recommended,
        }
    )


@app.post("/api/vet/cases")
@require_auth(("Technician", "PI", "Admin"))
def create_vet_case() -> Response:
    payload = request.get_json(force=True)
    cage_id = payload.get("cageId")
    animal_id = payload.get("animalId")
    lab_id = int(payload.get("labId", g.user.lab_id or 1))
    if not is_admin(g.user) and lab_id != int(g.user.lab_id or -1):
        return jsonify({"error": "Forbidden"}), 403
    cur = db().execute(
        """
        INSERT INTO vet_cases (cage_id, animal_id, lab_id, case_status, severity, opened_at, opened_by, notes)
        VALUES (?, ?, ?, 'open', ?, ?, ?, ?)
        """,
        (cage_id, animal_id, lab_id, payload.get("severity"), now_iso(), g.user.user_id, payload.get("notes")),
    )
    db().commit()
    audit_log(g.user.user_id, "vet_case", cur.lastrowid, "create", None, payload)
    return jsonify({"id": cur.lastrowid}), 201


@app.get("/api/vet/cases")
@require_auth()
def list_vet_cases() -> Response:
    status = request.args.get("status", "").strip()
    clauses = []
    params: list[Any] = []
    if status:
        clauses.append("vc.case_status = ?")
        params.append(status)
    if not is_admin(g.user):
        clauses.append("vc.lab_id = ?")
        params.append(g.user.lab_id)
    where = f"WHERE {' AND '.join(clauses)}" if clauses else ""
    rows = db().execute(
        f"""
        SELECT vc.id, vc.case_status, vc.severity, vc.opened_at, vc.closed_at, vc.notes,
               c.cage_code, a.animal_code, l.name AS lab_name
        FROM vet_cases vc
        LEFT JOIN cages c ON c.id = vc.cage_id
        LEFT JOIN animals a ON a.id = vc.animal_id
        LEFT JOIN labs l ON l.id = vc.lab_id
        {where}
        ORDER BY vc.id DESC
        LIMIT 500
        """,
        params,
    ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.post("/api/vet/cases/<int:case_id>/treatments")
@require_auth(("Technician", "PI", "Admin"))
def add_treatment(case_id: int) -> Response:
    vc = ensure_vet_case_scope(case_id, g.user)
    if not vc:
        return jsonify({"error": "Not found"}), 404
    payload = request.get_json(force=True)
    cur = db().execute(
        """
        INSERT INTO vet_treatments (case_id, treatment_name, schedule_rule, next_due_on, status, created_by, created_at)
        VALUES (?, ?, ?, ?, 'active', ?, ?)
        """,
        (case_id, payload["treatmentName"], payload.get("scheduleRule"), payload.get("nextDueOn"), g.user.user_id, now_iso()),
    )
    db().commit()
    return jsonify({"id": cur.lastrowid}), 201


@app.post("/api/tasks/assign")
@require_auth(("PI", "Admin"))
def assign_task() -> Response:
    payload = request.get_json(force=True)
    assignee = payload.get("assignedTo")
    required = payload.get("requiredQualification")
    if assignee and required:
        qual = db().execute(
            "SELECT 1 FROM staff_qualifications WHERE user_id = ? AND qualification_code = ?",
            (assignee, required),
        ).fetchone()
        if not qual:
            return jsonify({"error": "Assignee missing required qualification"}), 409
    cur = db().execute(
        """
        INSERT INTO task_assignments (task_type, cage_id, due_on, assigned_to, required_qualification, status, created_by, created_at)
        VALUES (?, ?, ?, ?, ?, 'pending', ?, ?)
        """,
        (
            payload["taskType"],
            payload.get("cageId"),
            payload["dueOn"],
            assignee,
            required,
            g.user.user_id,
            now_iso(),
        ),
    )
    db().commit()
    return jsonify({"id": cur.lastrowid}), 201


@app.post("/api/staff/qualifications")
@require_auth(("Admin",))
def add_staff_qualification() -> Response:
    payload = request.get_json(force=True)
    db().execute(
        storage.sql_upsert(
            "staff_qualifications",
            ["user_id", "qualification_code", "granted_on", "expires_on"],
            ["user_id", "qualification_code"],
            ["granted_on", "expires_on"],
        ),
        (
            int(payload["userId"]),
            str(payload["qualificationCode"]),
            payload.get("grantedOn", datetime.now(UTC).date().isoformat()),
            payload.get("expiresOn"),
        ),
    )
    db().commit()
    return jsonify({"ok": True})


@app.post("/api/attachments")
@require_auth(("Technician", "PI", "Admin"))
def upload_attachment() -> Response:
    entity_type = request.form.get("entityType", "").strip()
    entity_id = request.form.get("entityId", "").strip()
    f = request.files.get("file")
    if not entity_type or not entity_id or f is None:
        return jsonify({"error": "entityType, entityId and file are required"}), 400
    ATTACHMENT_DIR.mkdir(parents=True, exist_ok=True)
    safe_name = f.filename or "attachment.bin"
    unique_name = f"{datetime.now(UTC).strftime('%Y%m%d%H%M%S')}_{secrets.token_hex(6)}_{safe_name}"
    out_path = ATTACHMENT_DIR / unique_name
    f.save(out_path)
    content_type = f.mimetype or mimetypes.guess_type(safe_name)[0]
    cur = db().execute(
        """
        INSERT INTO record_attachments (entity_type, entity_id, filename, file_path, content_type, uploaded_by, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        (entity_type, entity_id, safe_name, str(out_path), content_type, g.user.user_id, now_iso()),
    )
    db().commit()
    return jsonify({"id": cur.lastrowid}), 201


@app.get("/api/attachments")
@require_auth()
def list_attachments() -> Response:
    entity_type = request.args.get("entityType", "").strip()
    entity_id = request.args.get("entityId", "").strip()
    if not entity_type or not entity_id:
        return jsonify({"error": "entityType and entityId required"}), 400
    rows = db().execute(
        """
        SELECT id, entity_type, entity_id, filename, content_type, uploaded_by, created_at
        FROM record_attachments
        WHERE entity_type = ? AND entity_id = ?
        ORDER BY id DESC
        """,
        (entity_type, entity_id),
    ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.get("/api/attachments/<int:attachment_id>/download")
@require_auth()
def download_attachment(attachment_id: int) -> Response:
    row = db().execute("SELECT * FROM record_attachments WHERE id = ?", (attachment_id,)).fetchone()
    if not row:
        return jsonify({"error": "Not found"}), 404
    path = Path(row["file_path"])
    if not path.exists():
        return jsonify({"error": "File missing"}), 404
    return send_file(path, as_attachment=True, download_name=row["filename"], mimetype=row["content_type"] or "application/octet-stream")


@app.post("/api/sign")
@require_auth(("Technician", "PI", "Admin"))
def e_sign() -> Response:
    payload = request.get_json(force=True)
    password = str(payload.get("password", ""))
    user = db().execute("SELECT password_hash FROM users WHERE id = ?", (g.user.user_id,)).fetchone()
    if not user or not check_password_hash(user["password_hash"], password):
        return jsonify({"error": "Signature authentication failed"}), 403
    signature_basis = f"{g.user.user_id}:{payload.get('entityType')}:{payload.get('entityId')}:{payload.get('action')}:{now_iso()}"
    sig_hash = hashlib.sha256(signature_basis.encode("utf-8")).hexdigest()
    cur = db().execute(
        """
        INSERT INTO e_signatures (signer_user_id, entity_type, entity_id, action, reason, signature_hash, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        (
            g.user.user_id,
            payload["entityType"],
            str(payload["entityId"]),
            payload["action"],
            payload.get("reason"),
            sig_hash,
            now_iso(),
        ),
    )
    db().commit()
    return jsonify({"id": cur.lastrowid, "signatureHash": sig_hash})


@app.get("/api/tasks")
@require_auth()
def list_tasks() -> Response:
    status = request.args.get("status", "").strip()
    clauses = []
    params: list[Any] = []
    if status:
        clauses.append("t.status = ?")
        params.append(status)
    if not is_admin(g.user):
        clauses.append("(c.lab_id = ? OR t.assigned_to = ?)")
        params.extend([g.user.lab_id, g.user.user_id])
    where = f"WHERE {' AND '.join(clauses)}" if clauses else ""
    rows = db().execute(
        f"""
        SELECT t.id, t.task_type, t.due_on, t.assigned_to, t.required_qualification, t.status,
               c.cage_code, u.full_name AS assignee_name
        FROM task_assignments t
        LEFT JOIN cages c ON c.id = t.cage_id
        LEFT JOIN users u ON u.id = t.assigned_to
        {where}
        ORDER BY t.due_on ASC, t.id DESC
        LIMIT 1000
        """,
        params,
    ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.post("/api/tasks/<int:task_id>/status")
@require_auth(("Technician", "PI", "Admin"))
def update_task_status(task_id: int) -> Response:
    payload = request.get_json(force=True)
    status = str(payload.get("status", "")).strip()
    if status not in {"pending", "in_progress", "done", "blocked"}:
        return jsonify({"error": "Invalid status"}), 400
    row = db().execute(
        """
        SELECT t.id, t.assigned_to, c.lab_id
        FROM task_assignments t
        LEFT JOIN cages c ON c.id = t.cage_id
        WHERE t.id = ?
        """,
        (task_id,),
    ).fetchone()
    if not row:
        return jsonify({"error": "Not found"}), 404
    if not is_admin(g.user):
        allowed = row["assigned_to"] == g.user.user_id or row["lab_id"] == g.user.lab_id
        if not allowed:
            return jsonify({"error": "Forbidden"}), 403
    db().execute("UPDATE task_assignments SET status = ? WHERE id = ?", (status, task_id))
    db().commit()
    return jsonify({"ok": True})


@app.get("/api/staff/qualification-alerts")
@require_auth(("PI", "Admin"))
def qualification_alerts() -> Response:
    cutoff = (datetime.now(UTC).date() + timedelta(days=30)).isoformat()
    rows = db().execute(
        """
        SELECT q.user_id, u.full_name, q.qualification_code, q.expires_on
        FROM staff_qualifications q
        JOIN users u ON u.id = q.user_id
        WHERE q.expires_on IS NOT NULL AND q.expires_on <= ?
        """
        + ("" if is_admin(g.user) else " AND u.lab_id = ? ")
        + """
        ORDER BY q.expires_on ASC
        """,
        (cutoff,) if is_admin(g.user) else (cutoff, g.user.lab_id),
    ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.post("/api/health/rounds")
@require_auth(("Technician", "PI", "Admin"))
def start_health_round() -> Response:
    payload = request.get_json(force=True)
    cur = db().execute(
        "INSERT INTO health_rounds (room_id, performed_by, started_at, status, notes) VALUES (?, ?, ?, 'active', ?)",
        (payload.get("roomId"), g.user.user_id, now_iso(), payload.get("notes")),
    )
    db().commit()
    return jsonify({"id": cur.lastrowid}), 201


@app.post("/api/health/rounds/<int:round_id>/observe")
@require_auth(("Technician", "PI", "Admin"))
def add_health_observation(round_id: int) -> Response:
    r = db().execute("SELECT * FROM health_rounds WHERE id = ?", (round_id,)).fetchone()
    if not r:
        return jsonify({"error": "Round not found"}), 404
    if r["status"] != "active":
        return jsonify({"error": "Round is not active"}), 409
    payload = request.get_json(force=True)
    cage_id = int(payload["cageId"])
    cage = ensure_cage_scope(cage_id, g.user)
    if not cage:
        return jsonify({"error": "Not found"}), 404
    cur = db().execute(
        """
        INSERT INTO health_observations (round_id, cage_id, finding, severity, action_taken, observed_at, observed_by)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        (
            round_id,
            cage_id,
            str(payload.get("finding", "")).strip(),
            payload.get("severity"),
            payload.get("actionTaken"),
            now_iso(),
            g.user.user_id,
        ),
    )
    db().commit()
    audit_log(g.user.user_id, "health_observation", cur.lastrowid, "create", None, payload)
    return jsonify({"id": cur.lastrowid}), 201


@app.post("/api/health/rounds/<int:round_id>/complete")
@require_auth(("Technician", "PI", "Admin"))
def complete_health_round(round_id: int) -> Response:
    r = db().execute("SELECT * FROM health_rounds WHERE id = ?", (round_id,)).fetchone()
    if not r:
        return jsonify({"error": "Round not found"}), 404
    db().execute("UPDATE health_rounds SET status = 'completed', completed_at = ? WHERE id = ?", (now_iso(), round_id))
    db().commit()
    return jsonify({"ok": True})


@app.get("/api/health/rounds/<int:round_id>")
@require_auth()
def get_health_round(round_id: int) -> Response:
    r = db().execute("SELECT * FROM health_rounds WHERE id = ?", (round_id,)).fetchone()
    if not r:
        return jsonify({"error": "Round not found"}), 404
    obs = db().execute(
        """
        SELECT o.id, o.finding, o.severity, o.action_taken, o.observed_at, c.cage_code
        FROM health_observations o
        JOIN cages c ON c.id = o.cage_id
        WHERE o.round_id = ?
        ORDER BY o.id ASC
        """,
        (round_id,),
    ).fetchall()
    return jsonify({"round": dict(r), "observations": [dict(x) for x in obs]})


@app.post("/api/compliance/deviations")
@require_auth(("Technician", "PI", "Admin"))
def report_deviation() -> Response:
    payload = request.get_json(force=True)
    protocol_id = int(payload["protocolId"])
    protocol = db().execute("SELECT * FROM iacuc_protocols WHERE id = ?", (protocol_id,)).fetchone()
    if not protocol:
        return jsonify({"error": "Protocol not found"}), 404
    if not is_admin(g.user) and int(protocol["lab_id"]) != int(g.user.lab_id or -1):
        return jsonify({"error": "Forbidden"}), 403
    cur = db().execute(
        """
        INSERT INTO protocol_deviations (protocol_id, cage_id, reported_by, reported_at, severity, summary, capa_plan, status)
        VALUES (?, ?, ?, ?, ?, ?, ?, 'open')
        """,
        (
            protocol_id,
            payload.get("cageId"),
            g.user.user_id,
            now_iso(),
            payload.get("severity", "medium"),
            payload.get("summary", ""),
            payload.get("capaPlan"),
        ),
    )
    db().commit()
    audit_log(g.user.user_id, "protocol_deviation", cur.lastrowid, "create", None, payload)
    return jsonify({"id": cur.lastrowid}), 201


@app.get("/api/compliance/deviations")
@require_auth(("Technician", "PI", "Admin"))
def list_deviations() -> Response:
    status = request.args.get("status", "").strip()
    clauses = []
    params: list[Any] = []
    if status:
        clauses.append("d.status = ?")
        params.append(status)
    if not is_admin(g.user):
        clauses.append("p.lab_id = ?")
        params.append(g.user.lab_id)
    where = f"WHERE {' AND '.join(clauses)}" if clauses else ""
    rows = db().execute(
        f"""
        SELECT d.id, d.reported_at, d.severity, d.summary, d.status, d.resolved_at,
               p.protocol_number, c.cage_code
        FROM protocol_deviations d
        JOIN iacuc_protocols p ON p.id = d.protocol_id
        LEFT JOIN cages c ON c.id = d.cage_id
        {where}
        ORDER BY d.id DESC
        LIMIT 500
        """,
        params,
    ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.post("/api/compliance/deviations/<int:deviation_id>/status")
@require_auth(("PI", "Admin"))
def update_deviation_status(deviation_id: int) -> Response:
    payload = request.get_json(force=True)
    status = str(payload.get("status", "")).strip()
    if status not in {"open", "under_review", "closed"}:
        return jsonify({"error": "Invalid status"}), 400
    row = db().execute(
        """
        SELECT d.*, p.lab_id
        FROM protocol_deviations d
        JOIN iacuc_protocols p ON p.id = d.protocol_id
        WHERE d.id = ?
        """,
        (deviation_id,),
    ).fetchone()
    if not row:
        return jsonify({"error": "Not found"}), 404
    if not is_admin(g.user) and int(row["lab_id"]) != int(g.user.lab_id or -1):
        return jsonify({"error": "Forbidden"}), 403
    db().execute(
        "UPDATE protocol_deviations SET status = ?, capa_plan = COALESCE(?, capa_plan), resolved_at = CASE WHEN ? = 'closed' THEN ? ELSE resolved_at END, resolved_by = CASE WHEN ? = 'closed' THEN ? ELSE resolved_by END WHERE id = ?",
        (status, payload.get("capaPlan"), status, now_iso(), status, g.user.user_id, deviation_id),
    )
    db().commit()
    return jsonify({"ok": True})


@app.post("/api/cages/<int:cage_id>/euthanasia")
@require_auth(("Technician", "PI", "Admin"))
def record_euthanasia(cage_id: int) -> Response:
    cage = ensure_cage_scope(cage_id, g.user)
    if not cage:
        return jsonify({"error": "Not found"}), 404
    payload = request.get_json(force=True)
    animal_id = payload.get("animalId")
    if animal_id:
        animal = db().execute("SELECT * FROM animals WHERE id = ? AND cage_id = ?", (animal_id, cage_id)).fetchone()
        if not animal:
            return jsonify({"error": "Animal not found in cage"}), 404
        db().execute("UPDATE animals SET status = 'Euthanized', updated_at = ? WHERE id = ?", (now_iso(), animal_id))
    else:
        male = int(payload.get("male", 0))
        female = int(payload.get("female", 0))
        db().execute(
            """
            UPDATE cages
            SET male_count = CASE WHEN male_count - ? > 0 THEN male_count - ? ELSE 0 END,
                female_count = CASE WHEN female_count - ? > 0 THEN female_count - ? ELSE 0 END,
                updated_at = ?
            WHERE id = ?
            """,
            (male, male, female, female, now_iso(), cage_id),
        )
    cur = db().execute(
        """
        INSERT INTO euthanasia_records (animal_id, cage_id, protocol_id, method, reason, disposition, performed_by, performed_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            animal_id,
            cage_id,
            cage["protocol_id"],
            payload.get("method", "CO2"),
            payload.get("reason"),
            payload.get("disposition"),
            g.user.user_id,
            now_iso(),
        ),
    )
    db().commit()
    audit_log(g.user.user_id, "euthanasia", cur.lastrowid, "create", None, payload)
    return jsonify({"id": cur.lastrowid}), 201


@app.get("/api/reports/euthanasia.csv")
@require_auth(("PI", "Admin"))
def report_euthanasia() -> Response:
    rows = db().execute(
        """
        SELECT e.id, e.performed_at, e.method, e.reason, e.disposition, c.cage_code, a.animal_code
        FROM euthanasia_records e
        LEFT JOIN cages c ON c.id = e.cage_id
        LEFT JOIN animals a ON a.id = e.animal_id
        """
        + ("" if is_admin(g.user) else " WHERE c.lab_id = ? ")
        + """
        ORDER BY e.id DESC
        LIMIT 2000
        """,
        () if is_admin(g.user) else (g.user.lab_id,),
    ).fetchall()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["id", "performed_at", "cage_code", "animal_code", "method", "reason", "disposition"])
    for r in rows:
        writer.writerow([r["id"], r["performed_at"], r["cage_code"], r["animal_code"], r["method"], r["reason"], r["disposition"]])
    return Response(output.getvalue(), mimetype="text/csv", headers={"Content-Disposition": "attachment; filename=euthanasia_report.csv"})


@app.post("/api/cages/<int:cage_id>/wash")
@require_auth(("Technician", "PI", "Admin"))
def queue_cage_wash(cage_id: int) -> Response:
    cage = ensure_cage_scope(cage_id, g.user)
    if not cage:
        return jsonify({"error": "Not found"}), 404
    cur = db().execute(
        "INSERT INTO cage_wash_events (cage_id, status, requested_by, requested_at) VALUES (?, 'queued', ?, ?)",
        (cage_id, g.user.user_id, now_iso()),
    )
    db().commit()
    return jsonify({"id": cur.lastrowid}), 201


@app.post("/api/wash-events/<int:event_id>/status")
@require_auth(("Technician", "PI", "Admin"))
def update_wash_status(event_id: int) -> Response:
    payload = request.get_json(force=True)
    status = str(payload.get("status", "")).strip()
    if status not in {"queued", "in_wash", "returned"}:
        return jsonify({"error": "Invalid status"}), 400
    row = db().execute(
        """
        SELECT w.id, c.lab_id
        FROM cage_wash_events w
        JOIN cages c ON c.id = w.cage_id
        WHERE w.id = ?
        """,
        (event_id,),
    ).fetchone()
    if not row:
        return jsonify({"error": "Not found"}), 404
    if not is_admin(g.user) and int(row["lab_id"]) != int(g.user.lab_id or -1):
        return jsonify({"error": "Forbidden"}), 403
    db().execute(
        "UPDATE cage_wash_events SET status = ?, completed_at = CASE WHEN ? = 'returned' THEN ? ELSE completed_at END WHERE id = ?",
        (status, status, now_iso(), event_id),
    )
    db().commit()
    return jsonify({"ok": True})


@app.get("/api/wash-events")
@require_auth()
def list_wash_events() -> Response:
    status = request.args.get("status", "").strip()
    clauses = []
    params: list[Any] = []
    if status:
        clauses.append("w.status = ?")
        params.append(status)
    if not is_admin(g.user):
        clauses.append("c.lab_id = ?")
        params.append(g.user.lab_id)
    where = f"WHERE {' AND '.join(clauses)}" if clauses else ""
    rows = db().execute(
        f"""
        SELECT w.id, w.status, w.requested_at, w.completed_at, c.cage_code
        FROM cage_wash_events w
        JOIN cages c ON c.id = w.cage_id
        {where}
        ORDER BY w.id DESC
        LIMIT 1000
        """,
        params,
    ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.post("/api/quarantine/intakes")
@require_auth(("Technician", "PI", "Admin"))
def create_quarantine_intake() -> Response:
    payload = request.get_json(force=True)
    quantity = int(payload.get("quantity", 0))
    if quantity <= 0:
        return jsonify({"error": "quantity must be > 0"}), 400
    lab_id = int(payload.get("labId") or g.user.lab_id or 0)
    if lab_id <= 0:
        return jsonify({"error": "labId is required"}), 400
    if not is_admin(g.user) and int(g.user.lab_id or -1) != lab_id:
        return jsonify({"error": "Forbidden"}), 403
    project_id = payload.get("projectId")
    if project_id:
        project = ensure_project_scope(int(project_id), g.user)
        if not project:
            return jsonify({"error": "Project not found"}), 404
    cage_id = payload.get("cageId")
    if cage_id:
        cage = ensure_cage_scope(int(cage_id), g.user)
        if not cage:
            return jsonify({"error": "Cage not found"}), 404
    now = now_iso()
    cur = db().execute(
        """
        INSERT INTO quarantine_intakes (
            lab_id, project_id, cage_id, vendor, strain, sex, quantity, arrival_date,
            quarantine_end_on, status, notes, created_by, created_at, updated_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'planned', ?, ?, ?, ?)
        """,
        (
            lab_id,
            project_id,
            cage_id,
            payload.get("vendor"),
            payload.get("strain"),
            payload.get("sex"),
            quantity,
            str(payload.get("arrivalDate") or date.today().isoformat()),
            payload.get("quarantineEndOn"),
            payload.get("notes"),
            g.user.user_id,
            now,
            now,
        ),
    )
    db().commit()
    audit_log(g.user.user_id, "quarantine_intake", cur.lastrowid, "create", None, payload)
    return jsonify({"id": cur.lastrowid}), 201


@app.get("/api/quarantine/intakes")
@require_auth()
def list_quarantine_intakes() -> Response:
    status = request.args.get("status", "").strip()
    clauses = []
    params: list[Any] = []
    if status:
        clauses.append("q.status = ?")
        params.append(status)
    if not is_admin(g.user):
        clauses.append("q.lab_id = ?")
        params.append(g.user.lab_id)
    where = f"WHERE {' AND '.join(clauses)}" if clauses else ""
    rows = db().execute(
        f"""
        SELECT q.id, q.status, q.arrival_date, q.quarantine_end_on, q.quantity, q.vendor, q.strain, q.sex,
               q.notes, l.name AS lab_name, p.project_code, c.cage_code
        FROM quarantine_intakes q
        JOIN labs l ON l.id = q.lab_id
        LEFT JOIN projects p ON p.id = q.project_id
        LEFT JOIN cages c ON c.id = q.cage_id
        {where}
        ORDER BY q.id DESC
        LIMIT 1000
        """,
        params,
    ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.post("/api/quarantine/intakes/<int:intake_id>/status")
@require_auth(("PI", "Admin"))
def update_quarantine_status(intake_id: int) -> Response:
    payload = request.get_json(force=True)
    status = str(payload.get("status", "")).strip()
    if status not in {"planned", "arrived", "in_quarantine", "cleared", "blocked"}:
        return jsonify({"error": "Invalid status"}), 400
    row = db().execute("SELECT * FROM quarantine_intakes WHERE id = ?", (intake_id,)).fetchone()
    if not row:
        return jsonify({"error": "Not found"}), 404
    if not is_admin(g.user) and int(row["lab_id"]) != int(g.user.lab_id or -1):
        return jsonify({"error": "Forbidden"}), 403
    now = now_iso()
    db().execute(
        """
        UPDATE quarantine_intakes
        SET status = ?, quarantine_end_on = COALESCE(?, quarantine_end_on), notes = COALESCE(?, notes),
            reviewed_by = ?, updated_at = ?
        WHERE id = ?
        """,
        (status, payload.get("quarantineEndOn"), payload.get("notes"), g.user.user_id, now, intake_id),
    )
    db().commit()
    audit_log(g.user.user_id, "quarantine_intake", intake_id, "status_update", {"status": row["status"]}, {"status": status})
    return jsonify({"ok": True})


@app.get("/api/compliance/quarantine-alerts")
@require_auth(("PI", "Admin"))
def quarantine_alerts() -> Response:
    today = date.today().isoformat()
    rows = db().execute(
        """
        SELECT q.id, q.status, q.arrival_date, q.quarantine_end_on, q.quantity, l.name AS lab_name, c.cage_code
        FROM quarantine_intakes q
        JOIN labs l ON l.id = q.lab_id
        LEFT JOIN cages c ON c.id = q.cage_id
        WHERE q.status IN ('arrived', 'in_quarantine')
          AND (q.quarantine_end_on IS NULL OR q.quarantine_end_on <= ?)
        """
        + ("" if is_admin(g.user) else " AND q.lab_id = ? ")
        + """
        ORDER BY q.quarantine_end_on ASC, q.arrival_date ASC
        LIMIT 500
        """,
        (today,) if is_admin(g.user) else (today, g.user.lab_id),
    ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.post("/api/cages/<int:cage_id>/mortality")
@require_auth(("Technician", "PI", "Admin"))
def record_mortality(cage_id: int) -> Response:
    cage = ensure_cage_scope(cage_id, g.user)
    if not cage:
        return jsonify({"error": "Not found"}), 404
    payload = request.get_json(force=True)
    animal_id = payload.get("animalId")
    male = int(payload.get("male", 0))
    female = int(payload.get("female", 0))
    necropsy_required = bool(payload.get("necropsyRequired", False))
    if not animal_id and male <= 0 and female <= 0:
        return jsonify({"error": "Provide animalId or male/female counts"}), 400
    if animal_id:
        animal = db().execute("SELECT * FROM animals WHERE id = ? AND cage_id = ?", (animal_id, cage_id)).fetchone()
        if not animal:
            return jsonify({"error": "Animal not found in cage"}), 404
        db().execute("UPDATE animals SET status = 'Dead', updated_at = ? WHERE id = ?", (now_iso(), animal_id))
    else:
        db().execute(
            """
            UPDATE cages
            SET male_count = CASE WHEN male_count - ? > 0 THEN male_count - ? ELSE 0 END,
                female_count = CASE WHEN female_count - ? > 0 THEN female_count - ? ELSE 0 END,
                updated_at = ?
            WHERE id = ?
            """,
            (male, male, female, female, now_iso(), cage_id),
        )
    cur = db().execute(
        """
        INSERT INTO mortality_records (
            animal_id, cage_id, protocol_id, count_male, count_female, cause, found_at,
            reported_by, necropsy_required, necropsy_status, notes
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            animal_id,
            cage_id,
            cage["protocol_id"],
            0 if animal_id else male,
            0 if animal_id else female,
            payload.get("cause"),
            now_iso(),
            g.user.user_id,
            1 if necropsy_required else 0,
            "pending" if necropsy_required else "not_required",
            payload.get("notes"),
        ),
    )
    db().commit()
    audit_log(g.user.user_id, "mortality", cur.lastrowid, "create", None, payload)
    return jsonify({"id": cur.lastrowid}), 201


@app.get("/api/mortality")
@require_auth()
def list_mortality() -> Response:
    necropsy_status = request.args.get("necropsyStatus", "").strip()
    clauses = []
    params: list[Any] = []
    if necropsy_status:
        clauses.append("m.necropsy_status = ?")
        params.append(necropsy_status)
    if not is_admin(g.user):
        clauses.append("c.lab_id = ?")
        params.append(g.user.lab_id)
    where = f"WHERE {' AND '.join(clauses)}" if clauses else ""
    rows = db().execute(
        f"""
        SELECT m.id, m.found_at, m.cause, m.count_male, m.count_female, m.necropsy_required, m.necropsy_status,
               c.cage_code, a.animal_code
        FROM mortality_records m
        JOIN cages c ON c.id = m.cage_id
        LEFT JOIN animals a ON a.id = m.animal_id
        {where}
        ORDER BY m.id DESC
        LIMIT 2000
        """,
        params,
    ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.post("/api/mortality/<int:record_id>/necropsy")
@require_auth(("PI", "Admin"))
def update_mortality_necropsy(record_id: int) -> Response:
    payload = request.get_json(force=True)
    status = str(payload.get("status", "")).strip()
    if status not in {"pending", "completed"}:
        return jsonify({"error": "Invalid status"}), 400
    row = db().execute(
        """
        SELECT m.id, c.lab_id
        FROM mortality_records m
        JOIN cages c ON c.id = m.cage_id
        WHERE m.id = ?
        """,
        (record_id,),
    ).fetchone()
    if not row:
        return jsonify({"error": "Not found"}), 404
    if not is_admin(g.user) and int(row["lab_id"]) != int(g.user.lab_id or -1):
        return jsonify({"error": "Forbidden"}), 403
    db().execute("UPDATE mortality_records SET necropsy_status = ? WHERE id = ?", (status, record_id))
    db().commit()
    return jsonify({"ok": True})


@app.get("/api/reports/mortality.csv")
@require_auth(("PI", "Admin"))
def report_mortality() -> Response:
    rows = db().execute(
        """
        SELECT m.id, m.found_at, m.cause, m.count_male, m.count_female, m.necropsy_status, c.cage_code, a.animal_code
        FROM mortality_records m
        JOIN cages c ON c.id = m.cage_id
        LEFT JOIN animals a ON a.id = m.animal_id
        """
        + ("" if is_admin(g.user) else " WHERE c.lab_id = ? ")
        + """
        ORDER BY m.id DESC
        LIMIT 2000
        """,
        () if is_admin(g.user) else (g.user.lab_id,),
    ).fetchall()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(
        ["id", "found_at", "cage_code", "animal_code", "cause", "count_male", "count_female", "necropsy_status"]
    )
    for r in rows:
        writer.writerow(
            [
                r["id"],
                r["found_at"],
                r["cage_code"],
                r["animal_code"],
                r["cause"],
                r["count_male"],
                r["count_female"],
                r["necropsy_status"],
            ]
        )
    return Response(output.getvalue(), mimetype="text/csv", headers={"Content-Disposition": "attachment; filename=mortality_report.csv"})


@app.get("/api/alerts/feed")
@require_auth()
def alerts_feed() -> Response:
    status = request.args.get("status", "active").strip()
    severity = request.args.get("severity", "").strip()
    upsert_active_alerts(g.user)

    clauses = []
    params: list[Any] = []
    if status:
        if status not in {"active", "acknowledged", "resolved"}:
            return jsonify({"error": "Invalid status"}), 400
        clauses.append("a.status = ?")
        params.append(status)
    if severity:
        if severity not in {"low", "medium", "high"}:
            return jsonify({"error": "Invalid severity"}), 400
        clauses.append("a.severity = ?")
        params.append(severity)
    if not is_admin(g.user):
        clauses.append("a.lab_id = ?")
        params.append(g.user.lab_id)
    where_sql = f"WHERE {' AND '.join(clauses)}" if clauses else ""

    rows = db().execute(
        f"""
        SELECT a.id, a.alert_key, a.cage_id, a.severity, a.category, a.title, a.message, a.status,
               a.first_seen_at, a.last_seen_at, a.escalation_level, a.next_notify_at, c.cage_code
        FROM alert_notifications a
        LEFT JOIN cages c ON c.id = a.cage_id
        {where_sql}
        ORDER BY
          CASE a.severity WHEN 'high' THEN 3 WHEN 'medium' THEN 2 ELSE 1 END DESC,
          a.last_seen_at DESC
        LIMIT 500
        """,
        params,
    ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.post("/api/alerts/<int:alert_id>/ack")
@require_auth(("Technician", "PI", "Admin"))
def acknowledge_alert(alert_id: int) -> Response:
    row = db().execute("SELECT id, lab_id, status FROM alert_notifications WHERE id = ?", (alert_id,)).fetchone()
    if not row:
        return jsonify({"error": "Not found"}), 404
    if not is_admin(g.user) and int(row["lab_id"] or -1) != int(g.user.lab_id or -1):
        return jsonify({"error": "Forbidden"}), 403
    if row["status"] == "resolved":
        return jsonify({"error": "Alert already resolved"}), 409
    db().execute(
        "UPDATE alert_notifications SET status = 'acknowledged', acked_by = ?, acked_at = ? WHERE id = ?",
        (g.user.user_id, now_iso(), alert_id),
    )
    db().commit()
    audit_log(g.user.user_id, "alert_notification", alert_id, "acknowledge", None, {"status": "acknowledged"})
    return jsonify({"ok": True})


@app.get("/api/notifications/channels")
@require_auth(("PI", "Admin"))
def list_notification_channels() -> Response:
    rows = db().execute(
        """
        SELECT id, user_id, lab_id, channel_type, target, min_severity, is_active, created_at
        FROM notification_channels
        """
        + ("" if is_admin(g.user) else " WHERE lab_id = ? ")
        + " ORDER BY id DESC",
        () if is_admin(g.user) else (g.user.lab_id,),
    ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.post("/api/notifications/channels")
@require_auth(("PI", "Admin"))
def create_notification_channel() -> Response:
    payload = request.get_json(force=True)
    channel_type = str(payload.get("channelType", "")).strip()
    if channel_type not in {"in_app", "webhook", "email"}:
        return jsonify({"error": "Invalid channelType"}), 400
    min_severity = str(payload.get("minSeverity", "medium")).strip()
    if min_severity not in {"low", "medium", "high"}:
        return jsonify({"error": "Invalid minSeverity"}), 400
    lab_id = int(payload.get("labId") or g.user.lab_id or 0)
    if not is_admin(g.user) and int(g.user.lab_id or -1) != lab_id:
        return jsonify({"error": "Forbidden"}), 403
    target = (payload.get("target") or "").strip()
    if channel_type in {"webhook", "email"} and not target:
        return jsonify({"error": "target is required for webhook/email channels"}), 400
    cur = db().execute(
        """
        INSERT INTO notification_channels (user_id, lab_id, channel_type, target, min_severity, is_active, created_at)
        VALUES (?, ?, ?, ?, ?, 1, ?)
        """,
        (payload.get("userId"), lab_id, channel_type, target or None, min_severity, now_iso()),
    )
    db().commit()
    return jsonify({"id": cur.lastrowid}), 201


@app.post("/api/alerts/dispatch")
@require_auth(("PI", "Admin"))
def dispatch_alert_notifications() -> Response:
    upsert_active_alerts(g.user)
    now = now_iso()
    due_rows = db().execute(
        """
        SELECT *
        FROM alert_notifications
        WHERE status = 'active'
          AND (next_notify_at IS NULL OR next_notify_at <= ?)
        """
        + ("" if is_admin(g.user) else " AND lab_id = ? ")
        + """
        ORDER BY
          CASE severity WHEN 'high' THEN 3 WHEN 'medium' THEN 2 ELSE 1 END DESC,
          id ASC
        LIMIT 200
        """,
        (now,) if is_admin(g.user) else (now, g.user.lab_id),
    ).fetchall()
    if not due_rows:
        return jsonify({"dispatched": 0, "failed": 0, "simulated": 0, "alerts": 0})

    channels = db().execute(
        """
        SELECT id, lab_id, channel_type, target, min_severity
        FROM notification_channels
        WHERE is_active = 1
        """
        + ("" if is_admin(g.user) else " AND lab_id = ? "),
        () if is_admin(g.user) else (g.user.lab_id,),
    ).fetchall()
    sent = 0
    failed = 0
    simulated = 0

    for alert in due_rows:
        for channel in channels:
            if channel["lab_id"] is not None and alert["lab_id"] != channel["lab_id"]:
                continue
            if not severity_at_least(alert["severity"], channel["min_severity"]):
                continue

            status = "simulated"
            summary = "in-app queue"
            if channel["channel_type"] == "webhook":
                body = json.dumps(
                    {
                        "id": alert["id"],
                        "severity": alert["severity"],
                        "category": alert["category"],
                        "title": alert["title"],
                        "message": alert["message"],
                        "cageId": alert["cage_id"],
                    }
                ).encode("utf-8")
                req = urlrequest.Request(str(channel["target"]), data=body, method="POST")
                req.add_header("Content-Type", "application/json")
                try:
                    with urlrequest.urlopen(req, timeout=5) as resp:
                        code = getattr(resp, "status", 200)
                    status = "sent" if 200 <= int(code) < 300 else "failed"
                    summary = f"webhook status={code}"
                except (urlerror.URLError, TimeoutError) as exc:
                    status = "failed"
                    summary = str(exc)
            elif channel["channel_type"] == "email":
                status = "simulated"
                summary = f"email queued to {channel['target']}"

            db().execute(
                """
                INSERT INTO notification_dispatch_log (alert_id, channel_id, dispatched_at, status, response_summary)
                VALUES (?, ?, ?, ?, ?)
                """,
                (alert["id"], channel["id"], now_iso(), status, summary),
            )
            if status == "sent":
                sent += 1
            elif status == "failed":
                failed += 1
            else:
                simulated += 1

        level = int(alert["escalation_level"] or 0) + 1
        delay_min = escalation_delay_minutes(str(alert["severity"]), level)
        next_notify_at = (datetime.now(UTC) + timedelta(minutes=delay_min)).isoformat()
        db().execute(
            "UPDATE alert_notifications SET escalation_level = ?, last_notified_at = ?, next_notify_at = ? WHERE id = ?",
            (level, now_iso(), next_notify_at, alert["id"]),
        )
    db().commit()
    return jsonify({"dispatched": sent, "failed": failed, "simulated": simulated, "alerts": len(due_rows)})


@app.get("/api/alerts/stream")
@require_auth()
def alerts_stream() -> Response:
    interval_s = max(1, min(int(request.args.get("interval", "10")), 60))
    once = request.args.get("once", "0") == "1"
    user = g.user

    @stream_with_context
    def event_stream():
        yield "retry: 10000\n\n"
        while True:
            upsert_active_alerts(user)
            rows = db().execute(
                """
                SELECT a.id, a.alert_key, a.cage_id, a.severity, a.category, a.title, a.message, a.status, c.cage_code
                FROM alert_notifications a
                LEFT JOIN cages c ON c.id = a.cage_id
                WHERE a.status = 'active'
                """
                + ("" if is_admin(user) else " AND a.lab_id = ? ")
                + """
                ORDER BY CASE a.severity WHEN 'high' THEN 3 WHEN 'medium' THEN 2 ELSE 1 END DESC, a.last_seen_at DESC
                LIMIT 250
                """,
                () if is_admin(user) else (user.lab_id,),
            ).fetchall()
            payload = json.dumps([dict(r) for r in rows], default=str)
            yield f"event: alerts\ndata: {payload}\n\n"
            if once:
                break
            time.sleep(interval_s)

    return Response(
        event_stream(),
        mimetype="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@app.post("/api/breeding/pairs")
@require_auth(("Technician", "PI", "Admin"))
def create_breeding_pair() -> Response:
    payload = request.get_json(force=True)
    sire_id = int(payload.get("sireId", 0))
    dam_id = int(payload.get("damId", 0))
    cage_id = int(payload.get("cageId", 0))
    if sire_id <= 0 or dam_id <= 0 or cage_id <= 0:
        return jsonify({"error": "sireId, damId, and cageId are required"}), 400
    if sire_id == dam_id:
        return jsonify({"error": "sireId and damId cannot be the same"}), 400
    cage = ensure_cage_scope(cage_id, g.user)
    if not cage:
        return jsonify({"error": "Cage not found"}), 404
    sire = ensure_animal_scope(sire_id, g.user)
    dam = ensure_animal_scope(dam_id, g.user)
    if not sire or not dam:
        return jsonify({"error": "Animal not found"}), 404
    if int(sire["cage_id"] or -1) != cage_id or int(dam["cage_id"] or -1) != cage_id:
        return jsonify({"error": "Both animals must be in the selected cage"}), 409
    cur = db().execute(
        """
        INSERT INTO breeding_pairs (sire_id, dam_id, cage_id, lab_id, status, started_on, notes, created_by, created_at, updated_at)
        VALUES (?, ?, ?, ?, 'active', ?, ?, ?, ?, ?)
        """,
        (
            sire_id,
            dam_id,
            cage_id,
            cage["lab_id"],
            str(payload.get("startedOn") or date.today().isoformat()),
            payload.get("notes"),
            g.user.user_id,
            now_iso(),
            now_iso(),
        ),
    )
    db().commit()
    audit_log(g.user.user_id, "breeding_pair", cur.lastrowid, "create", None, payload)
    return jsonify({"id": cur.lastrowid}), 201


@app.get("/api/breeding/pairs")
@require_auth()
def list_breeding_pairs() -> Response:
    status = request.args.get("status", "").strip()
    clauses = []
    params: list[Any] = []
    if status:
        clauses.append("bp.status = ?")
        params.append(status)
    if not is_admin(g.user):
        clauses.append("bp.lab_id = ?")
        params.append(g.user.lab_id)
    where_sql = f"WHERE {' AND '.join(clauses)}" if clauses else ""
    rows = db().execute(
        f"""
        SELECT bp.id, bp.status, bp.started_on, bp.ended_on, bp.notes,
               c.cage_code, sire.animal_code AS sire_code, dam.animal_code AS dam_code
        FROM breeding_pairs bp
        JOIN cages c ON c.id = bp.cage_id
        JOIN animals sire ON sire.id = bp.sire_id
        JOIN animals dam ON dam.id = bp.dam_id
        {where_sql}
        ORDER BY bp.id DESC
        LIMIT 500
        """,
        params,
    ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.post("/api/breeding/pairs/<int:pair_id>/status")
@require_auth(("Technician", "PI", "Admin"))
def update_breeding_pair_status(pair_id: int) -> Response:
    payload = request.get_json(force=True)
    status = str(payload.get("status", "")).strip()
    if status not in {"active", "paused", "retired"}:
        return jsonify({"error": "Invalid status"}), 400
    row = db().execute(
        "SELECT * FROM breeding_pairs WHERE id = ?" + ("" if is_admin(g.user) else " AND lab_id = ?"),
        (pair_id,) if is_admin(g.user) else (pair_id, g.user.lab_id),
    ).fetchone()
    if not row:
        return jsonify({"error": "Not found"}), 404
    ended_on = payload.get("endedOn") if status == "retired" else None
    db().execute(
        "UPDATE breeding_pairs SET status = ?, ended_on = COALESCE(?, ended_on), notes = COALESCE(?, notes), updated_at = ? WHERE id = ?",
        (status, ended_on, payload.get("notes"), now_iso(), pair_id),
    )
    db().commit()
    audit_log(g.user.user_id, "breeding_pair", pair_id, "status_update", {"status": row["status"]}, {"status": status})
    return jsonify({"ok": True})


@app.get("/api/breeding/pairs/<int:pair_id>/productivity")
@require_auth()
def breeding_pair_productivity(pair_id: int) -> Response:
    pair = db().execute(
        "SELECT * FROM breeding_pairs WHERE id = ?" + ("" if is_admin(g.user) else " AND lab_id = ?"),
        (pair_id,) if is_admin(g.user) else (pair_id, g.user.lab_id),
    ).fetchone()
    if not pair:
        return jsonify({"error": "Not found"}), 404
    start = pair["started_on"]
    end = pair["ended_on"] or date.today().isoformat()
    stats = db().execute(
        """
        SELECT COUNT(*) AS litter_count, COALESCE(AVG(survived_count), 0) AS avg_survived, COALESCE(SUM(survived_count), 0) AS total_survived
        FROM litters
        WHERE cage_id = ? AND birth_date BETWEEN ? AND ?
        """,
        (pair["cage_id"], start, end),
    ).fetchone()
    return jsonify(
        {
            "pairId": pair_id,
            "status": pair["status"],
            "window": {"start": start, "end": end},
            "litterCount": int(stats["litter_count"] or 0),
            "avgSurvived": round(float(stats["avg_survived"] or 0), 2),
            "totalSurvived": int(stats["total_survived"] or 0),
        }
    )


@app.post("/api/animals/<int:animal_id>/tags")
@require_auth(("Technician", "PI", "Admin"))
def create_animal_tag(animal_id: int) -> Response:
    animal = ensure_animal_scope(animal_id, g.user)
    if not animal:
        return jsonify({"error": "Not found"}), 404
    payload = request.get_json(force=True)
    tag_type = str(payload.get("tagType", "")).strip()
    tag_value = str(payload.get("tagValue", "")).strip()
    if tag_type not in {"ear_tag", "microchip", "tube", "well", "custom"}:
        return jsonify({"error": "Invalid tagType"}), 400
    if not tag_value:
        return jsonify({"error": "tagValue is required"}), 400
    try:
        cur = db().execute(
            """
            INSERT INTO animal_tags (animal_id, tag_type, tag_value, is_active, applied_on, applied_by)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (
                animal_id,
                tag_type,
                tag_value,
                1 if bool(payload.get("active", True)) else 0,
                str(payload.get("appliedOn") or date.today().isoformat()),
                g.user.user_id,
            ),
        )
        db().commit()
    except storage.IntegrityError:
        return jsonify({"error": "Tag already in use"}), 409
    audit_log(g.user.user_id, "animal_tag", cur.lastrowid, "create", None, payload)
    return jsonify({"id": cur.lastrowid}), 201


@app.get("/api/animals/<int:animal_id>/tags")
@require_auth()
def list_animal_tags(animal_id: int) -> Response:
    animal = ensure_animal_scope(animal_id, g.user)
    if not animal:
        return jsonify({"error": "Not found"}), 404
    rows = db().execute(
        "SELECT id, tag_type, tag_value, is_active, applied_on, applied_by FROM animal_tags WHERE animal_id = ? ORDER BY id DESC",
        (animal_id,),
    ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.post("/api/samples")
@require_auth(("Technician", "PI", "Admin"))
def create_sample_record() -> Response:
    payload = request.get_json(force=True)
    animal_id = int(payload.get("animalId", 0))
    if animal_id <= 0:
        return jsonify({"error": "animalId is required"}), 400
    animal = ensure_animal_scope(animal_id, g.user)
    if not animal:
        return jsonify({"error": "Animal not found"}), 404
    sample_type = str(payload.get("sampleType", "")).strip()
    if not sample_type:
        return jsonify({"error": "sampleType is required"}), 400
    sample_code = str(payload.get("sampleCode") or f"SMP-{secrets.token_hex(5)}").strip()
    status = str(payload.get("status", "collected")).strip()
    if status not in {"collected", "shipped", "received", "resulted", "rejected"}:
        return jsonify({"error": "Invalid status"}), 400
    try:
        cur = db().execute(
            """
            INSERT INTO sample_records (animal_id, cage_id, sample_type, sample_code, provider, status, tracking_number, collected_on, collected_by, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                animal_id,
                animal["cage_id"],
                sample_type,
                sample_code,
                payload.get("provider"),
                status,
                payload.get("trackingNumber"),
                str(payload.get("collectedOn") or date.today().isoformat()),
                g.user.user_id,
                payload.get("notes"),
            ),
        )
        sample_id = cur.lastrowid
        db().execute(
            """
            INSERT INTO sample_events (sample_id, event_type, event_time, actor_user_id, details_json)
            VALUES (?, ?, ?, ?, ?)
            """,
            (sample_id, status, now_iso(), g.user.user_id, json.dumps(payload.get("eventDetails", {}))),
        )
        db().commit()
    except storage.IntegrityError:
        return jsonify({"error": "sampleCode already exists"}), 409
    audit_log(g.user.user_id, "sample_record", sample_id, "create", None, payload)
    return jsonify({"id": sample_id, "sampleCode": sample_code}), 201


@app.get("/api/samples")
@require_auth()
def list_sample_records() -> Response:
    status = request.args.get("status", "").strip()
    provider = request.args.get("provider", "").strip()
    clauses = []
    params: list[Any] = []
    if status:
        clauses.append("s.status = ?")
        params.append(status)
    if provider:
        clauses.append("COALESCE(s.provider, '') = ?")
        params.append(provider)
    if not is_admin(g.user):
        clauses.append("c.lab_id = ?")
        params.append(g.user.lab_id)
    where_sql = f"WHERE {' AND '.join(clauses)}" if clauses else ""
    rows = db().execute(
        f"""
        SELECT s.id, s.sample_code, s.sample_type, s.provider, s.status, s.collected_on, s.tracking_number,
               s.notes, a.id AS animal_id, a.animal_code, c.id AS cage_id, c.cage_code
        FROM sample_records s
        JOIN animals a ON a.id = s.animal_id
        LEFT JOIN cages c ON c.id = s.cage_id
        {where_sql}
        ORDER BY s.id DESC
        LIMIT 1000
        """,
        params,
    ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.post("/api/samples/<int:sample_id>/status")
@require_auth(("Technician", "PI", "Admin"))
def update_sample_status(sample_id: int) -> Response:
    payload = request.get_json(force=True)
    status = str(payload.get("status", "")).strip()
    if status not in {"collected", "shipped", "received", "resulted", "rejected"}:
        return jsonify({"error": "Invalid status"}), 400
    row = db().execute(
        """
        SELECT s.id, s.status, c.lab_id
        FROM sample_records s
        LEFT JOIN cages c ON c.id = s.cage_id
        WHERE s.id = ?
        """,
        (sample_id,),
    ).fetchone()
    if not row:
        return jsonify({"error": "Not found"}), 404
    if not is_admin(g.user) and int(row["lab_id"] or -1) != int(g.user.lab_id or -1):
        return jsonify({"error": "Forbidden"}), 403
    db().execute(
        "UPDATE sample_records SET status = ?, tracking_number = COALESCE(?, tracking_number), notes = COALESCE(?, notes) WHERE id = ?",
        (status, payload.get("trackingNumber"), payload.get("notes"), sample_id),
    )
    db().execute(
        "INSERT INTO sample_events (sample_id, event_type, event_time, actor_user_id, details_json) VALUES (?, ?, ?, ?, ?)",
        (sample_id, status, now_iso(), g.user.user_id, json.dumps(payload.get("eventDetails", {}))),
    )
    db().commit()
    audit_log(g.user.user_id, "sample_record", sample_id, "status_update", {"status": row["status"]}, {"status": status})
    return jsonify({"ok": True})


@app.get("/api/samples/<int:sample_id>/events")
@require_auth()
def sample_event_history(sample_id: int) -> Response:
    sample = db().execute(
        """
        SELECT s.id, c.lab_id
        FROM sample_records s
        LEFT JOIN cages c ON c.id = s.cage_id
        WHERE s.id = ?
        """,
        (sample_id,),
    ).fetchone()
    if not sample:
        return jsonify({"error": "Not found"}), 404
    if not is_admin(g.user) and int(sample["lab_id"] or -1) != int(g.user.lab_id or -1):
        return jsonify({"error": "Forbidden"}), 403
    rows = db().execute(
        "SELECT id, event_type, event_time, actor_user_id, details_json FROM sample_events WHERE sample_id = ? ORDER BY id ASC",
        (sample_id,),
    ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.get("/api/animals/<int:animal_id>/genotypes")
@require_auth()
def animal_genotype_history(animal_id: int) -> Response:
    animal = ensure_animal_scope(animal_id, g.user)
    if not animal:
        return jsonify({"error": "Not found"}), 404
    rows = db().execute(
        """
        SELECT id, result, source, created_at
        FROM genotype_results
        WHERE animal_id = ?
        ORDER BY id DESC
        LIMIT 100
        """,
        (animal_id,),
    ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.post("/api/genotyping/orders")
@require_auth(("Technician", "PI", "Admin"))
def create_genotyping_order() -> Response:
    payload = request.get_json(force=True)
    provider = str(payload.get("provider", "")).strip()
    if not provider:
        return jsonify({"error": "provider is required"}), 400
    project_id = payload.get("projectId")
    if project_id:
        project = ensure_project_scope(int(project_id), g.user)
        if not project:
            return jsonify({"error": "Project not found"}), 404
        lab_id = int(project["lab_id"])
    else:
        lab_id = int(payload.get("labId") or g.user.lab_id or 0)
    if lab_id <= 0:
        return jsonify({"error": "labId is required"}), 400
    if not is_admin(g.user) and int(g.user.lab_id or -1) != lab_id:
        return jsonify({"error": "Forbidden"}), 403

    order_ref = str(payload.get("orderRef") or f"GENO-{datetime.now(UTC).strftime('%Y%m%d')}-{secrets.token_hex(3)}").strip()
    sample_ids = [int(x) for x in payload.get("sampleIds", [])]
    marker_panel = payload.get("markerPanel")
    now = now_iso()
    try:
        cur = db().execute(
            """
            INSERT INTO genotyping_orders (lab_id, project_id, provider, order_ref, status, requested_by, created_at, updated_at, payload_json)
            VALUES (?, ?, ?, ?, 'draft', ?, ?, ?, ?)
            """,
            (lab_id, project_id, provider, order_ref, g.user.user_id, now, now, json.dumps(payload, default=str)),
        )
    except storage.IntegrityError:
        return jsonify({"error": "orderRef already exists"}), 409
    order_id = cur.lastrowid
    inserted_items = 0
    for sid in sample_ids:
        sample = db().execute(
            """
            SELECT s.id, s.animal_id, c.lab_id
            FROM sample_records s
            LEFT JOIN cages c ON c.id = s.cage_id
            WHERE s.id = ?
            """,
            (sid,),
        ).fetchone()
        if not sample:
            continue
        if not is_admin(g.user) and int(sample["lab_id"] or -1) != int(g.user.lab_id or -1):
            continue
        db().execute(
            """
            INSERT INTO genotyping_order_items (order_id, sample_id, animal_id, marker_panel)
            VALUES (?, ?, ?, ?)
            """,
            (order_id, sid, sample["animal_id"], marker_panel),
        )
        inserted_items += 1
    db().commit()
    audit_log(g.user.user_id, "genotyping_order", order_id, "create", None, {"sampleIds": sample_ids, "items": inserted_items})
    return jsonify({"id": order_id, "orderRef": order_ref, "items": inserted_items}), 201


@app.get("/api/genotyping/orders")
@require_auth()
def list_genotyping_orders() -> Response:
    status = request.args.get("status", "").strip()
    clauses = []
    params: list[Any] = []
    if status:
        clauses.append("o.status = ?")
        params.append(status)
    if not is_admin(g.user):
        clauses.append("o.lab_id = ?")
        params.append(g.user.lab_id)
    where_sql = f"WHERE {' AND '.join(clauses)}" if clauses else ""
    rows = db().execute(
        f"""
        SELECT o.id, o.provider, o.order_ref, o.status, o.created_at, o.updated_at,
               COUNT(i.id) AS item_count,
               SUM(CASE WHEN i.result IS NOT NULL THEN 1 ELSE 0 END) AS resulted_count
        FROM genotyping_orders o
        LEFT JOIN genotyping_order_items i ON i.order_id = o.id
        {where_sql}
        GROUP BY o.id
        ORDER BY o.id DESC
        LIMIT 500
        """,
        params,
    ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.get("/api/genotyping/orders/<int:order_id>")
@require_auth()
def genotyping_order_detail(order_id: int) -> Response:
    order = _order_scope_row(order_id, g.user)
    if not order:
        return jsonify({"error": "Not found"}), 404
    items = _order_items_with_sample_context(order_id)
    reconciliation = _serialize_genotyping_reconciliation(order, items)
    return jsonify({"order": dict(order), "items": items, "reconciliation": reconciliation})


@app.get("/api/genotyping/orders/<int:order_id>/reconciliation")
@require_auth()
def genotyping_order_reconciliation(order_id: int) -> Response:
    order = _order_scope_row(order_id, g.user)
    if not order:
        return jsonify({"error": "Not found"}), 404
    items = _order_items_with_sample_context(order_id)
    return jsonify(_serialize_genotyping_reconciliation(order, items))


@app.get("/api/genotyping/orders/<int:order_id>/provider-template.csv")
@require_auth()
def genotyping_order_provider_template(order_id: int) -> Response:
    order = _order_scope_row(order_id, g.user)
    if not order:
        return jsonify({"error": "Not found"}), 404
    items = _order_items_with_sample_context(order_id)
    preset = _provider_preset_by_name(order["provider"])
    columns = list(preset.get("exportColumns") or ["order_ref", "sample_code", "animal_code", "marker_panel", "result"])
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=columns)
    writer.writeheader()
    for item in items:
        source_row = {
            "order_ref": order["order_ref"],
            "provider": order["provider"],
            "sample_code": item.get("sample_code") or "",
            "tube_id": item.get("sample_code") or "",
            "animal_code": item.get("animal_code") or "",
            "animal_id": item.get("animal_code") or "",
            "marker_panel": item.get("marker_panel") or "",
            "panel_name": item.get("marker_panel") or "",
            "target_assay": item.get("marker_panel") or "",
            "well_position": "",
            "result": item.get("result") or "",
        }
        writer.writerow({column: source_row.get(column, "") for column in columns})
    filename = f"{order['order_ref']}_provider_template.csv"
    return Response(output.getvalue(), mimetype="text/csv", headers={"Content-Disposition": f"attachment; filename={filename}"})


@app.post("/api/genotyping/orders/<int:order_id>/submit")
@require_auth(("Technician", "PI", "Admin"))
def submit_genotyping_order(order_id: int) -> Response:
    row = db().execute(
        "SELECT * FROM genotyping_orders WHERE id = ?" + ("" if is_admin(g.user) else " AND lab_id = ?"),
        (order_id,) if is_admin(g.user) else (order_id, g.user.lab_id),
    ).fetchone()
    if not row:
        return jsonify({"error": "Not found"}), 404
    if row["status"] not in {"draft", "failed"}:
        return jsonify({"error": "Order not submittable in current state"}), 409
    db().execute("UPDATE genotyping_orders SET status = 'submitted', updated_at = ? WHERE id = ?", (now_iso(), order_id))
    db().commit()
    audit_log(g.user.user_id, "genotyping_order", order_id, "submit", {"status": row["status"]}, {"status": "submitted"})
    return jsonify({"ok": True})


@app.post("/api/genotyping/orders/<int:order_id>/import-results")
@require_auth(("PI", "Admin"))
def import_genotyping_order_results(order_id: int) -> Response:
    order = _order_scope_row(order_id, g.user)
    if not order:
        return jsonify({"error": "Not found"}), 404
    if "file" not in request.files:
        return jsonify({"error": "Upload a CSV file"}), 400
    content = request.files["file"].read().decode("utf-8")
    reader = csv.DictReader(io.StringIO(content))
    preset = _provider_preset_by_name(order["provider"])
    aliases = preset.get("importAliases") or {}
    results = []
    for row in reader:
        result = _csv_pick(row, list(aliases.get("result") or ["result", "genotype_result"]))
        sample_code = _csv_pick(row, list(aliases.get("sampleCode") or ["sample_code"]))
        animal_code = _csv_pick(row, list(aliases.get("animalCode") or ["animal_code"]))
        if not result or (not sample_code and not animal_code):
            continue
        results.append(
            {
                "sampleCode": sample_code or None,
                "animalCode": animal_code or None,
                "result": result,
                "markerPanel": _csv_pick(row, list(aliases.get("markerPanel") or ["marker_panel"])) or None,
            }
        )
    status = str(request.form.get("status") or "received").strip()
    if status not in {"received", "closed", "failed"}:
        return jsonify({"error": "Invalid status"}), 400
    updated = _apply_order_results(
        order,
        results,
        order_status=status,
        genotype_source=f"csv_import:{order['order_ref']}",
        actor_user_id=g.user.user_id,
    )
    db().commit()
    audit_log(g.user.user_id, "genotyping_order", order_id, "import_results", None, {"status": status, "updatedAnimals": updated})
    return jsonify({"ok": True, "updatedAnimals": updated})


@app.post("/api/genotyping/orders/callback")
def genotyping_order_callback() -> Response:
    expected = os.getenv("MURISPHERE_PROVIDER_CALLBACK_TOKEN", "dev-callback-token")
    token = request.headers.get("X-Provider-Token", "").strip()
    if expected and token != expected:
        return jsonify({"error": "Forbidden"}), 403
    payload = request.get_json(force=True)
    order_ref = str(payload.get("orderRef", "")).strip()
    if not order_ref:
        return jsonify({"error": "orderRef is required"}), 400
    order = db().execute("SELECT * FROM genotyping_orders WHERE order_ref = ?", (order_ref,)).fetchone()
    if not order:
        return jsonify({"error": "Order not found"}), 404
    status = str(payload.get("status", "received")).strip()
    if status not in {"submitted", "received", "closed", "failed"}:
        return jsonify({"error": "Invalid status"}), 400
    updated = _apply_order_results(
        order,
        [dict(item) for item in payload.get("results", [])],
        order_status=status,
        genotype_source=f"callback:{order_ref}",
        actor_user_id=None,
    )
    db().commit()
    audit_log(None, "genotyping_order", order["id"], "callback", None, {"status": status, "updatedAnimals": updated})
    return jsonify({"ok": True, "updatedAnimals": updated})


@app.post("/api/recommendations/generate")
@require_auth(("PI", "Admin"))
def generate_recommendations() -> Response:
    generated = 0
    now = now_iso()
    low_density = db().execute(
        """
        SELECT c.id, c.lab_id, c.strain, c.genotype_summary, (c.male_count + c.female_count) AS n
        FROM cages c
        WHERE (c.male_count + c.female_count) <= 2
        """
        + ("" if is_admin(g.user) else " AND c.lab_id = ? ")
        + """
        ORDER BY n ASC, c.id ASC
        LIMIT 200
        """,
        () if is_admin(g.user) else (g.user.lab_id,),
    ).fetchall()
    for c in low_density:
        exists = db().execute(
            "SELECT id FROM workflow_recommendations WHERE rec_type = 'consolidate_cage' AND cage_id = ? AND status IN ('open', 'accepted', 'adjusted')",
            (c["id"],),
        ).fetchone()
        if exists:
            continue
        db().execute(
            """
            INSERT INTO workflow_recommendations (rec_type, lab_id, cage_id, status, rationale, payload_json, created_by, created_at, updated_at)
            VALUES ('consolidate_cage', ?, ?, 'open', ?, ?, ?, ?, ?)
            """,
            (
                c["lab_id"],
                c["id"],
                "Low-density cage; consider consolidation.",
                json.dumps({"totalAnimals": c["n"], "strain": c["strain"], "genotypeSummary": c["genotype_summary"]}),
                g.user.user_id,
                now,
                now,
            ),
        )
        generated += 1
    db().commit()
    return jsonify({"generated": generated})


@app.get("/api/recommendations")
@require_auth()
def list_recommendations() -> Response:
    status = request.args.get("status", "").strip()
    rec_type = request.args.get("type", "").strip()
    clauses = []
    params: list[Any] = []
    if status:
        clauses.append("r.status = ?")
        params.append(status)
    if rec_type:
        clauses.append("r.rec_type = ?")
        params.append(rec_type)
    if not is_admin(g.user):
        clauses.append("r.lab_id = ?")
        params.append(g.user.lab_id)
    where_sql = f"WHERE {' AND '.join(clauses)}" if clauses else ""
    rows = db().execute(
        f"""
        SELECT r.id, r.rec_type, r.status, r.rationale, r.created_at, r.acted_at, c.cage_code
        FROM workflow_recommendations r
        LEFT JOIN cages c ON c.id = r.cage_id
        {where_sql}
        ORDER BY r.id DESC
        LIMIT 500
        """,
        params,
    ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.post("/api/recommendations/<int:recommendation_id>/decision")
@require_auth(("PI", "Admin"))
def decide_recommendation(recommendation_id: int) -> Response:
    payload = request.get_json(force=True)
    decision = str(payload.get("decision", "")).strip()
    if decision not in {"accepted", "adjusted", "ignored", "completed"}:
        return jsonify({"error": "Invalid decision"}), 400
    row = ensure_recommendation_scope(recommendation_id, g.user)
    if not row:
        return jsonify({"error": "Not found"}), 404
    merged_payload = {}
    if row["payload_json"]:
        try:
            merged_payload = json.loads(row["payload_json"])
        except json.JSONDecodeError:
            merged_payload = {}
    if decision == "adjusted":
        merged_payload["adjustment"] = payload.get("adjustment", {})
    if payload.get("note"):
        merged_payload["decisionNote"] = payload.get("note")
    db().execute(
        "UPDATE workflow_recommendations SET status = ?, payload_json = ?, acted_by = ?, acted_at = ?, updated_at = ? WHERE id = ?",
        (decision, json.dumps(merged_payload), g.user.user_id, now_iso(), now_iso(), recommendation_id),
    )
    db().commit()
    audit_log(g.user.user_id, "recommendation", recommendation_id, "decision", {"status": row["status"]}, {"status": decision})
    return jsonify({"ok": True})


@app.get("/api/recommendations/outcomes")
@require_auth()
def recommendation_outcomes() -> Response:
    rows = db().execute(
        """
        SELECT rec_type, status, COUNT(*) AS n
        FROM workflow_recommendations
        """
        + ("" if is_admin(g.user) else " WHERE lab_id = ? ")
        + """
        GROUP BY rec_type, status
        ORDER BY rec_type, status
        """,
        () if is_admin(g.user) else (g.user.lab_id,),
    ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.post("/api/planner/scenarios")
@require_auth(("PI", "Admin"))
def create_planner_scenario() -> Response:
    payload = request.get_json(force=True)
    name = str(payload.get("name", "")).strip()
    if not name:
        return jsonify({"error": "name is required"}), 400
    target_animals = int(payload.get("targetAnimals", 0))
    if target_animals <= 0:
        return jsonify({"error": "targetAnimals must be > 0"}), 400
    lab_id = int(payload.get("labId") or g.user.lab_id or 0)
    if lab_id <= 0:
        return jsonify({"error": "labId is required"}), 400
    if not is_admin(g.user) and int(g.user.lab_id or -1) != lab_id:
        return jsonify({"error": "Forbidden"}), 403
    cur = db().execute(
        """
        INSERT INTO planner_scenarios (lab_id, name, needed_by, target_animals, max_new_cages, assumptions_json, status, created_by, created_at, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, 'draft', ?, ?, ?)
        """,
        (
            lab_id,
            name,
            payload.get("neededBy"),
            target_animals,
            payload.get("maxNewCages"),
            json.dumps(payload.get("assumptions", {})),
            g.user.user_id,
            now_iso(),
            now_iso(),
        ),
    )
    db().commit()
    return jsonify({"id": cur.lastrowid}), 201


@app.get("/api/planner/scenarios")
@require_auth()
def list_planner_scenarios() -> Response:
    rows = db().execute(
        """
        SELECT ps.id, ps.lab_id, ps.name, ps.needed_by, ps.target_animals, ps.max_new_cages, ps.status, ps.created_at, ps.updated_at,
               l.name AS lab_name
        FROM planner_scenarios ps
        JOIN labs l ON l.id = ps.lab_id
        """
        + ("" if is_admin(g.user) else " WHERE ps.lab_id = ? ")
        + """
        ORDER BY ps.id DESC
        LIMIT 200
        """,
        () if is_admin(g.user) else (g.user.lab_id,),
    ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.get("/api/planner/scenarios/<int:scenario_id>")
@require_auth()
def planner_scenario_detail(scenario_id: int) -> Response:
    scenario = ensure_planner_scenario_scope(scenario_id, g.user)
    if not scenario:
        return jsonify({"error": "Not found"}), 404
    detail = db().execute(
        """
        SELECT ps.id, ps.lab_id, ps.name, ps.needed_by, ps.target_animals, ps.max_new_cages, ps.status, ps.assumptions_json, ps.created_at, ps.updated_at,
               l.name AS lab_name
        FROM planner_scenarios ps
        JOIN labs l ON l.id = ps.lab_id
        WHERE ps.id = ?
        """,
        (scenario_id,),
    ).fetchone()
    projects = db().execute(
        """
        SELECT p.id, p.project_code, p.title, p.status, sp.animals_needed, sp.priority
        FROM planner_scenario_projects sp
        JOIN projects p ON p.id = sp.project_id
        WHERE sp.scenario_id = ?
        ORDER BY sp.priority ASC, p.project_code ASC
        """,
        (scenario_id,),
    ).fetchall()
    return jsonify({"scenario": dict(detail), "projects": [dict(r) for r in projects]})


@app.post("/api/planner/scenarios/<int:scenario_id>/projects")
@require_auth(("PI", "Admin"))
def add_planner_scenario_projects(scenario_id: int) -> Response:
    scenario = ensure_planner_scenario_scope(scenario_id, g.user)
    if not scenario:
        return jsonify({"error": "Not found"}), 404
    payload = request.get_json(force=True)
    projects = payload.get("projects", [])
    inserted = 0
    for p in projects:
        project_id = int(p.get("projectId", 0))
        if project_id <= 0:
            continue
        project = ensure_project_scope(project_id, g.user)
        if not project:
            continue
        try:
            db().execute(
                """
                INSERT INTO planner_scenario_projects (scenario_id, project_id, animals_needed, priority)
                VALUES (?, ?, ?, ?)
                """,
                (scenario_id, project_id, int(p.get("animalsNeeded", 0)), int(p.get("priority", 3))),
            )
            inserted += 1
        except storage.IntegrityError:
            db().execute(
                "UPDATE planner_scenario_projects SET animals_needed = ?, priority = ? WHERE scenario_id = ? AND project_id = ?",
                (int(p.get("animalsNeeded", 0)), int(p.get("priority", 3)), scenario_id, project_id),
            )
    db().execute("UPDATE planner_scenarios SET updated_at = ? WHERE id = ?", (now_iso(), scenario_id))
    db().commit()
    return jsonify({"upserted": inserted})


@app.post("/api/planner/scenarios/<int:scenario_id>/evaluate")
@require_auth(("PI", "Admin"))
def evaluate_planner_scenario(scenario_id: int) -> Response:
    scenario = ensure_planner_scenario_scope(scenario_id, g.user)
    if not scenario:
        return jsonify({"error": "Not found"}), 404
    proj_rows = db().execute(
        "SELECT project_id, animals_needed, priority FROM planner_scenario_projects WHERE scenario_id = ? ORDER BY priority ASC, project_id ASC",
        (scenario_id,),
    ).fetchall()
    if proj_rows:
        target = sum(int(r["animals_needed"] or 0) for r in proj_rows)
        current = 0
        for r in proj_rows:
            got = db().execute(
                """
                SELECT COUNT(*) AS n
                FROM animals a
                JOIN cages c ON c.id = a.cage_id
                JOIN project_cages pc ON pc.cage_id = c.id
                WHERE a.status = 'Active' AND pc.project_id = ?
                """,
                (r["project_id"],),
            ).fetchone()
            current += int(got["n"] or 0)
    else:
        target = int(scenario["target_animals"] or 0)
        got = db().execute(
            """
            SELECT COUNT(*) AS n
            FROM animals a
            JOIN cages c ON c.id = a.cage_id
            WHERE a.status = 'Active' AND c.lab_id = ?
            """,
            (scenario["lab_id"],),
        ).fetchone()
        current = int(got["n"] or 0)
    deficit = max(target - current, 0)
    estimated_litters = (deficit + 3) // 4
    estimated_cages = (deficit + 4) // 5
    max_new_cages = int(scenario["max_new_cages"] or 0)
    if max_new_cages > 0 and estimated_cages > int(max_new_cages * 1.25):
        risk = "high"
    elif max_new_cages > 0 and estimated_cages > max_new_cages:
        risk = "medium"
    else:
        risk = "low"
    recommendation = {
        "targetAnimals": target,
        "currentActiveAnimals": current,
        "projectedDeficit": deficit,
        "estimatedLitters": estimated_litters,
        "estimatedCages": estimated_cages,
        "riskLevel": risk,
    }
    cur = db().execute(
        """
        INSERT INTO planner_plans (scenario_id, estimated_litters, estimated_cages, projected_deficit, risk_level, recommendation_json, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        (scenario_id, estimated_litters, estimated_cages, deficit, risk, json.dumps(recommendation), now_iso()),
    )
    if deficit > 0:
        exists = db().execute(
            "SELECT id FROM workflow_recommendations WHERE rec_type = 'planner_supply_gap' AND lab_id = ? AND status IN ('open', 'accepted', 'adjusted')",
            (scenario["lab_id"],),
        ).fetchone()
        if not exists:
            db().execute(
                """
                INSERT INTO workflow_recommendations (rec_type, lab_id, project_id, status, rationale, payload_json, created_by, created_at, updated_at)
                VALUES ('planner_supply_gap', ?, NULL, 'open', ?, ?, ?, ?, ?)
                """,
                (
                    scenario["lab_id"],
                    "Planner found projected supply gap against target demand.",
                    json.dumps(recommendation),
                    g.user.user_id,
                    now_iso(),
                    now_iso(),
                ),
            )
    db().execute("UPDATE planner_scenarios SET updated_at = ? WHERE id = ?", (now_iso(), scenario_id))
    db().commit()
    return jsonify({"planId": cur.lastrowid, **recommendation})


@app.get("/api/planner/scenarios/<int:scenario_id>/plans")
@require_auth()
def list_planner_scenario_plans(scenario_id: int) -> Response:
    scenario = ensure_planner_scenario_scope(scenario_id, g.user)
    if not scenario:
        return jsonify({"error": "Not found"}), 404
    rows = db().execute(
        """
        SELECT id, estimated_litters, estimated_cages, projected_deficit, risk_level, recommendation_json, created_at
        FROM planner_plans
        WHERE scenario_id = ?
        ORDER BY id DESC
        LIMIT 100
        """,
        (scenario_id,),
    ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.get("/api/audit")
@require_auth(("Admin",))
def audit_list() -> Response:
    rows = db().execute(
        """
        SELECT a.id, u.full_name AS actor, a.entity_type, a.entity_id, a.action, a.created_at
        FROM audit_logs a
        LEFT JOIN users u ON a.actor_user_id = u.id
        ORDER BY a.id DESC
        LIMIT 250
        """
    ).fetchall()
    return jsonify([dict(r) for r in rows])


def serve(host: str | None = None, port: int | None = None) -> None:
    init_db()
    app.run(
        host=host or DEFAULT_BIND_HOST,
        port=port or DEFAULT_BIND_PORT,
        debug=os.getenv("FLASK_DEBUG", "0") == "1",
    )


if __name__ == "__main__":
    serve()
